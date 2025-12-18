import argparse
import re
import sys
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Ajusta estos valores si tu archivo usa otros nombres.
CODE_COLUMN_NAME = "CRM"
SHEET_NAME = "Export"
OUTPUT_SHEET_NAME = "Plantilla alta de clases"
OUTPUT_FILENAME = "PlantillaClases.xlsx"
OUTPUT_DIR = Path("salidas")
EXPECTED_HEADERS: Set[str] = {
    "CRM",
    "Institucion",
    "Nivel Educativo",
    "Grado",
    "Asignatura Producto",
    "Producto",
    "Plataforma",
    "Razon Estado",
}


def _normalize_key(valor: str) -> str:
    """Normaliza texto: sin tildes, mayusculas y sin espacios extremos."""
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFD", valor)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return texto.strip().upper()


def _detectar_fila_encabezado(
    df_raw: pd.DataFrame, columna_codigo: str
) -> Optional[int]:
    """
    Identifica la fila que contiene la mayoría de encabezados conocidos.
    Devuelve el índice de fila o None si no se detecta.
    """
    candidatos = {columna_codigo.lower()} | {col.lower() for col in EXPECTED_HEADERS}
    mejor_fila, mejores_aciertos = None, 0

    for idx, row in df_raw.head(30).iterrows():
        valores = {str(val).strip().lower() for val in row if pd.notna(val)}
        aciertos = len(valores & candidatos)
        if aciertos > mejores_aciertos:
            mejor_fila, mejores_aciertos = idx, aciertos

    return mejor_fila if mejores_aciertos >= 3 else None


def cargar_excel(
    ruta_archivo: Path, hoja: str = SHEET_NAME, columna_codigo: str = CODE_COLUMN_NAME
) -> pd.DataFrame:
    """Carga la hoja solicitada; si es necesario, detecta la fila de encabezados."""
    if not ruta_archivo.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {ruta_archivo}")
    try:
        df = pd.read_excel(
            ruta_archivo,
            sheet_name=hoja,
            dtype=str,
            engine="openpyxl",
        )
    except ValueError as exc:
        raise ValueError(
            f"No se encontró la hoja '{hoja}' en el archivo. Error: {exc}"
        ) from exc
    except Exception as exc:  # pragma: no cover - defensa general
        raise RuntimeError(f"No se pudo leer el Excel: {exc}") from exc

    encabezados_ok = (
        columna_codigo in df.columns or EXPECTED_HEADERS.issubset(set(df.columns))
    )
    if encabezados_ok:
        return df

    try:
        df_raw = pd.read_excel(
            ruta_archivo,
            sheet_name=hoja,
            dtype=str,
            header=None,
            engine="openpyxl",
        )
    except Exception as exc:  # pragma: no cover - defensa general
        raise RuntimeError(
            f"No se pudo leer el Excel para detectar encabezados: {exc}"
        ) from exc

    fila_encabezado = _detectar_fila_encabezado(df_raw, columna_codigo)
    if fila_encabezado is None:
        return df

    df = pd.read_excel(
        ruta_archivo,
        sheet_name=hoja,
        dtype=str,
        engine="openpyxl",
        header=fila_encabezado,
    )
    print(f"Encabezados detectados en la fila {fila_encabezado + 1}.", file=sys.stderr)
    return df


def filtrar_codigo(df: pd.DataFrame, codigo: str, columna_codigo: str) -> pd.DataFrame:
    """Filtra las filas por el código exacto."""
    if columna_codigo not in df.columns:
        columnas = ", ".join(df.columns)
        raise KeyError(
            f"La columna '{columna_codigo}' no existe. Columnas disponibles: {columnas}"
        )
    codigo = codigo.strip()
    filtro = df[columna_codigo].astype(str).str.strip() == codigo
    return df.loc[filtro].copy()


def mostrar_resumen(df: pd.DataFrame) -> None:
    """Imprime la tabla de resumen solicitada si hay columnas disponibles."""
    columnas_deseadas = [
        "Institucion",
        "Nivel Educativo",
        "Grado",
        "Asignatura Producto",
        "Producto",
        "Plataforma",
        "Razon Estado",
    ]
    columnas_presentes = [col for col in columnas_deseadas if col in df.columns]
    if not columnas_presentes:
        print(
            "No se pueden mostrar columnas solicitadas; ninguna está presente en el DataFrame."
        )
        return
    resumen = df[columnas_presentes]
    print("\nResumen filtrado por código:\n")
    print(resumen.to_string(index=False))


def _mapear_nivel(valor: str) -> str:
    normalizado = _normalize_key(valor)
    mapa = {
        "EDUCACION PRIMARIA": "Primaria",
        "EDUCACION SECUNDARIA": "Secundaria",
    }
    return mapa.get(normalizado, "")


def _mapear_grado(valor: str) -> Tuple[str, int]:
    """Devuelve (texto_grado, numero) o ('', 0) si no coincide."""
    limpio = valor.replace("°", "º").replace("�", "º")
    normalizado = _normalize_key(limpio)
    match = re.match(r"^(\d+)\s*º?\s*(PRIMARIA|SECUNDARIA)$", normalizado)
    if not match:
        return ("", 0)

    numero = int(match.group(1))
    nivel = match.group(2)

    grados_primaria: Dict[int, str] = {
        1: "Primer grado de primaria",
        2: "Segundo grado de primaria",
        3: "Tercer grado de primaria",
        4: "Cuarto grado de primaria",
        5: "Quinto grado de primaria",
        6: "Sexto grado de primaria",
    }
    grados_secundaria: Dict[int, str] = {
        1: "Primer año de secundaria",
        2: "Segundo año de secundaria",
        3: "Tercer año de secundaria",
        4: "Cuarto año de secundaria",
        5: "Quinto año de secundaria",
    }

    if nivel == "PRIMARIA":
        return (grados_primaria.get(numero, ""), numero if numero <= 6 else 0)
    return (grados_secundaria.get(numero, ""), numero if numero <= 5 else 0)


def _mapear_materia(valor: str) -> Tuple[str, str]:
    """Devuelve (materia_legible, sufijo) o ('', '') si no coincide."""
    normalizado = _normalize_key(valor)
    mapa: Dict[str, Tuple[str, str]] = {
        "COMUNICACION": ("Comunicación", "CO"),
        "MATEMATICAS": ("Matemática", "MA"),
        "MATEMATICAS": ("Informática", "IN"),
    }
    materia = mapa.get(normalizado)
    return materia if materia else ("", "")


def transformar(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica filtros y transforma los datos para la exportación."""
    columnas_necesarias = {"Plataforma", "Asignatura Producto", "Nivel Educativo", "Grado"}
    faltantes = [col for col in columnas_necesarias if col not in df.columns]
    if faltantes:
        raise KeyError(f"Faltan columnas necesarias para transformar: {', '.join(faltantes)}")

    trabajo = df.copy().fillna("")
    trabajo = trabajo.assign(
        Plataforma=trabajo["Plataforma"].astype(str).str.strip(),
        AsignaturaProducto=trabajo["Asignatura Producto"].astype(str).str.strip(),
        NivelEducativo=trabajo["Nivel Educativo"].astype(str).str.strip(),
        GradoVal=trabajo["Grado"].astype(str).str.strip(),
    )
    if "Razon Estado" in trabajo.columns:
        trabajo["RazonEstado"] = trabajo["Razon Estado"].astype(str).str.strip()
    else:
        trabajo["RazonEstado"] = ""

    trabajo["NivelEducativoNorm"] = trabajo["NivelEducativo"].apply(_normalize_key)
    trabajo["AsignaturaProductoNorm"] = trabajo["AsignaturaProducto"].apply(_normalize_key)

    trabajo = trabajo[trabajo["Plataforma"] == "Compartir Aprendizajes"]
    trabajo = trabajo[
        (trabajo["RazonEstado"] == "Validado") | (trabajo["RazonEstado"] == "")
    ]

    if trabajo.empty:
        return pd.DataFrame()

    trabajo = trabajo[
        (trabajo["AsignaturaProductoNorm"] != "NO APLICA")
        & (trabajo["NivelEducativoNorm"] != "EDUCACION INICIAL")
    ]

    registros: List[Dict[str, str]] = []
    for _, fila in trabajo.iterrows():
        nivel_legible = _mapear_nivel(fila["NivelEducativo"])
        grado_legible, grado_num = _mapear_grado(fila["GradoVal"])
        materia_legible, sufijo = _mapear_materia(fila["AsignaturaProducto"])

        if not (nivel_legible and grado_legible and grado_num and materia_legible and sufijo):
            continue

        nombre_clase = f"{materia_legible} {grado_num}{sufijo}"
        registros.append(
            {
                "Nivel": nivel_legible,
                "Grado": grado_legible,
                "Grupo": "Grupo A",
                "Nombre de Clase": nombre_clase,
                "Clase Clave": nombre_clase,
                "Alias Clase": "",
                "Materias": materia_legible,
                "_orden_nivel": 0 if nivel_legible == "Primaria" else 1,
                "_orden_grado": grado_num,
                "_orden_materia": 0 if materia_legible == "Comunicación" else 1,
            }
        )

    if not registros:
        return pd.DataFrame()

    salida = pd.DataFrame(registros)
    salida = salida.sort_values(
        by=["_orden_nivel", "_orden_grado", "_orden_materia", "Materias"],
        ascending=[True, True, True, True],
    ).drop(columns=["_orden_nivel", "_orden_grado", "_orden_materia"])

    return salida.reset_index(drop=True)


def exportar_excel(
    df: pd.DataFrame,
    ruta_salida: Path,
    codigo: str,
    plantilla: Path = Path(OUTPUT_FILENAME),
) -> None:
    """
    Exporta el DataFrame a Excel:
    - Parte de la plantilla PlantillaClases.xlsx (si existe); si no, crea un libro nuevo.
    - Mantiene el formato de la hoja existente y limpia solo las filas de datos.
    - Guarda todo en ruta_salida.
    """
    try:
        wb = load_workbook(plantilla)
    except Exception:
        wb = Workbook()
        wb.remove(wb.active)

    if OUTPUT_SHEET_NAME in wb.sheetnames:
        ws_out = wb[OUTPUT_SHEET_NAME]
        # Limpiar filas de datos (dejamos encabezados)
        if ws_out.max_row > 1:
            ws_out.delete_rows(2, ws_out.max_row - 1)
        headers = [cell.value for cell in ws_out[1]]
    else:
        ws_out = wb.create_sheet(OUTPUT_SHEET_NAME)
        headers = list(df.columns)
        ws_out.append(headers)

    # Aseguramos el orden de columnas según los headers existentes
    try:
        df_to_write = df[headers]
    except KeyError as exc:
        raise KeyError(f"No coinciden los encabezados de la plantilla con las columnas generadas: {exc}")

    for row in dataframe_to_rows(df_to_write, index=False, header=False):
        ws_out.append(row)

    try:
        wb.save(ruta_salida)
    except Exception as exc:  # pragma: no cover - defensa general
        raise RuntimeError(f"No se pudo exportar el Excel: {exc}") from exc


def _parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filtra y transforma datos de la hoja de detalle."
    )
    parser.add_argument("ruta_excel", help="Ruta del archivo XLSX de entrada")
    parser.add_argument("codigo", help="Código a filtrar (se respetan ceros a la izquierda)")
    parser.add_argument(
        "--columna-codigo",
        dest="columna_codigo",
        default=CODE_COLUMN_NAME,
        help=f"Nombre de la columna que contiene el código (default: {CODE_COLUMN_NAME})",
    )
    parser.add_argument(
        "--hoja",
        dest="hoja",
        default=SHEET_NAME,
        help=f"Nombre de la hoja a leer (default: {SHEET_NAME})",
    )
    return parser.parse_args(argv[1:])


def main(argv: List[str]) -> int:
    args = _parse_args(argv)
    ruta_archivo = Path(args.ruta_excel)

    try:
        df = cargar_excel(ruta_archivo, hoja=args.hoja, columna_codigo=args.columna_codigo)
        df_filtrado = filtrar_codigo(df, args.codigo, args.columna_codigo)

        if df_filtrado.empty:
            print(f"No se encontraron filas para el código {args.codigo}.")
            return 0

        mostrar_resumen(df_filtrado)

        df_transformado = transformar(df_filtrado)
        if df_transformado.empty:
            print(
                "No hay filas que cumplan con los filtros de 'Compartir Aprendizajes' y reglas de transformación."
            )
            return 0

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        archivo_salida = OUTPUT_DIR / f"{OUTPUT_FILENAME.split('.xlsx')[0]}_{args.codigo}.xlsx"
        exportar_excel(
            df_transformado,
            ruta_salida=archivo_salida,
            codigo=args.codigo,
            plantilla=Path(OUTPUT_FILENAME),
        )
        print(f"\nArchivo generado: {archivo_salida.resolve()}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
