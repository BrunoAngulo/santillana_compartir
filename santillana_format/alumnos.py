import re
import unicodedata
from io import BytesIO
from typing import Callable, Dict, List, Optional, Sequence, Tuple

import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_EMPRESA_ID = 11
DEFAULT_CICLO_ID = 207
BASE_URL = (
    "https://www.uno-internacional.com/pegasus-api/censo/empresas/{empresa_id}"
    "/ciclos/{ciclo_id}/colegios/{colegio_id}/alumnos"
)
PLANTILLA_EDICION_URL = (
    "https://www.uno-internacional.com/pegasus-api/censo/empresas/{empresa_id}"
    "/ciclos/{ciclo_id}/colegios/{colegio_id}/descargarPlantillaEdicionMasiva"
)

NIVEL_MAP = {
    "Inicial": 38,
    "Primaria": 39,
    "Secundaria": 40,
}

GRADOS_POR_NIVEL = {
    38: {
        113: "3 anos",
        114: "4 anos",
        115: "5 anos",
        116: "6 anos",
        117: "7 anos",
        118: "8 anos",
    },
    39: {
        119: "1ro primaria",
        120: "2do primaria",
        121: "3ro primaria",
        122: "4to primaria",
        123: "5to primaria",
    },
    40: {
        126: "1ro secundaria",
        127: "2do secundaria",
        128: "3ro secundaria",
        129: "4to secundaria",
        130: "5to secundaria",
    },
}

GRUPO_ID_TO_LETRA = {
    661: "A",
    662: "B",
    663: "C",
    664: "D",
    665: "E",
    666: "F",
    667: "G",
    668: "H",
    669: "I",
    670: "J",
    671: "K",
}
GRUPO_LETRA_TO_ID = {letra: grupo_id for grupo_id, letra in GRUPO_ID_TO_LETRA.items()}

NIVEL_ORDER = {38: 0, 39: 1, 40: 2}
NIVEL_NAME_ORDER = {"inicial": 0, "primaria": 1, "secundaria": 2}
ORDINAL_MAP = {
    "primer": 1,
    "primero": 1,
    "segundo": 2,
    "tercer": 3,
    "tercero": 3,
    "cuarto": 4,
    "quinto": 5,
    "sexto": 6,
    "septimo": 7,
    "octavo": 8,
}

PLANTILLA_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "NUI",
    "Id Alumno",
    "Activo",
    "Nombre",
    "Apellido Paterno",
    "Apellido materno",
    "Sexo",
    "Fecha de Nacimiento",
    "Extranjero",
    "NUIP",
    "Login",
    "Password",
]
PLANTILLA_ACTUALIZADA_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "Nombre",
    "Apellido Paterno",
    "Apellido materno",
    "Sexo",
    "Fecha de Nacimiento",
    "NUIP",
    "Login",
    "Password",
]

GRADO_OPTIONS = [
    "2 años",
    "3 años",
    "4 años",
    "5 años",
    "Primer grado de primaria",
    "Segundo grado de primaria",
    "Tercer grado de primaria",
    "Cuarto grado de primaria",
    "Quinto grado de primaria",
    "Sexto grado de primaria",
    "Primer año de secundaria",
    "Segundo año de secundaria",
    "Tercer año de secundaria",
    "Cuarto año de secundaria",
    "Quinto año de secundaria",
]

GRUPO_OPTIONS = (
    [f"Grupo {letra}" for letra in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]
    + ["Grupo BM"]
)
ACTIVO_OPTIONS = ["Si", "No"]
SEXO_OPTIONS = ["F", "M"]
NIVEL_OPTIONS = ["Inicial", "Primaria", "Secundaria"]

ALUMNO_COLUMNS = [
    "colegio_id",
    "nivel_id",
    "nivel",
    "grado_id",
    "grado",
    "grupo_id",
    "grupo",
    "grupo_clave",
    "alumno_id",
    "persona_id",
    "nombre_completo",
    "nombre",
    "apellido_paterno",
    "apellido_materno",
    "sexo",
    "id_oficial",
    "fecha_nacimiento",
    "activo",
    "alumno_clave",
    "fecha_desde",
    "fecha_validado",
]

ERROR_COLUMNS = [
    "colegio_id",
    "nivel_id",
    "grado_id",
    "grupo_id",
    "url",
    "status_code",
    "error",
]


def build_alumnos_filename(colegio_ids: Sequence[int]) -> str:
    if not colegio_ids:
        return "alumnos.xlsx"
    ids = sorted({int(colegio_id) for colegio_id in colegio_ids})
    if len(ids) == 1:
        return f"alumnos_{ids[0]}.xlsx"
    joined = "_".join(str(colegio_id) for colegio_id in ids)
    return f"alumnos_{joined}.xlsx"


def parse_id_list(text: str) -> List[int]:
    if not text:
        return []
    ids: List[int] = []
    for token in re.split(r"[\s,]+", text.strip()):
        if not token:
            continue
        if "-" in token:
            inicio, fin = token.split("-", 1)
            if inicio.isdigit() and fin.isdigit():
                ids.extend(range(int(inicio), int(fin) + 1))
            continue
        if token.isdigit():
            ids.append(int(token))
    return sorted(set(ids))


def build_request_contexts(
    colegio_ids: Sequence[int],
    nivel_ids: Sequence[int],
    grupo_ids: Sequence[int],
    empresa_id: int = DEFAULT_EMPRESA_ID,
    ciclo_id: int = DEFAULT_CICLO_ID,
) -> List[Dict[str, int]]:
    contexts: List[Dict[str, int]] = []
    for colegio_id in colegio_ids:
        for nivel_id in nivel_ids:
            grados = GRADOS_POR_NIVEL.get(nivel_id, {})
            for grado_id in grados.keys():
                for grupo_id in grupo_ids:
                    contexts.append(
                        {
                            "colegio_id": colegio_id,
                            "nivel_id": nivel_id,
                            "grado_id": grado_id,
                            "grupo_id": grupo_id,
                            "empresa_id": empresa_id,
                            "ciclo_id": ciclo_id,
                        }
                    )
    return contexts


def _build_url(context: Dict[str, int]) -> Tuple[str, Dict[str, int]]:
    url = BASE_URL.format(
        empresa_id=context["empresa_id"],
        ciclo_id=context["ciclo_id"],
        colegio_id=context["colegio_id"],
    )
    params = {
        "nivelId": context["nivel_id"],
        "gradoId": context["grado_id"],
        "grupoId": context["grupo_id"],
    }
    return url, params


def _fetch_alumnos(
    session: requests.Session,
    token: str,
    context: Dict[str, int],
    timeout: int = 30,
) -> Tuple[List[Dict[str, object]], Optional[str], Optional[int], str]:
    url, params = _build_url(context)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }
    try:
        response = session.get(url, headers=headers, params=params, timeout=timeout)
    except requests.RequestException as exc:
        return [], str(exc), None, url

    status_code = response.status_code
    try:
        payload = response.json()
    except ValueError:
        return [], f"Respuesta no JSON (status {status_code})", status_code, url

    if not response.ok:
        message = ""
        if isinstance(payload, dict):
            message = payload.get("message") or payload.get("error") or ""
        return [], message or f"HTTP {status_code}", status_code, url

    if not isinstance(payload, dict) or not payload.get("success", False):
        message = payload.get("message") if isinstance(payload, dict) else "Respuesta invalida"
        return [], message or "Respuesta invalida", status_code, url

    data = payload.get("data") or []
    if not isinstance(data, list):
        return [], "Campo data no es lista", status_code, url
    return data, None, status_code, url


def _flatten_alumno(item: Dict[str, object], context: Dict[str, int]) -> Dict[str, object]:
    persona = item.get("persona") or {}
    nivel = item.get("nivel") or {}
    grado = item.get("grado") or {}
    grupo = item.get("grupo") or {}
    return {
        "colegio_id": context["colegio_id"],
        "nivel_id": nivel.get("nivelId", context["nivel_id"]),
        "nivel": nivel.get("nivel", ""),
        "grado_id": grado.get("gradoId", context["grado_id"]),
        "grado": grado.get("grado", ""),
        "grupo_id": grupo.get("grupoId", context["grupo_id"]),
        "grupo": grupo.get("grupo", ""),
        "grupo_clave": grupo.get("grupoClave", ""),
        "alumno_id": item.get("alumnoId", ""),
        "persona_id": persona.get("personaId", ""),
        "nombre_completo": persona.get("nombreCompleto", ""),
        "nombre": persona.get("nombre", ""),
        "apellido_paterno": persona.get("apellidoPaterno", ""),
        "apellido_materno": persona.get("apellidoMaterno", ""),
        "sexo": persona.get("sexoMoral", ""),
        "id_oficial": persona.get("idOficial", ""),
        "fecha_nacimiento": persona.get("fechaNacimiento", ""),
        "activo": item.get("activo", ""),
        "alumno_clave": item.get("alumnoClave", ""),
        "fecha_desde": item.get("fechaDesde", ""),
        "fecha_validado": item.get("fechaValidado", ""),
    }


def _ensure_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=columns)
    return df.reindex(columns=columns)


def export_alumnos_excel(
    alumnos: List[Dict[str, object]],
    errores: List[Dict[str, object]],
) -> bytes:
    output = BytesIO()
    df_alumnos = _ensure_columns(pd.DataFrame(alumnos), ALUMNO_COLUMNS)
    df_errores = _ensure_columns(pd.DataFrame(errores), ERROR_COLUMNS)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_alumnos.to_excel(writer, index=False, sheet_name="Alumnos")
        df_errores.to_excel(writer, index=False, sheet_name="Errores")
        for sheet_name, df in (("Alumnos", df_alumnos), ("Errores", df_errores)):
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for idx, col in enumerate(df.columns, start=1):
                sample = df[col].astype(str).head(200).tolist()
                max_len = max([len(str(col))] + [len(val) for val in sample])
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    output.seek(0)
    return output.getvalue()


def descargar_plantilla_edicion_masiva(
    token: str,
    colegio_id: int,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    ciclo_id: int = DEFAULT_CICLO_ID,
    timeout: int = 30,
) -> Tuple[bytes, Dict[str, int]]:
    url = PLANTILLA_EDICION_URL.format(
        empresa_id=empresa_id,
        ciclo_id=ciclo_id,
        colegio_id=colegio_id,
    )
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    try:
        response = requests.get(url, headers=headers, params={"descargar": 0}, timeout=timeout)
    except requests.RequestException as exc:
        raise RuntimeError(f"Error de red: {exc}") from exc

    status_code = response.status_code
    try:
        payload = response.json()
    except ValueError as exc:
        raise RuntimeError(f"Respuesta no JSON (status {status_code})") from exc

    if not response.ok:
        message = payload.get("message") if isinstance(payload, dict) else ""
        raise RuntimeError(message or f"HTTP {status_code}")

    if not isinstance(payload, dict) or not payload.get("success", False):
        message = payload.get("message") if isinstance(payload, dict) else "Respuesta invalida"
        raise RuntimeError(message or "Respuesta invalida")

    data = payload.get("data") or []
    if not isinstance(data, list):
        raise RuntimeError("Campo data no es lista")

    rows = _sort_plantilla_rows(data)
    mapped_rows = [_map_plantilla_row(item) for item in rows]
    df = pd.DataFrame(mapped_rows, columns=PLANTILLA_COLUMNS)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name_bd = "Plantilla_BD"
        sheet_name_actualizada = "Plantilla_Actualizada"
        df_actualizada = pd.DataFrame(columns=PLANTILLA_ACTUALIZADA_COLUMNS)
        df.to_excel(writer, index=False, sheet_name=sheet_name_bd)
        df_actualizada.to_excel(writer, index=False, sheet_name=sheet_name_actualizada)
        workbook = writer.book
        catalog = _build_catalogos(workbook)

        for sheet_name, df_sheet in (
            (sheet_name_bd, df),
            (sheet_name_actualizada, df_actualizada),
        ):
            ws = workbook[sheet_name]
            _apply_plantilla_validations(ws, df_sheet, catalog)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for idx, col in enumerate(df_sheet.columns, start=1):
                sample = df_sheet[col].astype(str).head(200).tolist()
                max_len = max([len(str(col))] + [len(val) for val in sample])
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    output.seek(0)
    summary = {"alumnos_total": len(rows)}
    return output.getvalue(), summary


def listar_alumnos(
    token: str,
    colegio_ids: Sequence[int],
    nivel_ids: Optional[Sequence[int]] = None,
    grupo_ids: Optional[Sequence[int]] = None,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    ciclo_id: int = DEFAULT_CICLO_ID,
    timeout: int = 30,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[bytes, Dict[str, int]]:
    nivel_ids = list(nivel_ids) if nivel_ids else list(NIVEL_MAP.values())
    grupo_ids = list(grupo_ids) if grupo_ids else list(GRUPO_ID_TO_LETRA.keys())

    contexts = build_request_contexts(
        colegio_ids=colegio_ids,
        nivel_ids=nivel_ids,
        grupo_ids=grupo_ids,
        empresa_id=empresa_id,
        ciclo_id=ciclo_id,
    )

    alumnos: List[Dict[str, object]] = []
    errores: List[Dict[str, object]] = []

    with requests.Session() as session:
        for index, context in enumerate(contexts, start=1):
            data, error, status_code, url = _fetch_alumnos(
                session=session,
                token=token,
                context=context,
                timeout=timeout,
            )
            if error:
                errores.append(
                    {
                        "colegio_id": context["colegio_id"],
                        "nivel_id": context["nivel_id"],
                        "grado_id": context["grado_id"],
                        "grupo_id": context["grupo_id"],
                        "url": url,
                        "status_code": status_code or "",
                        "error": error,
                    }
                )
            else:
                for item in data:
                    alumnos.append(_flatten_alumno(item, context))

            if on_progress:
                on_progress(index, len(contexts))

    output_bytes = export_alumnos_excel(alumnos, errores)
    summary = {
        "solicitudes_total": len(contexts),
        "solicitudes_error": len(errores),
        "alumnos_total": len(alumnos),
    }
    return output_bytes, summary


def _sort_plantilla_rows(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    def _safe_int(value: object) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            try:
                return int(value)
            except (TypeError, ValueError):
                return None
        text = re.sub(r"\D", "", str(value))
        if not text:
            return None
        try:
            return int(text)
        except ValueError:
            return None

    def _normalize(value: object) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFD", text)
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        return text

    def _nivel_order(item: Dict[str, object]) -> int:
        nivel_id = _safe_int(item.get("nivelId"))
        if nivel_id in NIVEL_ORDER:
            return NIVEL_ORDER[nivel_id]
        nivel_name = _normalize(item.get("nivel"))
        return NIVEL_NAME_ORDER.get(nivel_name, 99)

    def _grado_order(item: Dict[str, object]) -> int:
        grado_id = _safe_int(item.get("gradoId"))
        if grado_id is not None:
            return grado_id
        grado_text = _normalize(item.get("grado"))
        digits = re.findall(r"\d+", grado_text)
        if digits:
            return int(digits[0])
        for key, value in ORDINAL_MAP.items():
            if key in grado_text:
                return value
        return 999

    def _grupo_order(item: Dict[str, object]) -> int:
        grupo_id = _safe_int(item.get("grupoId"))
        if grupo_id is not None:
            return grupo_id
        grupo_text = _normalize(item.get("grupo"))
        match = re.search(r"\b([a-z])\b", grupo_text)
        if match:
            return ord(match.group(1).upper()) - ord("A") + 1
        return 999

    def _name_key(item: Dict[str, object]) -> Tuple[str, str, str]:
        ap_pat = _normalize(item.get("apellidoPaterno") or item.get("apellido_paterno"))
        ap_mat = _normalize(item.get("apellidoMaterno") or item.get("apellido_materno"))
        nombre = _normalize(item.get("nombre"))
        return ap_pat, ap_mat, nombre

    return sorted(
        rows,
        key=lambda item: (
            _grado_order(item),
            _grupo_order(item),
            _nivel_order(item),
            *_name_key(item),
        ),
    )


def _map_plantilla_row(item: Dict[str, object]) -> Dict[str, object]:
    activo = item.get("activo", "")
    if isinstance(activo, bool):
        activo = "Si" if activo else "No"
    return {
        "Nivel": item.get("nivel", ""),
        "Grado": item.get("grado", ""),
        "Grupo": item.get("grupo", ""),
        "NUI": item.get("personaId", ""),
        "Id Alumno": item.get("alumnoId", ""),
        "Activo": activo,
        "Nombre": item.get("nombre", ""),
        "Apellido Paterno": item.get("apellidoPaterno", ""),
        "Apellido materno": item.get("apellidoMaterno", ""),
        "Sexo": item.get("sexo", ""),
        "Fecha de Nacimiento": item.get("fechaNacimiento", ""),
        "Extranjero": item.get("extranjero", ""),
        "NUIP": item.get("idOficial", ""),
        "Login": item.get("login", ""),
        "Password": item.get("password", ""),
    }


def _build_catalogos(workbook):
    if "Catalogos" in workbook.sheetnames:
        del workbook["Catalogos"]
    catalog = workbook.create_sheet("Catalogos")
    _write_list_column(catalog, 1, "Nivel", NIVEL_OPTIONS)
    _write_list_column(catalog, 2, "Grado", GRADO_OPTIONS)
    _write_list_column(catalog, 3, "Grupo", GRUPO_OPTIONS)
    _write_list_column(catalog, 4, "Activo", ACTIVO_OPTIONS)
    _write_list_column(catalog, 5, "Sexo", SEXO_OPTIONS)
    catalog.sheet_state = "hidden"
    return catalog


def _apply_plantilla_validations(ws, df: pd.DataFrame, catalog) -> None:
    header_map = {name: idx + 1 for idx, name in enumerate(df.columns)}
    _add_list_validation(ws, catalog, header_map.get("Nivel"), 1, len(NIVEL_OPTIONS), df)
    _add_list_validation(ws, catalog, header_map.get("Grado"), 2, len(GRADO_OPTIONS), df)
    _add_list_validation(ws, catalog, header_map.get("Grupo"), 3, len(GRUPO_OPTIONS), df)
    _add_list_validation(ws, catalog, header_map.get("Activo"), 4, len(ACTIVO_OPTIONS), df)
    _add_list_validation(ws, catalog, header_map.get("Sexo"), 5, len(SEXO_OPTIONS), df)


def _write_list_column(ws, col: int, title: str, values: List[str]) -> None:
    ws.cell(row=1, column=col, value=title)
    for idx, value in enumerate(values, start=2):
        ws.cell(row=idx, column=col, value=value)


def _add_list_validation(
    ws,
    catalog_ws,
    target_col: Optional[int],
    catalog_col: int,
    total: int,
    df: pd.DataFrame,
) -> None:
    if not target_col or total <= 0:
        return
    max_row = max(int(df.shape[0]) + 1, 2)
    col_letter = get_column_letter(target_col)
    catalog_letter = get_column_letter(catalog_col)
    formula = f"=Catalogos!${catalog_letter}$2:${catalog_letter}${total + 1}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Selecciona un valor de la lista."
    dv.errorTitle = "Valor invalido"
    dv.prompt = "Selecciona un valor permitido."
    dv.promptTitle = "Lista"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")
