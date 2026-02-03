import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter

DEFAULT_SHEET_BD = "Plantilla_BD"
DEFAULT_SHEET_ACTUALIZADA = "Plantilla_Actualizada"
ALUMNOS_CREAR_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "Nombre",
    "Apellido Paterno",
    "Apellido materno",
    "Sexo",
    "Fecha de Nacimiento",
    "NUIP",
    "Login",
    "Password",
]
ALUMNOS_COMPARACION_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "NUI",
    "Id Alumno",
    "Activo",
    "Nombre",
    "Apellido Paterno",
    "Apellido materno",
    "Sexo",
    "Fecha de Nacimiento",
    "Extranjero",
    "NUIP",
    "Login",
    "Password",
    "Nuevo Nivel",
    "Nuevo Grado",
    "Nuevo Grupo",
]

BASE_OUTPUT_MAP = {
    "nivel": "Nivel",
    "grado": "Grado",
    "grupo": "Grupo",
    "nui": "NUI",
    "id_alumno": "Id Alumno",
    "activo": "Activo",
    "nombre": "Nombre",
    "apellido_paterno": "Apellido Paterno",
    "apellido_materno": "Apellido materno",
    "sexo": "Sexo",
    "fecha_nacimiento": "Fecha de Nacimiento",
    "extranjero": "Extranjero",
    "nuip": "NUIP",
    "login": "Login",
    "password": "Password",
}

ACTUALIZADA_OUTPUT_MAP = {
    "nivel": "Nuevo Nivel",
    "grado": "Nuevo Grado",
    "grupo": "Nuevo Grupo",
}

ACTUALIZADA_BASE_FIELDS = {
    "nombre": "Nombre",
    "apellido_paterno": "Apellido Paterno",
    "apellido_materno": "Apellido materno",
    "sexo": "Sexo",
    "fecha_nacimiento": "Fecha de Nacimiento",
    "nuip": "NUIP",
    "login": "Login",
    "password": "Password",
}

HEADER_ALIASES = {
    "nivel": "nivel",
    "grado": "grado",
    "grupo": "grupo",
    "nui": "nui",
    "id alumno": "id_alumno",
    "idalumno": "id_alumno",
    "activo": "activo",
    "nombre": "nombre",
    "apellido paterno": "apellido_paterno",
    "apellido materno": "apellido_materno",
    "apellido materno ": "apellido_materno",
    "sexo": "sexo",
    "fecha de nacimiento": "fecha_nacimiento",
    "fecha nacimiento": "fecha_nacimiento",
    "extranjero": "extranjero",
    "nuip": "nuip",
    "dni": "nuip",
    "login": "login",
    "password": "password",
}

MATCH_TIPO_N1 = "N1_NUIP"
MATCH_TIPO_N2 = "N2_APELLIDOS"
MATCH_TIPO_N3 = "N3_APELLIDOS_INICIAL"


def comparar_plantillas(
    excel_path: Path,
    sheet_bd: str = DEFAULT_SHEET_BD,
    sheet_actualizada: str = DEFAULT_SHEET_ACTUALIZADA,
) -> Tuple[bytes, Dict[str, int]]:
    df_bd = _read_sheet(excel_path, sheet_bd)
    df_act = _read_sheet(excel_path, sheet_actualizada)

    df_bd = _canonicalize_columns(df_bd)
    df_act = _canonicalize_columns(df_act)

    summary = _build_login_summary(df_bd, df_act)
    comparacion = _build_comparacion_bd(df_bd, df_act)

    output = Path(excel_path).name
    output_bytes = _export_comparacion(comparacion)
    summary.update(
        {
            "actualizados_total": len(df_act),
            "base_total": len(df_bd),
            "archivo_base": output,
        }
    )
    return output_bytes, summary


def _read_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")
    with pd.ExcelFile(excel_path, engine="openpyxl") as excel:
        if sheet_name == DEFAULT_SHEET_BD:
            resolved = _resolve_sheet_name_fallback(
                excel.sheet_names,
                sheet_name,
                ["Plantilla edicion masiva", "Plantilla edición masiva"],
            )
        else:
            resolved = _resolve_sheet_name(excel.sheet_names, sheet_name)
        df = pd.read_excel(excel, sheet_name=resolved, dtype=str)
    return df.fillna("")


def _resolve_sheet_name(available: Sequence[str], desired: str) -> str:
    if desired in available:
        return desired
    desired_lower = desired.lower()
    for sheet in available:
        if sheet.lower() == desired_lower:
            return sheet
    desired_norm = _normalize_header(desired)
    for sheet in available:
        if _normalize_header(sheet) == desired_norm:
            return sheet
    available_list = ", ".join(available) if available else "(sin hojas)"
    raise ValueError(
        f"No se encontro la hoja '{desired}'. Hojas disponibles: {available_list}."
    )


def _resolve_sheet_name_fallback(
    available: Sequence[str], desired: str, fallbacks: Sequence[str]
) -> str:
    try:
        return _resolve_sheet_name(available, desired)
    except ValueError:
        for fallback in fallbacks:
            try:
                return _resolve_sheet_name(available, fallback)
            except ValueError:
                continue
    available_list = ", ".join(available) if available else "(sin hojas)"
    raise ValueError(
        f"No se encontro la hoja '{desired}'. Hojas disponibles: {available_list}."
    )


def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    used = set()
    for col in df.columns:
        key = _normalize_header(col)
        canonical = HEADER_ALIASES.get(key)
        if canonical and canonical not in used:
            mapping[col] = canonical
            used.add(canonical)
    return df.rename(columns=mapping)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text)
    return text.strip().lower()


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _normalize_date(value: object) -> str:
    text = str(value or "")
    digits = re.sub(r"\D", "", text)
    return digits


def _normalize_nuip(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _first_initial(value: object) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    return text[0]


def _comment_for_match(match_type: Optional[str]) -> str:
    if match_type == MATCH_TIPO_N1:
        return "Nivel 1: NUIP"
    if match_type == MATCH_TIPO_N2:
        return "Nivel 2: Apellidos"
    if match_type == MATCH_TIPO_N3:
        return "Nivel 3: Apellidos + inicial"
    return "No encontrado"


def _format_match_flag(value: Optional[bool]) -> str:
    if value is True:
        return "Si"
    if value is False:
        return "No"
    return ""


def _build_base_indexes(df_bd: pd.DataFrame) -> Dict[str, Dict[str, List[int]]]:
    indexes = {"nuip": {}, "apellidos": {}, "apellidos_inicial": {}}
    for idx, row in df_bd.iterrows():
        nuip = _normalize_nuip(row.get("nuip"))
        if nuip:
            indexes["nuip"].setdefault(nuip, []).append(idx)
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        ap_mat = _normalize_text(row.get("apellido_materno"))
        if not (ap_pat and ap_mat):
            continue
        key = f"{ap_pat}|{ap_mat}"
        indexes["apellidos"].setdefault(key, []).append(idx)
        inicial = _first_initial(row.get("nombre"))
        if inicial:
            key_inicial = f"{ap_pat}|{ap_mat}|{inicial}"
            indexes["apellidos_inicial"].setdefault(key_inicial, []).append(idx)
    return indexes


def _clasificar_actualizados(
    df_act: pd.DataFrame,
    df_bd: pd.DataFrame,
    indexes: Dict[str, Dict[str, List[int]]],
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    rows = []
    counts = {
        "nivel1": 0,
        "nivel2": 0,
        "nivel3": 0,
        "sin_match": 0,
        "grado_no": 0,
        "grupo_no": 0,
    }

    for _idx, row in df_act.iterrows():
        match_type, match_indices = _match_row(row, indexes)
        enriched = row.copy()
        enriched["Match Tipo"] = match_type or ""
        bd_grado = _collect_matches(df_bd, match_indices, "grado")
        bd_grupo = _collect_matches(df_bd, match_indices, "grupo")
        grado_ok = _match_field(row.get("grado"), df_bd, match_indices, "grado")
        grupo_ok = _match_field(row.get("grupo"), df_bd, match_indices, "grupo")
        enriched["BD Grado"] = bd_grado
        enriched["BD Grupo"] = bd_grupo
        enriched["Grado coincide"] = _format_match_flag(grado_ok)
        enriched["Grupo coincide"] = _format_match_flag(grupo_ok)
        comentario = _comment_for_match(match_type)
        if match_type:
            if grado_ok is False:
                comentario = f"{comentario} | Grado no coincide"
                counts["grado_no"] += 1
            if grupo_ok is False:
                comentario = f"{comentario} | Grupo no coincide"
                counts["grupo_no"] += 1
        enriched["Comentario"] = comentario

        if match_type == MATCH_TIPO_N1:
            counts["nivel1"] += 1
        elif match_type == MATCH_TIPO_N2:
            counts["nivel2"] += 1
        elif match_type == MATCH_TIPO_N3:
            counts["nivel3"] += 1
        else:
            counts["sin_match"] += 1
        rows.append(enriched)

    return pd.DataFrame(rows), counts


def _match_row(
    row: pd.Series, indexes: Dict[str, Dict[str, List[int]]]
) -> Tuple[Optional[str], List[int]]:
    nuip = _normalize_nuip(row.get("nuip"))
    if nuip and nuip in indexes["nuip"]:
        return MATCH_TIPO_N1, indexes["nuip"][nuip]
    ap_pat = _normalize_text(row.get("apellido_paterno"))
    ap_mat = _normalize_text(row.get("apellido_materno"))
    if ap_pat and ap_mat:
        key = f"{ap_pat}|{ap_mat}"
        if key in indexes["apellidos"]:
            return MATCH_TIPO_N2, indexes["apellidos"][key]
        inicial = _first_initial(row.get("nombre"))
        if inicial:
            key_inicial = f"{ap_pat}|{ap_mat}|{inicial}"
            if key_inicial in indexes["apellidos_inicial"]:
                return MATCH_TIPO_N3, indexes["apellidos_inicial"][key_inicial]
    return None, []


def _collect_matches(
    df_bd: pd.DataFrame, indices: Sequence[int], column: str
) -> str:
    if not indices:
        return ""
    values = []
    for idx in indices:
        value = str(df_bd.loc[idx].get(column, "") or "").strip()
        if value and value not in values:
            values.append(value)
    return ", ".join(values)


def _match_field(
    value: object, df_bd: pd.DataFrame, indices: Sequence[int], column: str
) -> Optional[bool]:
    if not indices:
        return None
    current = _normalize_text(value)
    if not current:
        return None
    for idx in indices:
        bd_value = _normalize_text(df_bd.loc[idx].get(column))
        if bd_value and bd_value == current:
            return True
    return False


def _export_comparacion(comparacion: pd.DataFrame) -> bytes:
    from io import BytesIO

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, "Plantilla edición masiva", comparacion)
    output.seek(0)
    return output.getvalue()


def _write_sheet(writer: pd.ExcelWriter, name: str, df: pd.DataFrame) -> None:
    df = df.copy()
    df.to_excel(writer, index=False, sheet_name=name)
    ws = writer.book[name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _apply_date_format(ws, df, "Fecha de Nacimiento")
    for idx, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).head(200).tolist()
        max_len = max([len(str(col))] + [len(val) for val in sample])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)


def _apply_date_format(ws, df: pd.DataFrame, column_name: str) -> None:
    if column_name not in df.columns:
        return
    col_idx = list(df.columns).index(column_name) + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            cell.number_format = "dd/mm/yyyy"


def _build_alumnos_crear(nuevos: pd.DataFrame) -> pd.DataFrame:
    if nuevos.empty:
        return pd.DataFrame(columns=ALUMNOS_CREAR_COLUMNS)
    rename_map = {
        "nivel": "Nivel",
        "grado": "Grado",
        "grupo": "Grupo",
        "nombre": "Nombre",
        "apellido_paterno": "Apellido Paterno",
        "apellido_materno": "Apellido materno",
        "sexo": "Sexo",
        "fecha_nacimiento": "Fecha de Nacimiento",
        "nuip": "NUIP",
        "login": "Login",
        "password": "Password",
    }
    df = nuevos.rename(columns=rename_map)
    for col in ALUMNOS_CREAR_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df.loc[:, ALUMNOS_CREAR_COLUMNS].copy()
    if "Fecha de Nacimiento" in df.columns:
        df["Fecha de Nacimiento"] = df["Fecha de Nacimiento"].apply(_parse_fecha_excel)
    cleaned = df.astype(str).apply(lambda col: col.str.strip().replace("nan", ""))
    mask = (cleaned != "").any(axis=1)
    return df.loc[mask].reset_index(drop=True)


def _clean_cell_value(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _normalize_grupo(value: object) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"^grupo", "", text)
    match = re.search(r"[a-z]", text)
    if match:
        return match.group(0)
    digits = re.findall(r"\d+", text)
    if digits:
        return digits[0]
    return text


def _normalize_login(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    return text


def _build_login_index(df: pd.DataFrame) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for idx, row in df.iterrows():
        login = _normalize_login(row.get("login"))
        if not login:
            continue
        if login not in index:
            index[login] = idx
    return index


def _build_login_summary(df_bd: pd.DataFrame, df_act: pd.DataFrame) -> Dict[str, int]:
    bd_logins = set()
    for _idx, row in df_bd.iterrows():
        login = _normalize_login(row.get("login"))
        if login:
            bd_logins.add(login)

    act_logins = set()
    for _idx, row in df_act.iterrows():
        login = _normalize_login(row.get("login"))
        if login:
            act_logins.add(login)

    matched = bd_logins.intersection(act_logins)
    return {
        "login_bd_total": len(bd_logins),
        "login_actualizada_total": len(act_logins),
        "login_match": len(matched),
        "login_sin_bd": len(act_logins - bd_logins),
        "login_sin_actualizada": len(bd_logins - act_logins),
    }


def _is_changed(new_value: object, old_value: object, normalizer) -> bool:
    new_norm = normalizer(new_value)
    if not new_norm:
        return False
    old_norm = normalizer(old_value)
    if not old_norm:
        return True
    return new_norm != old_norm


def _pick_best_match(
    base_row: pd.Series, df_act: pd.DataFrame, indices: Sequence[int]
) -> Optional[int]:
    if not indices:
        return None
    if len(indices) == 1:
        return indices[0]

    target_nombre = _normalize_text(base_row.get("nombre"))
    target_fecha = _normalize_date(base_row.get("fecha_nacimiento"))
    target_sexo = _normalize_text(base_row.get("sexo"))

    best_idx = indices[0]
    best_score = -1
    for idx in indices:
        candidate = df_act.loc[idx]
        score = 0
        if target_nombre and _normalize_text(candidate.get("nombre")) == target_nombre:
            score += 2
        if target_fecha and _normalize_date(candidate.get("fecha_nacimiento")) == target_fecha:
            score += 1
        if target_sexo and _normalize_text(candidate.get("sexo")) == target_sexo:
            score += 1
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _build_comparacion_bd(df_bd: pd.DataFrame, df_act: pd.DataFrame) -> pd.DataFrame:
    if df_bd.empty or df_act.empty:
        return pd.DataFrame(columns=ALUMNOS_COMPARACION_COLUMNS)
    bd_index = _build_login_index(df_bd)
    rows: List[Dict[str, object]] = []

    for _idx, act_row in df_act.iterrows():
        login_norm = _normalize_login(act_row.get("login"))
        if not login_norm:
            continue
        bd_idx = bd_index.get(login_norm)
        if bd_idx is None:
            continue
        bd_row = df_bd.loc[bd_idx]

        row_out: Dict[str, object] = {}
        for source, target in BASE_OUTPUT_MAP.items():
            row_out[target] = _clean_cell_value(bd_row.get(source, ""))

        for source, target in ACTUALIZADA_BASE_FIELDS.items():
            row_out[target] = _clean_cell_value(act_row.get(source, ""))

        row_out["Activo"] = "Si"

        nuevo_nivel = _clean_cell_value(act_row.get("nivel", ""))
        nuevo_grado = _clean_cell_value(act_row.get("grado", ""))
        nuevo_grupo = _clean_cell_value(act_row.get("grupo", ""))
        bd_nivel = _clean_cell_value(bd_row.get("nivel", ""))
        bd_grado = _clean_cell_value(bd_row.get("grado", ""))
        bd_grupo = _clean_cell_value(bd_row.get("grupo", ""))

        nivel_changed = _is_changed(nuevo_nivel, bd_nivel, _normalize_text)
        grado_changed = _is_changed(nuevo_grado, bd_grado, _normalize_text)
        grupo_changed = _is_changed(nuevo_grupo, bd_grupo, _normalize_grupo)

        if nivel_changed or grado_changed or grupo_changed:
            row_out["Nuevo Nivel"] = nuevo_nivel
            row_out["Nuevo Grado"] = nuevo_grado
            row_out["Nuevo Grupo"] = nuevo_grupo
        else:
            row_out["Nuevo Nivel"] = ""
            row_out["Nuevo Grado"] = ""
            row_out["Nuevo Grupo"] = ""

        rows.append(row_out)

    df_out = pd.DataFrame(rows, columns=ALUMNOS_COMPARACION_COLUMNS)
    cleaned = df_out.astype(str).apply(lambda col: col.str.strip().replace("nan", ""))
    mask = (cleaned != "").any(axis=1)
    df_out = df_out.loc[mask].reset_index(drop=True)
    if "Fecha de Nacimiento" in df_out.columns:
        df_out["Fecha de Nacimiento"] = df_out["Fecha de Nacimiento"].apply(
            _parse_fecha_excel
        )
    return df_out


def _parse_fecha_excel(value: object):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if abs(numeric - int(numeric)) < 1e-6:
            compact = _parse_compact_date(str(int(numeric)))
            if compact:
                return compact
        parsed = _parse_excel_serial(numeric)
        return parsed or ""
    text = str(value or "").strip()
    if not text:
        return ""
    compact = _parse_compact_date(text)
    if compact:
        return compact
    num = _parse_numeric_string(text)
    if num is not None:
        parsed = _parse_excel_serial(num)
        if parsed:
            return parsed
        return ""
    # Remove time portion if present
    text = re.split(r"[T\s]", text)[0]
    match = re.fullmatch(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", text)
    if match:
        year, month, day = match.groups()
        return date(int(year), int(month), int(day))
    match = re.fullmatch(r"(\d{2})[-/\.](\d{2})[-/\.](\d{4})", text)
    if match:
        day, month, year = match.groups()
        return date(int(year), int(month), int(day))
    return text


def _parse_numeric_string(text: str) -> Optional[float]:
    cleaned = text.strip()
    if not re.fullmatch(r"-?\d+([.,]\d+)?", cleaned):
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_compact_date(text: str) -> Optional[date]:
    digits = re.sub(r"\D", "", text)
    if len(digits) != 8:
        return None
    year_first = int(digits[:4])
    month_first = int(digits[4:6])
    day_first = int(digits[6:8])
    if 1900 <= year_first <= 2099:
        parsed = _safe_date(year_first, month_first, day_first)
        if parsed:
            return parsed
    day_last = int(digits[:2])
    month_last = int(digits[2:4])
    year_last = int(digits[4:8])
    if 1900 <= year_last <= 2099:
        parsed = _safe_date(year_last, month_last, day_last)
        if parsed:
            return parsed
    return None


def _safe_date(year: int, month: int, day: int) -> Optional[date]:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_excel_serial(value: float) -> Optional[date]:
    try:
        if value < 0:
            return None
    except TypeError:
        return None
    try:
        ts = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
    except Exception:
        return None
    if pd.isna(ts):
        return None
    try:
        return ts.date()
    except AttributeError:
        return None
