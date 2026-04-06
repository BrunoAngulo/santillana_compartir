import re
import unicodedata
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter

DEFAULT_SHEET_BD = "Plantilla_BD"
DEFAULT_SHEET_ACTUALIZADA = "Plantilla_Actualizada"
COMPARE_MODE_DNI = "dni"
COMPARE_MODE_APELLIDOS = "apellidos"
COMPARE_MODE_AMBOS = "ambos"
ALUMNOS_CREAR_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "Nombre",
    "Apellido Paterno",
    "Apellido Materno",
    "Sexo",
    "Fecha de Nacimiento",
    "NUIP",
    "Login",
    "Password",
]
ALUMNOS_COMPARACION_COLUMNS = [
    "Nivel",
    "Grado",
    "Grupo",
    "NUI",
    "Id Alumno",
    "Activo",
    "Nombre",
    "Apellido Paterno",
    "Apellido materno",
    "Sexo",
    "Fecha de Nacimiento",
    "Extranjero",
    "NUIP",
    "Login",
    "Password",
    "Nuevo Nivel",
    "Nuevo Grado",
    "Nuevo Grupo",
]

BASE_OUTPUT_MAP = {
    "nivel": "Nivel",
    "grado": "Grado",
    "grupo": "Grupo",
    "nui": "NUI",
    "id_alumno": "Id Alumno",
    "activo": "Activo",
    "nombre": "Nombre",
    "apellido_paterno": "Apellido Paterno",
    "apellido_materno": "Apellido materno",
    "sexo": "Sexo",
    "fecha_nacimiento": "Fecha de Nacimiento",
    "extranjero": "Extranjero",
    "nuip": "NUIP",
    "login": "Login",
    "password": "Password",
}

ACTUALIZADA_OUTPUT_MAP = {
    "nivel": "Nuevo Nivel",
    "grado": "Nuevo Grado",
    "grupo": "Nuevo Grupo",
}

ACTUALIZADA_BASE_FIELDS = {
    "nombre": "Nombre",
    "apellido_paterno": "Apellido Paterno",
    "apellido_materno": "Apellido materno",
    "sexo": "Sexo",
    "fecha_nacimiento": "Fecha de Nacimiento",
    "nuip": "NUIP",
    "login": "Login",
    "password": "Password",
}

HEADER_ALIASES = {
    "nivel": "nivel",
    "grado": "grado",
    "grupo": "grupo",
    "seccion": "grupo",
    "seccion grupo": "grupo",
    "gruposeccion": "grupo",
    "nui": "nui",
    "id alumno": "id_alumno",
    "idalumno": "id_alumno",
    "activo": "activo",
    "nombre": "nombre",
    "apellido paterno": "apellido_paterno",
    "apellido materno": "apellido_materno",
    "apellido materno ": "apellido_materno",
    "sexo": "sexo",
    "fecha de nacimiento": "fecha_nacimiento",
    "fecha nacimiento": "fecha_nacimiento",
    "extranjero": "extranjero",
    "nuip": "nuip",
    "dni": "nuip",
    "login": "login",
    "password": "password",
}

MATCH_TIPO_N1 = "N1_NUIP"
MATCH_TIPO_N2 = "N2_APELLIDOS"
MATCH_TIPO_N3 = "N3_APELLIDOS_INICIAL"
APELLIDO_MIN_SCORE = 0.85
APELLIDO_SHORT_LEN = 4
DEFAULT_BIRTHDATE = date(2000, 1, 1)


def comparar_plantillas(
    excel_path: Path,
    sheet_bd: str = DEFAULT_SHEET_BD,
    sheet_actualizada: str = DEFAULT_SHEET_ACTUALIZADA,
    compare_mode: str = COMPARE_MODE_AMBOS,
) -> Tuple[bytes, Dict[str, int]]:
    result = comparar_plantillas_detalle(
        excel_path=excel_path,
        sheet_bd=sheet_bd,
        sheet_actualizada=sheet_actualizada,
        compare_mode=compare_mode,
    )
    return result["resultado_bytes"], result["summary"]


def comparar_plantillas_detalle(
    excel_path: Path,
    sheet_bd: str = DEFAULT_SHEET_BD,
    sheet_actualizada: str = DEFAULT_SHEET_ACTUALIZADA,
    compare_mode: str = COMPARE_MODE_AMBOS,
) -> Dict[str, object]:
    df_bd = _read_sheet(excel_path, sheet_bd)
    df_act = _read_sheet(excel_path, sheet_actualizada)

    df_bd = _canonicalize_columns(df_bd)
    df_act = _canonicalize_columns(df_act)
    df_bd = _sanitize_student_name_columns(df_bd)
    df_act = _sanitize_student_name_columns(df_act)

    summary = _build_compare_summary(df_bd, df_act, compare_mode=compare_mode)
    comparacion, nuevos, coincidencias, sin_referencia = _build_comparacion_bd(
        df_bd,
        df_act,
        compare_mode=compare_mode,
    )

    output = Path(excel_path).name
    output_bytes = _export_comparacion(comparacion, nuevos)
    actualizacion_bytes = export_alumnos_actualizacion_excel(comparacion)
    alta_bytes = export_alumnos_crear_excel(nuevos)
    summary.update(
        {
            "actualizados_total": len(df_act),
            "base_total": len(df_bd),
            "archivo_base": output,
            "nuevos_total": len(nuevos),
            "inactivados_total": _count_inactivos(comparacion),
            "compare_mode": compare_mode,
        }
    )
    return {
        "resultado_bytes": output_bytes,
        "actualizacion_bytes": actualizacion_bytes,
        "alta_bytes": alta_bytes,
        "summary": summary,
        "coincidencias_rows": coincidencias.to_dict("records"),
        "sin_referencia_rows": sin_referencia.to_dict("records"),
    }


def _read_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")
    with pd.ExcelFile(excel_path, engine="openpyxl") as excel:
        if sheet_name == DEFAULT_SHEET_BD:
            resolved = _resolve_sheet_name_fallback(
                excel.sheet_names,
                sheet_name,
                ["Plantilla edicion masiva", "Plantilla edición masiva"],
            )
        else:
            resolved = _resolve_sheet_name(excel.sheet_names, sheet_name)
        df = pd.read_excel(excel, sheet_name=resolved, dtype=str)
    return df.fillna("")


def _resolve_sheet_name(available: Sequence[str], desired: str) -> str:
    if desired in available:
        return desired
    desired_lower = desired.lower()
    for sheet in available:
        if sheet.lower() == desired_lower:
            return sheet
    desired_norm = _normalize_header(desired)
    for sheet in available:
        if _normalize_header(sheet) == desired_norm:
            return sheet
    available_list = ", ".join(available) if available else "(sin hojas)"
    raise ValueError(
        f"No se encontro la hoja '{desired}'. Hojas disponibles: {available_list}."
    )


def _resolve_sheet_name_fallback(
    available: Sequence[str], desired: str, fallbacks: Sequence[str]
) -> str:
    try:
        return _resolve_sheet_name(available, desired)
    except ValueError:
        for fallback in fallbacks:
            try:
                return _resolve_sheet_name(available, fallback)
            except ValueError:
                continue
    available_list = ", ".join(available) if available else "(sin hojas)"
    raise ValueError(
        f"No se encontro la hoja '{desired}'. Hojas disponibles: {available_list}."
    )


def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    used = set()
    for col in df.columns:
        key = _normalize_header(col)
        canonical = HEADER_ALIASES.get(key)
        if canonical and canonical not in used:
            mapping[col] = canonical
            used.add(canonical)
    return df.rename(columns=mapping)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text)
    return text.strip().lower()


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.replace(" ", "")


def _sanitize_student_name_value(value: object) -> str:
    text = _clean_cell_value(value)
    if not text:
        return ""
    text = re.sub(r"[\'´`\-’]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _sanitize_student_name_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    sanitized = df.copy()
    for column in ("nombre", "apellido_paterno", "apellido_materno"):
        if column not in sanitized.columns:
            continue
        sanitized[column] = sanitized[column].apply(
            _sanitize_student_name_value
        )
    return sanitized


def _char_match_ratio(text_a: str, text_b: str) -> float:
    if not text_a or not text_b:
        return 0.0
    if text_a == text_b:
        return 1.0
    return SequenceMatcher(None, text_a, text_b).ratio()


def _apellido_match_ok(text_a: str, text_b: str, score: float) -> bool:
    max_len = max(len(text_a), len(text_b))
    if max_len <= APELLIDO_SHORT_LEN:
        return score == 1.0
    return score >= APELLIDO_MIN_SCORE


def _normalize_date(value: object) -> str:
    parsed = _parse_fecha_excel(value)
    if isinstance(parsed, date):
        return parsed.strftime("%Y%m%d")
    text = str(parsed or "").strip()
    if not text:
        return ""
    compact = _parse_compact_date(text)
    if compact:
        return compact.strftime("%Y%m%d")
    return re.sub(r"\D", "", text)


def _normalize_nuip(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _first_initial(value: object) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    return text[0]


def _normalize_tokens(value: object) -> List[str]:
    if value is None:
        return []
    text = str(value or "").strip().lower()
    if not text:
        return []
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    return [token for token in text.split(" ") if token]


def _is_name_subset_match(tokens_a: Sequence[str], tokens_b: Sequence[str]) -> bool:
    if not tokens_a or not tokens_b:
        return False
    if tokens_a[0] != tokens_b[0]:
        return False
    set_a = set(tokens_a)
    set_b = set(tokens_b)
    return set_a.issubset(set_b) or set_b.issubset(set_a)


def _score_name_match(value_a: object, value_b: object) -> int:
    tokens_a = _normalize_tokens(value_a)
    tokens_b = _normalize_tokens(value_b)
    if not tokens_a or not tokens_b:
        return 0

    compact_a = "".join(tokens_a)
    compact_b = "".join(tokens_b)
    if compact_a == compact_b:
        return 4
    if _is_name_subset_match(tokens_a, tokens_b):
        return 3
    if tokens_a[0] == tokens_b[0]:
        return 1

    overlap = len(set(tokens_a) & set(tokens_b))
    if overlap > 0:
        return 1
    return 0


def _comment_for_match(match_type: Optional[str]) -> str:
    if match_type == MATCH_TIPO_N1:
        return "Nivel 1: NUIP"
    if match_type == MATCH_TIPO_N2:
        return "Nivel 2: Apellidos"
    if match_type == MATCH_TIPO_N3:
        return "Nivel 3: Apellidos + inicial"
    return "No encontrado"


def _format_match_flag(value: Optional[bool]) -> str:
    if value is True:
        return "Si"
    if value is False:
        return "No"
    return ""


def _build_base_indexes(df_bd: pd.DataFrame) -> Dict[str, Dict[str, List[int]]]:
    indexes = {"nuip": {}, "apellidos": {}, "apellidos_inicial": {}}
    for idx, row in df_bd.iterrows():
        nuip = _normalize_nuip(row.get("nuip"))
        if nuip:
            indexes["nuip"].setdefault(nuip, []).append(idx)
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        ap_mat = _normalize_text(row.get("apellido_materno"))
        if not (ap_pat and ap_mat):
            continue
        key = f"{ap_pat}|{ap_mat}"
        indexes["apellidos"].setdefault(key, []).append(idx)
        inicial = _first_initial(row.get("nombre"))
        if inicial:
            key_inicial = f"{ap_pat}|{ap_mat}|{inicial}"
            indexes["apellidos_inicial"].setdefault(key_inicial, []).append(idx)
    return indexes


def _clasificar_actualizados(
    df_act: pd.DataFrame,
    df_bd: pd.DataFrame,
    indexes: Dict[str, Dict[str, List[int]]],
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    rows = []
    counts = {
        "nivel1": 0,
        "nivel2": 0,
        "nivel3": 0,
        "sin_match": 0,
        "grado_no": 0,
        "grupo_no": 0,
    }

    for _idx, row in df_act.iterrows():
        match_type, match_indices = _match_row(row, indexes)
        enriched = row.copy()
        enriched["Match Tipo"] = match_type or ""
        bd_grado = _collect_matches(df_bd, match_indices, "grado")
        bd_grupo = _collect_matches(df_bd, match_indices, "grupo")
        grado_ok = _match_field(row.get("grado"), df_bd, match_indices, "grado")
        grupo_ok = _match_field(row.get("grupo"), df_bd, match_indices, "grupo")
        enriched["BD Grado"] = bd_grado
        enriched["BD Grupo"] = bd_grupo
        enriched["Grado coincide"] = _format_match_flag(grado_ok)
        enriched["Grupo coincide"] = _format_match_flag(grupo_ok)
        comentario = _comment_for_match(match_type)
        if match_type:
            if grado_ok is False:
                comentario = f"{comentario} | Grado no coincide"
                counts["grado_no"] += 1
            if grupo_ok is False:
                comentario = f"{comentario} | Grupo no coincide"
                counts["grupo_no"] += 1
        enriched["Comentario"] = comentario

        if match_type == MATCH_TIPO_N1:
            counts["nivel1"] += 1
        elif match_type == MATCH_TIPO_N2:
            counts["nivel2"] += 1
        elif match_type == MATCH_TIPO_N3:
            counts["nivel3"] += 1
        else:
            counts["sin_match"] += 1
        rows.append(enriched)

    return pd.DataFrame(rows), counts


def _match_row(
    row: pd.Series, indexes: Dict[str, Dict[str, List[int]]]
) -> Tuple[Optional[str], List[int]]:
    nuip = _normalize_nuip(row.get("nuip"))
    if nuip and nuip in indexes["nuip"]:
        return MATCH_TIPO_N1, indexes["nuip"][nuip]
    ap_pat = _normalize_text(row.get("apellido_paterno"))
    ap_mat = _normalize_text(row.get("apellido_materno"))
    if ap_pat and ap_mat:
        key = f"{ap_pat}|{ap_mat}"
        if key in indexes["apellidos"]:
            return MATCH_TIPO_N2, indexes["apellidos"][key]
        inicial = _first_initial(row.get("nombre"))
        if inicial:
            key_inicial = f"{ap_pat}|{ap_mat}|{inicial}"
            if key_inicial in indexes["apellidos_inicial"]:
                return MATCH_TIPO_N3, indexes["apellidos_inicial"][key_inicial]
    return None, []


def _collect_matches(
    df_bd: pd.DataFrame, indices: Sequence[int], column: str
) -> str:
    if not indices:
        return ""
    values = []
    for idx in indices:
        value = str(df_bd.loc[idx].get(column, "") or "").strip()
        if value and value not in values:
            values.append(value)
    return ", ".join(values)


def _match_field(
    value: object, df_bd: pd.DataFrame, indices: Sequence[int], column: str
) -> Optional[bool]:
    if not indices:
        return None
    normalizer = _normalize_grupo if column == "grupo" else _normalize_text
    current = normalizer(value)
    if not current:
        return None
    for idx in indices:
        bd_value = normalizer(df_bd.loc[idx].get(column))
        if bd_value and bd_value == current:
            return True
    return False


def _export_comparacion(comparacion: pd.DataFrame, nuevos: pd.DataFrame) -> bytes:
    from io import BytesIO

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, "Plantilla edición masiva", comparacion)
        _write_sheet(writer, "Plantilla alta de alumnos", nuevos)
    output.seek(0)
    return output.getvalue()


def export_alumnos_actualizacion_excel(comparacion: pd.DataFrame) -> bytes:
    from io import BytesIO

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, "Plantilla ediciÃ³n masiva", comparacion)
    output.seek(0)
    return output.getvalue()


def export_alumnos_crear_excel(nuevos: pd.DataFrame) -> bytes:
    from io import BytesIO

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, "Plantilla alta de alumnos", nuevos)
    output.seek(0)
    return output.getvalue()


def _write_sheet(writer: pd.ExcelWriter, name: str, df: pd.DataFrame) -> None:
    df = df.copy()
    df.to_excel(writer, index=False, sheet_name=name)
    ws = writer.book[name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _apply_date_format(ws, df, "Fecha de Nacimiento")
    for idx, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).head(200).tolist()
        max_len = max([len(str(col))] + [len(val) for val in sample])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)


def _apply_date_format(ws, df: pd.DataFrame, column_name: str) -> None:
    if column_name not in df.columns:
        return
    col_idx = list(df.columns).index(column_name) + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            cell.number_format = "dd/mm/yyyy"


def _build_alumnos_crear(nuevos: pd.DataFrame) -> pd.DataFrame:
    if nuevos.empty:
        return pd.DataFrame(columns=ALUMNOS_CREAR_COLUMNS)
    rename_map = {
        "nivel": "Nivel",
        "grado": "Grado",
        "grupo": "Grupo",
        "nombre": "Nombre",
        "apellido_paterno": "Apellido Paterno",
        "apellido_materno": "Apellido Materno",
        "sexo": "Sexo",
        "fecha_nacimiento": "Fecha de Nacimiento",
        "nuip": "NUIP",
        "login": "Login",
        "password": "Password",
    }
    df = nuevos.rename(columns=rename_map)
    for col in ALUMNOS_CREAR_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df.loc[:, ALUMNOS_CREAR_COLUMNS].copy()
    if "Fecha de Nacimiento" in df.columns:
        df["Fecha de Nacimiento"] = df["Fecha de Nacimiento"].apply(
            _parse_fecha_excel_with_default
        )
    cleaned = df.astype(str).apply(lambda col: col.str.strip().replace("nan", ""))
    mask = (cleaned != "").any(axis=1)
    return df.loc[mask].reset_index(drop=True)


def _clean_cell_value(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _merge_non_empty_compare_fields(
    row_out: Dict[str, object],
    act_row: pd.Series,
    fields_map: Dict[str, str],
) -> None:
    for source, target in fields_map.items():
        current_value = _clean_cell_value(row_out.get(target, ""))
        updated_value = _clean_cell_value(act_row.get(source, ""))
        row_out[target] = updated_value or current_value


def _compose_display_name(row: pd.Series) -> str:
    parts = [
        _clean_cell_value(row.get("nombre", "")),
        _clean_cell_value(row.get("apellido_paterno", "")),
        _clean_cell_value(row.get("apellido_materno", "")),
    ]
    return " ".join(part for part in parts if part).strip()


def _resolve_final_login(act_row: pd.Series, bd_row: Optional[pd.Series] = None) -> str:
    act_login = _clean_cell_value(act_row.get("login", ""))
    if act_login:
        return act_login
    if bd_row is not None:
        return _clean_cell_value(bd_row.get("login", ""))
    return ""


def _describe_match_reference(act_row: pd.Series, bd_row: pd.Series) -> str:
    act_nuip = _normalize_nuip(act_row.get("nuip"))
    bd_nuip = _normalize_nuip(bd_row.get("nuip"))
    if act_nuip and bd_nuip and act_nuip == bd_nuip:
        return "NUIP"

    act_nombre = _normalize_text(act_row.get("nombre"))
    bd_nombre = _normalize_text(bd_row.get("nombre"))
    act_ap_pat = _normalize_text(act_row.get("apellido_paterno"))
    bd_ap_pat = _normalize_text(bd_row.get("apellido_paterno"))
    act_ap_mat = _normalize_text(act_row.get("apellido_materno"))
    bd_ap_mat = _normalize_text(bd_row.get("apellido_materno"))

    if act_nombre and act_ap_pat and act_ap_mat:
        if (
            act_nombre == bd_nombre
            and act_ap_pat == bd_ap_pat
            and act_ap_mat == bd_ap_mat
        ):
            return "Nombre + apellidos"
    if act_nombre and act_ap_pat and act_nombre == bd_nombre and act_ap_pat == bd_ap_pat:
        return "Nombre + apellido paterno"
    if act_ap_pat and act_ap_mat and act_ap_pat == bd_ap_pat and act_ap_mat == bd_ap_mat:
        return "Apellidos"
    return "Referencia"


def _build_match_preview_row(act_row: pd.Series, bd_row: pd.Series) -> Dict[str, str]:
    return {
        "Alumno Plantilla_Actualizada": _compose_display_name(act_row),
        "NUIP Plantilla_Actualizada": _clean_cell_value(act_row.get("nuip", "")),
        "Login Plantilla_Actualizada": _clean_cell_value(act_row.get("login", "")),
        "Login final": _resolve_final_login(act_row, bd_row),
        "Alumno BD": _compose_display_name(bd_row),
        "NUIP BD": _clean_cell_value(bd_row.get("nuip", "")),
        "Login BD": _clean_cell_value(bd_row.get("login", "")),
        "Referencia por": _describe_match_reference(act_row, bd_row),
        "Nivel Actualizada": _clean_cell_value(act_row.get("nivel", "")),
        "Grado Actualizada": _clean_cell_value(act_row.get("grado", "")),
        "Grupo Actualizada": _clean_cell_value(act_row.get("grupo", "")),
        "Nivel BD": _clean_cell_value(bd_row.get("nivel", "")),
        "Grado BD": _clean_cell_value(bd_row.get("grado", "")),
        "Grupo BD": _clean_cell_value(bd_row.get("grupo", "")),
    }


def _build_unmatched_preview_row(act_row: pd.Series) -> Dict[str, str]:
    return {
        "Alumno Plantilla_Actualizada": _compose_display_name(act_row),
        "NUIP": _clean_cell_value(act_row.get("nuip", "")),
        "Login Plantilla_Actualizada": _clean_cell_value(act_row.get("login", "")),
        "Login final": _resolve_final_login(act_row),
        "Nivel": _clean_cell_value(act_row.get("nivel", "")),
        "Grado": _clean_cell_value(act_row.get("grado", "")),
        "Grupo": _clean_cell_value(act_row.get("grupo", "")),
    }


def _pick_row_value(row: pd.Series, aliases: Sequence[str]) -> object:
    normalized_aliases = {_normalize_header(alias) for alias in aliases if alias}
    for key in row.index:
        if _normalize_header(key) not in normalized_aliases:
            continue
        value = row.get(key, "")
        if isinstance(value, pd.Series):
            for item in value.tolist():
                cleaned = _clean_cell_value(item)
                if cleaned:
                    return cleaned
            continue
        cleaned = _clean_cell_value(value)
        if cleaned:
            return cleaned
    return ""


def _normalize_grupo(value: object) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"^(grupo|seccion)+", "", text)
    tail_match = re.search(r"([a-z]+|\d+)$", text)
    if tail_match:
        return tail_match.group(1)
    digits = re.findall(r"\d+", text)
    if digits:
        return digits[-1]
    letters = re.findall(r"[a-z]", text)
    if letters:
        return letters[-1]
    return text


def _build_nuip_index(df: pd.DataFrame) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        nuip = _normalize_nuip(row.get("nuip"))
        if not nuip:
            continue
        index.setdefault(nuip, []).append(int(idx))
    return index


def _resolve_nuip_match(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    index: Dict[str, List[int]],
    used_indices: Optional[set] = None,
) -> Optional[int]:
    nuip_norm = _normalize_nuip(act_row.get("nuip"))
    if not nuip_norm:
        return None

    candidates = _filter_unused_indices(index.get(nuip_norm) or [], used_indices)
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    return _pick_best_match(act_row, df_bd, candidates)


def _build_nombre_ap_pat_index(df: pd.DataFrame) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        nombre = _normalize_text(row.get("nombre"))
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        if not (nombre and ap_pat):
            continue
        key = f"{nombre}|{ap_pat}"
        index.setdefault(key, []).append(idx)
    return index


def _build_identidad_index(df: pd.DataFrame) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        nombre = _normalize_text(row.get("nombre"))
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        ap_mat = _normalize_text(row.get("apellido_materno"))
        if not (nombre and ap_pat and ap_mat):
            continue
        key = f"{nombre}|{ap_pat}|{ap_mat}"
        index.setdefault(key, []).append(idx)
    return index


def _filter_unused_indices(
    indices: Sequence[int], used_indices: Optional[set]
) -> List[int]:
    if not used_indices:
        return [int(idx) for idx in indices]
    return [int(idx) for idx in indices if int(idx) not in used_indices]


def _resolve_nombre_ap_pat_match(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    index: Dict[str, List[int]],
    used_indices: Optional[set] = None,
) -> Optional[int]:
    nombre = _normalize_text(act_row.get("nombre"))
    ap_pat = _normalize_text(act_row.get("apellido_paterno"))
    if not (nombre and ap_pat):
        return None

    key = f"{nombre}|{ap_pat}"
    candidates = _filter_unused_indices(index.get(key) or [], used_indices)
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    nuip_norm = _normalize_nuip(act_row.get("nuip"))
    if nuip_norm:
        for idx in candidates:
            bd_nuip = _normalize_nuip(df_bd.loc[idx].get("nuip"))
            if bd_nuip and bd_nuip == nuip_norm:
                return idx

    return _pick_best_match(act_row, df_bd, candidates)


def _resolve_identidad_match(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    index: Dict[str, List[int]],
    used_indices: Optional[set] = None,
) -> Optional[int]:
    nombre = _normalize_text(act_row.get("nombre"))
    ap_pat = _normalize_text(act_row.get("apellido_paterno"))
    ap_mat = _normalize_text(act_row.get("apellido_materno"))
    if not (nombre and ap_pat and ap_mat):
        return None

    key = f"{nombre}|{ap_pat}|{ap_mat}"
    candidates = _filter_unused_indices(index.get(key) or [], used_indices)
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    nuip_norm = _normalize_nuip(act_row.get("nuip"))
    if nuip_norm:
        for idx in candidates:
            bd_nuip = _normalize_nuip(df_bd.loc[idx].get("nuip"))
            if bd_nuip and bd_nuip == nuip_norm:
                return idx

    return _pick_best_match(act_row, df_bd, candidates)


def _build_apellidos_index(df: pd.DataFrame) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    for idx, row in df.iterrows():
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        ap_mat = _normalize_text(row.get("apellido_materno"))
        if not (ap_pat and ap_mat):
            continue
        key = f"{ap_pat}|{ap_mat}"
        index.setdefault(key, []).append(idx)
    return index


def _build_apellidos_cache(df: pd.DataFrame) -> List[Tuple[int, str, str]]:
    cache: List[Tuple[int, str, str]] = []
    for idx, row in df.iterrows():
        ap_pat = _normalize_text(row.get("apellido_paterno"))
        ap_mat = _normalize_text(row.get("apellido_materno"))
        if not (ap_pat and ap_mat):
            continue
        cache.append((idx, ap_pat, ap_mat))
    return cache


def _collect_apellidos_candidates(
    act_row: pd.Series,
    cache: Sequence[Tuple[int, str, str]],
    used_indices: Optional[set] = None,
) -> List[int]:
    ap_pat = _normalize_text(act_row.get("apellido_paterno"))
    ap_mat = _normalize_text(act_row.get("apellido_materno"))
    if not (ap_pat and ap_mat):
        return []

    best_score = -1.0
    candidates: List[int] = []
    for idx, bd_ap_pat, bd_ap_mat in cache:
        score_pat = _char_match_ratio(ap_pat, bd_ap_pat)
        if not _apellido_match_ok(ap_pat, bd_ap_pat, score_pat):
            continue
        score_mat = _char_match_ratio(ap_mat, bd_ap_mat)
        if not _apellido_match_ok(ap_mat, bd_ap_mat, score_mat):
            continue
        score = (score_pat + score_mat) / 2
        if score > best_score + 1e-6:
            best_score = score
            candidates = [int(idx)]
        elif abs(score - best_score) <= 1e-6:
            candidates.append(int(idx))

    return _filter_unused_indices(candidates, used_indices)


def _filter_candidates_by_nuip(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    candidates: Sequence[int],
) -> List[int]:
    nuip_norm = _normalize_nuip(act_row.get("nuip"))
    if not nuip_norm:
        return []

    matched: List[int] = []
    for idx in candidates:
        bd_nuip = _normalize_nuip(df_bd.loc[idx].get("nuip"))
        if bd_nuip and bd_nuip == nuip_norm:
            matched.append(int(idx))
    return matched


def _resolve_apellidos_match(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    cache: Sequence[Tuple[int, str, str]],
    used_indices: Optional[set] = None,
) -> Optional[int]:
    candidates = _collect_apellidos_candidates(
        act_row,
        cache,
        used_indices=used_indices,
    )
    if not candidates:
        return None

    nuip_candidates = _filter_candidates_by_nuip(act_row, df_bd, candidates)
    if len(nuip_candidates) == 1:
        return nuip_candidates[0]
    if len(nuip_candidates) > 1:
        best_by_identity = _pick_best_match(act_row, df_bd, nuip_candidates)
        if best_by_identity is not None:
            return best_by_identity

    # Si hay apellidos repetidos (ej. hermanos), usar identidad de persona
    # como desempate solo despues de intentar match por NUIP.
    best_by_identity = _pick_best_match(act_row, df_bd, candidates)
    if best_by_identity is not None:
        return best_by_identity

    return None


def _resolve_apellidos_then_nuip_match(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    nuip_index: Dict[str, List[int]],
    nombre_ap_pat_index: Dict[str, List[int]],
    identidad_index: Dict[str, List[int]],
    apellidos_cache: Sequence[Tuple[int, str, str]],
    used_indices: Optional[set] = None,
) -> Optional[int]:
    bd_idx = _resolve_identidad_match(
        act_row,
        df_bd,
        identidad_index,
        used_indices=used_indices,
    )
    if bd_idx is not None:
        return bd_idx

    bd_idx = _resolve_nombre_ap_pat_match(
        act_row,
        df_bd,
        nombre_ap_pat_index,
        used_indices=used_indices,
    )
    if bd_idx is not None:
        return bd_idx

    candidates = _collect_apellidos_candidates(
        act_row,
        apellidos_cache,
        used_indices=used_indices,
    )
    if candidates:
        nuip_candidates = _filter_candidates_by_nuip(act_row, df_bd, candidates)
        if len(nuip_candidates) == 1:
            return nuip_candidates[0]
        if len(nuip_candidates) > 1:
            best_by_identity = _pick_best_match(act_row, df_bd, nuip_candidates)
            if best_by_identity is not None:
                return best_by_identity

        best_by_identity = _pick_best_match(act_row, df_bd, candidates)
        if best_by_identity is not None:
            return best_by_identity

    return _resolve_nuip_match(
        act_row,
        df_bd,
        nuip_index,
        used_indices=used_indices,
    )


def _resolve_match_by_mode(
    act_row: pd.Series,
    df_bd: pd.DataFrame,
    nuip_index: Dict[str, List[int]],
    nombre_ap_pat_index: Dict[str, List[int]],
    identidad_index: Dict[str, List[int]],
    apellidos_cache: Sequence[Tuple[int, str, str]],
    used_indices: Optional[set],
    compare_mode: str,
) -> Optional[int]:
    if compare_mode == COMPARE_MODE_DNI:
        bd_idx = _resolve_nuip_match(
            act_row,
            df_bd,
            nuip_index,
            used_indices=used_indices,
        )
        if bd_idx is not None:
            return bd_idx
        return _resolve_identidad_match(
            act_row,
            df_bd,
            identidad_index,
            used_indices=used_indices,
        )
        
    if compare_mode == COMPARE_MODE_APELLIDOS:
        bd_idx = _resolve_identidad_match(
            act_row,
            df_bd,
            identidad_index,
            used_indices=used_indices,
        )
        if bd_idx is not None:
            return bd_idx
        bd_idx = _resolve_nombre_ap_pat_match(
            act_row,
            df_bd,
            nombre_ap_pat_index,
            used_indices=used_indices,
        )
        if bd_idx is not None:
            return bd_idx
        return _resolve_apellidos_match(
            act_row,
            df_bd,
            apellidos_cache,
            used_indices=used_indices,
        )
    if compare_mode == COMPARE_MODE_AMBOS:
        return _resolve_apellidos_then_nuip_match(
            act_row,
            df_bd,
            nuip_index,
            nombre_ap_pat_index,
            identidad_index,
            apellidos_cache,
            used_indices=used_indices,
        )
    raise ValueError(f"Modo de comparacion invalido: {compare_mode}")


def _build_compare_summary(
    df_bd: pd.DataFrame,
    df_act: pd.DataFrame,
    compare_mode: str,
) -> Dict[str, int]:
    bd_nuips = set()
    for _idx, row in df_bd.iterrows():
        nuip = _normalize_nuip(row.get("nuip"))
        if nuip:
            bd_nuips.add(nuip)

    act_nuips = set()
    for _idx, row in df_act.iterrows():
        nuip = _normalize_nuip(row.get("nuip"))
        if nuip:
            act_nuips.add(nuip)

    matched_total = 0
    if not df_bd.empty and not df_act.empty:
        nuip_index = _build_nuip_index(df_bd)
        nombre_ap_pat_index = _build_nombre_ap_pat_index(df_bd)
        identidad_index = _build_identidad_index(df_bd)
        apellidos_cache = _build_apellidos_cache(df_bd)
        used_indices = set()
        for _idx, act_row in df_act.iterrows():
            bd_idx = _resolve_match_by_mode(
                act_row,
                df_bd,
                nuip_index,
                nombre_ap_pat_index,
                identidad_index,
                apellidos_cache,
                used_indices=used_indices,
                compare_mode=compare_mode,
            )
            if bd_idx is None:
                continue
            used_indices.add(int(bd_idx))
            matched_total += 1

    return {
        "nuip_bd_total": len(bd_nuips),
        "nuip_actualizada_total": len(act_nuips),
        "match_total": int(matched_total),
        "nuip_sin_bd": len(act_nuips - bd_nuips),
        "nuip_sin_actualizada": len(bd_nuips - act_nuips),
    }


def _is_changed(new_value: object, old_value: object, normalizer) -> bool:
    new_norm = normalizer(new_value)
    if not new_norm:
        return False
    old_norm = normalizer(old_value)
    if not old_norm:
        return True
    return new_norm != old_norm


def _pick_best_match(
    base_row: pd.Series, df_act: pd.DataFrame, indices: Sequence[int]
) -> Optional[int]:
    if not indices:
        return None

    target_nombre = _normalize_text(base_row.get("nombre"))
    target_fecha = _normalize_date(base_row.get("fecha_nacimiento"))
    target_sexo = _normalize_text(base_row.get("sexo"))

    best_score = -1
    best_indices: List[int] = []
    for idx in indices:
        candidate = df_act.loc[idx]
        score = 0
        if target_nombre:
            score += _score_name_match(base_row.get("nombre"), candidate.get("nombre"))
        if target_fecha and _normalize_date(candidate.get("fecha_nacimiento")) == target_fecha:
            score += 3
        if target_sexo and _normalize_text(candidate.get("sexo")) == target_sexo:
            score += 1
        if score > best_score:
            best_score = score
            best_indices = [int(idx)]
        elif score == best_score:
            best_indices.append(int(idx))

    # Evita empates ambiguos y matches sin señales de identidad.
    if best_score < 3:
        return None
    if len(best_indices) != 1:
        return None
    return best_indices[0]


def _build_comparacion_bd(
    df_bd: pd.DataFrame,
    df_act: pd.DataFrame,
    compare_mode: str = COMPARE_MODE_AMBOS,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df_bd.empty and df_act.empty:
        return (
            pd.DataFrame(columns=ALUMNOS_COMPARACION_COLUMNS),
            pd.DataFrame(columns=ALUMNOS_CREAR_COLUMNS),
            pd.DataFrame(),
            pd.DataFrame(),
        )

    nuip_index = _build_nuip_index(df_bd) if not df_bd.empty else {}
    nombre_ap_pat_index = _build_nombre_ap_pat_index(df_bd) if not df_bd.empty else {}
    identidad_index = _build_identidad_index(df_bd) if not df_bd.empty else {}
    apellidos_cache = _build_apellidos_cache(df_bd) if not df_bd.empty else []
    rows: List[Dict[str, object]] = []
    nuevos_rows: List[pd.Series] = []
    coincidencias_rows: List[Dict[str, str]] = []
    sin_referencia_rows: List[Dict[str, str]] = []
    bd_matched_indices = set()
    bd_protected_indices = set()

    for _idx, act_row in df_act.iterrows():
        bd_idx: Optional[int] = None
        nuip_norm = ""
        if not df_bd.empty:
            nuip_norm = _normalize_nuip(act_row.get("nuip"))
            bd_idx = _resolve_match_by_mode(
                act_row,
                df_bd,
                nuip_index,
                nombre_ap_pat_index,
                identidad_index,
                apellidos_cache,
                used_indices=bd_matched_indices,
                compare_mode=compare_mode,
            )
        if bd_idx is None:
            protected_candidates: List[int] = []
            if compare_mode in {COMPARE_MODE_DNI, COMPARE_MODE_AMBOS} and nuip_norm:
                protected_candidates.extend(
                    _filter_unused_indices(
                        nuip_index.get(nuip_norm) or [],
                        bd_matched_indices,
                    )
                )
            if compare_mode in {COMPARE_MODE_APELLIDOS, COMPARE_MODE_AMBOS}:
                protected_candidates.extend(
                    _collect_apellidos_candidates(
                        act_row,
                        apellidos_cache,
                        used_indices=bd_matched_indices,
                    )
                )
            if protected_candidates:
                bd_protected_indices.update(
                    int(idx) for idx in protected_candidates if idx is not None
                )
            nuevos_rows.append(act_row)
            sin_referencia_rows.append(_build_unmatched_preview_row(act_row))
            continue
        bd_matched_indices.add(int(bd_idx))
        bd_row = df_bd.loc[bd_idx]
        coincidencias_rows.append(_build_match_preview_row(act_row, bd_row))

        row_out: Dict[str, object] = {}
        for source, target in BASE_OUTPUT_MAP.items():
            row_out[target] = _clean_cell_value(bd_row.get(source, ""))

        row_out["Activo"] = "Si"
        _merge_non_empty_compare_fields(
            row_out,
            act_row,
            {
                "login": "Login",
                "password": "Password",
            },
        )

        nuevo_nivel = _clean_cell_value(
            _pick_row_value(act_row, ("nivel", "Nivel", "Nuevo Nivel"))
        )
        nuevo_grado = _clean_cell_value(
            _pick_row_value(act_row, ("grado", "Grado", "Nuevo Grado"))
        )
        nuevo_grupo = _clean_cell_value(
            _pick_row_value(
                act_row,
                ("grupo", "Grupo", "Seccion", "Sección", "Nuevo Grupo", "Nueva Seccion", "Nueva Sección"),
            )
        )

        nivel_actual = _clean_cell_value(row_out.get("Nivel", ""))
        grado_actual = _clean_cell_value(row_out.get("Grado", ""))
        grupo_actual = _clean_cell_value(row_out.get("Grupo", ""))

        nivel_changed = _is_changed(nuevo_nivel, nivel_actual, _normalize_text)
        grado_changed = _is_changed(nuevo_grado, grado_actual, _normalize_text)
        grupo_changed = _is_changed(nuevo_grupo, grupo_actual, _normalize_grupo)
        ubicacion_changed = nivel_changed or grado_changed or grupo_changed

        if ubicacion_changed:
            row_out["Nuevo Nivel"] = nuevo_nivel or nivel_actual
            row_out["Nuevo Grado"] = nuevo_grado or grado_actual
            row_out["Nuevo Grupo"] = nuevo_grupo or grupo_actual
        else:
            row_out["Nuevo Nivel"] = ""
            row_out["Nuevo Grado"] = ""
            row_out["Nuevo Grupo"] = ""

        rows.append(row_out)

    # Todo alumno existente en BD que no aparezca en la plantilla actualizada
    # se envía como actualización de estado inactivo.
    if not df_bd.empty:
        for bd_idx, bd_row in df_bd.iterrows():
            if int(bd_idx) in bd_matched_indices:
                continue
            if int(bd_idx) in bd_protected_indices:
                continue
            row_out: Dict[str, object] = {}
            for source, target in BASE_OUTPUT_MAP.items():
                row_out[target] = _clean_cell_value(bd_row.get(source, ""))

            # Mantener datos base; solo cambia el estado.
            for source, target in ACTUALIZADA_BASE_FIELDS.items():
                row_out[target] = _clean_cell_value(bd_row.get(source, ""))

            row_out["Activo"] = "No"
            row_out["Nuevo Nivel"] = ""
            row_out["Nuevo Grado"] = ""
            row_out["Nuevo Grupo"] = ""
            rows.append(row_out)

    df_out = pd.DataFrame(rows, columns=ALUMNOS_COMPARACION_COLUMNS)
    cleaned = df_out.astype(str).apply(lambda col: col.str.strip().replace("nan", ""))
    mask = (cleaned != "").any(axis=1)
    df_out = df_out.loc[mask].reset_index(drop=True)
    if "Fecha de Nacimiento" in df_out.columns:
        df_out["Fecha de Nacimiento"] = df_out["Fecha de Nacimiento"].apply(
            _parse_fecha_excel_with_default
        )
    nuevos_df = _build_alumnos_crear(pd.DataFrame(nuevos_rows))
    coincidencias_df = pd.DataFrame(coincidencias_rows)
    sin_referencia_df = pd.DataFrame(sin_referencia_rows)
    return df_out, nuevos_df, coincidencias_df, sin_referencia_df


def _count_inactivos(comparacion: pd.DataFrame) -> int:
    if comparacion.empty or "Activo" not in comparacion.columns:
        return 0
    normalized = comparacion["Activo"].astype(str).apply(_normalize_text)
    return int((normalized == "no").sum())


def _parse_fecha_excel(value: object):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if abs(numeric - int(numeric)) < 1e-6:
            compact = _parse_compact_date(str(int(numeric)))
            if compact:
                return compact
        parsed = _parse_excel_serial(numeric)
        return parsed or ""
    text = str(value or "").strip()
    if not text:
        return ""
    compact = _parse_compact_date(text)
    if compact:
        return compact
    num = _parse_numeric_string(text)
    if num is not None:
        parsed = _parse_excel_serial(num)
        if parsed:
            return parsed
        return ""
    # Remove time portion if present
    text = re.split(r"[T\s]", text)[0]
    match = re.fullmatch(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", text)
    if match:
        year, month, day = match.groups()
        return date(int(year), int(month), int(day))
    match = re.fullmatch(r"(\d{2})[-/\.](\d{2})[-/\.](\d{4})", text)
    if match:
        day, month, year = match.groups()
        return date(int(year), int(month), int(day))
    return text


def _parse_fecha_excel_with_default(value: object):
    parsed = _parse_fecha_excel(value)
    if parsed is None or parsed == "":
        return DEFAULT_BIRTHDATE
    return parsed


def _parse_numeric_string(text: str) -> Optional[float]:
    cleaned = text.strip()
    if not re.fullmatch(r"-?\d+([.,]\d+)?", cleaned):
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_compact_date(text: str) -> Optional[date]:
    digits = re.sub(r"\D", "", text)
    if len(digits) != 8:
        return None
    year_first = int(digits[:4])
    month_first = int(digits[4:6])
    day_first = int(digits[6:8])
    if 1900 <= year_first <= 2099:
        parsed = _safe_date(year_first, month_first, day_first)
        if parsed:
            return parsed
    day_last = int(digits[:2])
    month_last = int(digits[2:4])
    year_last = int(digits[4:8])
    if 1900 <= year_last <= 2099:
        parsed = _safe_date(year_last, month_last, day_last)
        if parsed:
            return parsed
    return None


def _safe_date(year: int, month: int, day: int) -> Optional[date]:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_excel_serial(value: float) -> Optional[date]:
    try:
        if value < 0:
            return None
    except TypeError:
        return None
    try:
        ts = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
    except Exception:
        return None
    if pd.isna(ts):
        return None
    try:
        return ts.date()
    except AttributeError:
        return None
