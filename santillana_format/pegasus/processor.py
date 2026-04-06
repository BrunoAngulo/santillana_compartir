import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Constantes por defecto
CODE_COLUMN_NAME = "CRM"
SHEET_NAME = "Export"
OUTPUT_SHEET_NAME = "Plantilla alta de clases"
OUTPUT_FILENAME = "PlantillaClases.xlsx"
EXPECTED_HEADERS: Set[str] = {
    "CRM",
    "Institucion",
    "Nivel Educativo",
    "Grado",
    "Asignatura Producto",
    "Producto",
    "Plataforma",
    "Razon Estado",
}

# Reglas de materias para configuracion Compartir.
MATERIA_KEY_ALIASES = {
    "CIENCIA Y TECNOLOGIA - PERSONAL SOCIAL": "CIENCIAS INTEGRADAS",
    "CIENCIA Y TECNOLOGIA PERSONAL SOCIAL": "CIENCIAS INTEGRADAS",
    "CIENCIA TECNOLOGIA - PERSONAL SOCIAL": "CIENCIAS INTEGRADAS",
    "CIENCIA TECNOLOGIA PERSONAL SOCIAL": "CIENCIAS INTEGRADAS",
    "INFORMATICA": "TECNOLOGIA",
    "MATEMATICA": "MATEMATICAS",
    "DESARROLLO PERSONAL": "DPCC",
    "CIUDADANIA Y CIVICA": "DPCC",
    "CUIDADANIA Y CIVICA": "DPCC",
    "DESARROLLO PERSONAL CIUDADANIA Y CIVICA": "DPCC",
    "DESARROLLO PERSONAL CUIDADANIA Y CIVICA": "DPCC",
    "DPCC": "DPCC",
    "BIOLOGIA": "BIOLOGIA-FISICA-QUIMICA",
    "FISICA": "BIOLOGIA-FISICA-QUIMICA",
    "QUIMICA": "BIOLOGIA-FISICA-QUIMICA",
    "BIOLOGIA-FISICA-QUIMICA": "BIOLOGIA-FISICA-QUIMICA",
}

MATERIAS_COMPARTIR = {
    "Inicial": {"PREESCOLAR", "INGLES", "TECNOLOGIA"},
    "Primaria": {
        "MATEMATICAS",
        "COMUNICACION",
        "PERSONAL SOCIAL",
        "CIENCIAS SOCIALES",
        "CIENCIA Y TECNOLOGIA",
        "CIENCIAS INTEGRADAS",
        "RAZONAMIENTO MATEMATICO",
        "RAZONAMIENTO VERBAL",
        "RELIGION",
        "INGLES",
        "TECNOLOGIA",
    },
    "Secundaria": {
        "MATEMATICAS",
        "COMUNICACION",
        "CIENCIAS SOCIALES",
        "CIENCIA Y TECNOLOGIA",
        "RAZONAMIENTO MATEMATICO",
        "RAZONAMIENTO VERBAL",
        "RELIGION",
        "INGLES",
        "TECNOLOGIA",
        "DPCC",
        "BIOLOGIA-FISICA-QUIMICA",
    },
}


def _normalize_grupos(grupos: Optional[Sequence[str]]) -> List[str]:
    if not grupos:
        return ["A"]
    if isinstance(grupos, str):
        grupos_iter = re.split(r"[\s,]+", grupos.strip())
    else:
        grupos_iter = grupos
    letras: List[str] = []
    invalid: List[str] = []
    for grupo in grupos_iter:
        if grupo is None:
            continue
        upper = str(grupo).strip().upper()
        if not upper:
            continue
        if len(upper) != 1 or not upper.isalpha():
            invalid.append(str(grupo))
            continue
        if upper not in letras:
            letras.append(upper)
    if invalid:
        raise ValueError(f"Grupo invalido: {', '.join(invalid)}")
    return letras or ["A"]


def _apply_grupo(
    nombre_clase: str,
    grupo: str,
    grado_num: int,
    nivel_codigo: str,
    multi_grupo: bool,
) -> str:
    if grupo == "A":
        return nombre_clase
    if grado_num and nivel_codigo:
        suffix = f"{grado_num}{nivel_codigo}A"
        if nombre_clase.endswith(suffix):
            return f"{nombre_clase[:-len(suffix)]}{grado_num}{nivel_codigo}{grupo}"
    if multi_grupo:
        return f"{nombre_clase} {grupo}"
    return nombre_clase


def _normalize_key(valor: str) -> str:
    """Normaliza texto: sin tildes, mayúsculas y sin espacios extremos."""
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFD", valor)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return texto.strip().upper()


def _canonical_materia_key(valor: str) -> str:
    limpio = re.sub(r"\s*\([^)]*\)", "", valor or "").strip()
    key = _normalize_key(limpio)
    if "INGLES" in key:
        return "INGLES"
    return MATERIA_KEY_ALIASES.get(key, key)


def _materia_permitida(nivel_legible: str, materia_key: str, grado_num: int) -> bool:
    permitidas = MATERIAS_COMPARTIR.get(nivel_legible, set())
    if materia_key not in permitidas:
        return False
    if materia_key == "CIENCIAS INTEGRADAS" and not (
        nivel_legible == "Primaria" and grado_num in {1, 2}
    ):
        return False
    if materia_key == "CIENCIAS SOCIALES" and nivel_legible == "Primaria" and grado_num != 6:
        return False
    return True


def _detectar_fila_encabezado(
    df_raw: pd.DataFrame, columna_codigo: str
) -> Optional[int]:
    """
    Identifica la fila que contiene la mayoría de encabezados conocidos.
    Devuelve el índice de fila o None si no se detecta.
    """
    candidatos = {columna_codigo.lower()} | {col.lower() for col in EXPECTED_HEADERS}
    mejor_fila, mejores_aciertos = None, 0

    for idx, row in df_raw.head(30).iterrows():
        valores = {str(val).strip().lower() for val in row if pd.notna(val)}
        aciertos = len(valores & candidatos)
        if aciertos > mejores_aciertos:
            mejor_fila, mejores_aciertos = idx, aciertos

    return mejor_fila if mejores_aciertos >= 3 else None


def _to_bytes(excel_input) -> bytes:
    """Acepta Path, bytes o buffer; devuelve bytes."""
    if isinstance(excel_input, bytes):
        return excel_input
    if hasattr(excel_input, "read"):
        return excel_input.read()
    return Path(excel_input).read_bytes()


def cargar_excel(
    excel_input,
    hoja: str = SHEET_NAME,
    columna_codigo: str = CODE_COLUMN_NAME,
) -> pd.DataFrame:
    """
    Carga la hoja solicitada; si es necesario, detecta la fila de encabezados.
    excel_input puede ser bytes, Path o buffer.
    """
    raw_bytes = _to_bytes(excel_input)

    def read_bytes(header=None) -> pd.DataFrame:
        return pd.read_excel(
            BytesIO(raw_bytes),
            sheet_name=hoja,
            dtype=str,
            engine="openpyxl",
            header=header,
        )

    try:
        df = read_bytes()
    except ValueError as exc:
        raise ValueError(
            f"No se encontró la hoja '{hoja}' en el archivo. Error: {exc}"
        ) from exc
    except Exception as exc:  # pragma: no cover - defensa general
        raise RuntimeError(f"No se pudo leer el Excel: {exc}") from exc

    encabezados_ok = (
        columna_codigo in df.columns or EXPECTED_HEADERS.issubset(set(df.columns))
    )
    if encabezados_ok:
        return df

    try:
        df_raw = read_bytes(header=None)
    except Exception as exc:  # pragma: no cover - defensa general
        raise RuntimeError(
            f"No se pudo leer el Excel para detectar encabezados: {exc}"
        ) from exc

    fila_encabezado = _detectar_fila_encabezado(df_raw, columna_codigo)
    if fila_encabezado is None:
        return df

    df = read_bytes(header=fila_encabezado)
    return df


def filtrar_codigo(df: pd.DataFrame, codigo: str, columna_codigo: str) -> pd.DataFrame:
    """Filtra las filas por el código exacto."""
    if columna_codigo not in df.columns:
        columnas = ", ".join(df.columns)
        raise KeyError(
            f"La columna '{columna_codigo}' no existe. Columnas disponibles: {columnas}"
        )
    codigo = codigo.strip()
    filtro = df[columna_codigo].astype(str).str.strip() == codigo
    return df.loc[filtro].copy()


def _mapear_nivel(valor: str) -> str:
    normalizado = _normalize_key(valor)
    mapa = {
        "EDUCACION INICIAL": "Inicial",
        "EDUCACION PRIMARIA": "Primaria",
        "EDUCACION SECUNDARIA": "Secundaria",
    }
    return mapa.get(normalizado, "")


def _codigo_nivel(nivel_legible: str) -> str:
    """Devuelve el código de nivel usado en el nombre de clase (P o S)."""
    return {"Inicial": "I", "Primaria": "P", "Secundaria": "S"}.get(nivel_legible, "")


def _letra_tecpro(grado_num: int, nivel_legible: str) -> str:
    """Mapea grado a letra para nombres Tecpro."""
    if nivel_legible == "Primaria":
        return {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F"}.get(grado_num, "")
    if nivel_legible == "Secundaria":
        return {1: "G", 2: "H", 3: "I", 4: "J", 5: "K"}.get(grado_num, "")
    return ""


def _termina_con_seccion(nombre: str) -> bool:
    """Detecta sufijos tipo 1PA o P1A al final del nombre."""
    if not nombre:
        return False
    token = nombre.strip().split()[-1]
    token = re.sub(r"[^A-Za-z0-9]+", "", token).upper()
    if not token:
        return False
    if re.match(r"^(\d{1,2})([IPS])([A-Z])$", token):
        return True
    if re.match(r"^([IPS])(\d{1,2})([A-Z])$", token):
        return True
    return False


def _mapear_grado(valor: str) -> Tuple[str, int]:
    """Devuelve (texto_grado, numero) o ('', 0) si no coincide."""
    limpio = valor.replace("°", "º").replace("�", "º")
    normalizado = _normalize_key(limpio)
    match = re.match(r"^(\d+)\s*º?\s*(PRIMARIA|SECUNDARIA)$", normalizado)
    if match:
        numero = int(match.group(1))
        nivel = match.group(2)
    else:
        nivel = ""
        if "PRIMARIA" in normalizado:
            nivel = "PRIMARIA"
        elif "SECUNDARIA" in normalizado:
            nivel = "SECUNDARIA"
        if not nivel:
            return ("", 0)
        num_match = re.search(r"\d+", normalizado)
        if not num_match:
            return ("", 0)
        numero = int(num_match.group(0))

    grados_primaria: Dict[int, str] = {
        1: "Primer grado de primaria",
        2: "Segundo grado de primaria",
        3: "Tercer grado de primaria",
        4: "Cuarto grado de primaria",
        5: "Quinto grado de primaria",
        6: "Sexto grado de primaria",
    }
    grados_secundaria: Dict[int, str] = {
        1: "Primer año de secundaria",
        2: "Segundo año de secundaria",
        3: "Tercer año de secundaria",
        4: "Cuarto año de secundaria",
        5: "Quinto año de secundaria",
    }

    if nivel == "PRIMARIA":
        return (grados_primaria.get(numero, ""), numero if numero <= 6 else 0)
    return (grados_secundaria.get(numero, ""), numero if numero <= 5 else 0)


def _mapear_materia(valor: str) -> Tuple[str, str]:
    """Devuelve (materia_legible, sufijo) o ('', '') si no coincide."""
    normalizado = _canonical_materia_key(valor)
    mapa: Dict[str, Tuple[str, str]] = {
        "CALIGRAFIA": ("Caligrafía", "CA"),
        "CIENCIA Y TECNOLOGIA": ("Ciencia y Tecnología", "CT"),
        "CIENCIAS INTEGRADAS": ("Ciencias Integradas", "CI"),
        "CIENCIAS SOCIALES": ("Ciencias Sociales", "CS"),
        "CIUDADANIA Y CIVICA": ("Desarrollo Personal Ciudadanía y Cívica", "DPCC"),
        "COMUNICACION": ("Comunicación", "CO"),
        "DESARROLLO PERSONAL": ("Desarrollo Personal Ciudadanía y Cívica", "DPCC"),
        "DESARROLLO PERSONAL CIUDADANIA Y CIVICA": (
            "Desarrollo Personal Ciudadanía y Cívica",
            "DPCC",
        ),
        "DPCC": ("Desarrollo Personal Ciudadanía y Cívica", "DPCC"),
        "TECNOLOGIA": ("Tecnología", "TE"),
        "INFORMATICA": ("Tecnología", "TE"),
        "INGLES": ("Inglés", "IG"),
        "BIOLOGIA": ("Biología-Física-Química", "BFQ"),
        "FISICA": ("Biología-Física-Química", "BFQ"),
        "QUIMICA": ("Biología-Física-Química", "BFQ"),
        "BIOLOGIA-FISICA-QUIMICA": ("Biología-Física-Química", "BFQ"),
        "LECTURAS": ("Lecturas", "LE"),
        "MATEMATICAS": ("Matemática", "MA"),
        "MATEMATICA": ("Matemática", "MA"),
        "PERSONAL SOCIAL": ("Personal Social", "PS"),
        "PLAN LECTOR": ("Plan Lector", "PL"),
        "PREESCOLAR": ("Preescolar", "PE"),
        "RAZONAMIENTO MATEMATICO": ("Razonamiento Matemático", "RM"),
        "RAZONAMIENTO VERBAL": ("Razonamiento Verbal", "RV"),
        "RELIGION": ("Educación Religiosa", "RE"),
        "TUTORIA Y ORIENTACION EDUCATIVA": ("Tutoría y Orientación Educativa", "TO"),
    }
    materia = mapa.get(normalizado)
    return materia if materia else ("", "")


def _orden_materia(materia: str) -> int:
    orden = {
        "Ciencia y Tecnología": 0,
        "Ciencia y Tecnología - Personal Social": 1,
        "Ciencias Integradas": 2,
        "Ciencias Sociales": 3,
        "Biología-Física-Química": 4,
        "Comunicación": 5,
        "Desarrollo Personal Ciudadanía y Cívica": 6,
        "Tecnología": 7,
        "Inglés": 8,
        "Lecturas": 9,
        "Matemática": 10,
        "Personal Social": 11,
        "Plan Lector": 12,
        "Preescolar": 13,
        "Razonamiento Matemático": 14,
        "Razonamiento Verbal": 15,
        "Educación Religiosa": 16,
        "Tutoría y Orientación Educativa": 17,
        "Caligrafía": 18,
    }
    return orden.get(materia, len(orden))


def transformar(df: pd.DataFrame, grupos: Optional[Sequence[str]] = None) -> pd.DataFrame:
    """Aplica filtros y transforma los datos para la exportación."""
    columnas_necesarias = {
        "Plataforma",
        "Asignatura Producto",
        "Nivel Educativo",
        "Grado",
        "Producto",
    }
    faltantes = [col for col in columnas_necesarias if col not in df.columns]
    if faltantes:
        raise KeyError(f"Faltan columnas necesarias para transformar: {', '.join(faltantes)}")

    trabajo = df.copy().fillna("")
    for col in (
        "Institucion",
        "Nivel Educativo",
        "Grado",
        "Asignatura Producto",
        "Plataforma",
    ):
        if col in trabajo.columns:
            trabajo[col] = trabajo[col].replace("", pd.NA).ffill().fillna("")
    trabajo = trabajo.assign(
        Plataforma=trabajo["Plataforma"].astype(str).str.strip(),
        AsignaturaProducto=trabajo["Asignatura Producto"].astype(str).str.strip(),
        NivelEducativo=trabajo["Nivel Educativo"].astype(str).str.strip(),
        GradoVal=trabajo["Grado"].astype(str).str.strip(),
        ProductoVal=trabajo["Producto"].astype(str).str.strip(),
    )
    if "Razon Estado" in trabajo.columns:
        trabajo["RazonEstado"] = trabajo["Razon Estado"].astype(str).str.strip()
    else:
        trabajo["RazonEstado"] = ""

    trabajo["PlataformaNorm"] = trabajo["Plataforma"].apply(_normalize_key)

    if trabajo.empty:
        return pd.DataFrame()

    grupos = _normalize_grupos(grupos)
    multi_grupo = len(grupos) > 1
    grupo_order = {grupo: idx for idx, grupo in enumerate(grupos, start=1)}

    registros: List[Dict[str, str]] = []
    for _, fila in trabajo.iterrows():
        nivel_legible = _mapear_nivel(fila["NivelEducativo"])
        grado_legible, grado_num = _mapear_grado(fila["GradoVal"])
        materia_legible, sufijo = _mapear_materia(fila["AsignaturaProducto"])
        materia_key = _canonical_materia_key(fila["AsignaturaProducto"])
        nivel_codigo = _codigo_nivel(nivel_legible)
        plataforma_norm = fila["PlataformaNorm"]
        producto_val = fila["ProductoVal"]
        prod_upper = producto_val.upper()

        if (fila["RazonEstado"] not in {"Validado", ""}) or (plataforma_norm == "RLP"):
            continue

        es_richmond = "RICHMOND" in plataforma_norm
        if es_richmond and materia_key != "INGLES":
            if not producto_val.upper().startswith("CODIGO DE ACCESO AL SISTEMA"):
                continue

        if materia_legible == "Tecnología" and producto_val:
            if producto_val.upper().startswith("CODIGO DE ACCESO AL SISTEMA TECPRO"):
                letra = _letra_tecpro(grado_num, nivel_legible)
                if letra:
                    pref = "TECPRO MAX" if nivel_legible == "Secundaria" else "TECPRO"
                    producto_val = f"{pref} NIVEL {letra}"

        if (
            materia_legible == "Ciencias Integradas"
            and nivel_legible == "Primaria"
            and grado_num in {1, 2}
        ):
            materia_legible = "Ciencia y Tecnología - Personal Social"

        if materia_legible == "Preescolar":
            if not (
                prod_upper.startswith("CODIGO DE ACCESO AL SISTEMA")
                or "KURMI" in prod_upper
            ):
                continue
            grado_norm = _normalize_key(fila["GradoVal"])
            mapa_inicial = {
                "DOS": 2,
                "2": 2,
                "TRES": 3,
                "3": 3,
                "CUATRO": 4,
                "4": 4,
                "CINCO": 5,
                "5": 5,
                "UNO": 1,
                "1": 1,
            }
            if not grado_num:
                clave = grado_norm.split()[0] if grado_norm else ""
                grado_num = mapa_inicial.get(clave, 1)
            nombres_inicial = {
                1: "1 año",
                2: "2 años",
                3: "3 años",
                4: "4 años",
                5: "5 años",
            }
            if grado_num in nombres_inicial:
                grado_legible = nombres_inicial[grado_num]
            elif not grado_legible:
                grado_legible = fila["GradoVal"] or "Inicial"
            if not nivel_legible:
                nivel_legible = "Inicial"
                nivel_codigo = _codigo_nivel(nivel_legible)

        if not _materia_permitida(nivel_legible, materia_key, grado_num):
            continue

        if materia_legible == "Preescolar":
            nombre_clase = producto_val if producto_val else "Kurmi (1IA)"
            if "KURMI" in prod_upper:
                match = re.search(r"KURMI\s*\d+", producto_val, flags=re.IGNORECASE)
                if match:
                    nombre_clase = match.group(0).upper()
                else:
                    nombre_clase = producto_val.split("(", 1)[0].strip().upper() or nombre_clase
        elif (
            materia_legible == "Plan Lector"
            and grado_num
            and nivel_codigo
            and ("LQL" in prod_upper or "LOQUELEO" in prod_upper)
        ):
            nombre_clase = f"Loqueleo {grado_num}{nivel_codigo}A"
        elif materia_legible == "Tecnología" and producto_val:
            nombre_clase = producto_val
            if "TECPRO" in prod_upper and grado_num and nivel_codigo:
                if not _termina_con_seccion(nombre_clase):
                    nombre_clase = f"{nombre_clase} {grado_num}{nivel_codigo}A"
        else:
            nombre_clase = f"{materia_legible} {grado_num}{nivel_codigo}A"

        if not (
            nivel_legible
            and grado_legible
            and grado_num
            and materia_legible
            and sufijo
            and nivel_codigo
        ):
            continue

        for grupo in grupos:
            nombre_grupo = _apply_grupo(
                nombre_clase, grupo, grado_num, nivel_codigo, multi_grupo
            )
            registros.append(
                {
                    "Nivel": nivel_legible,
                    "Grado": grado_legible,
                    "Grupo": f"Grupo {grupo}",
                    "Nombre de Clase": nombre_grupo,
                    "Clase Clave": nombre_grupo,
                    "Alias Clase": "",
                    "Materias": materia_legible,
                    "_orden_nivel": 0
                    if nivel_legible == "Inicial"
                    else (1 if nivel_legible == "Primaria" else 2),
                    "_orden_grado": grado_num,
                    "_orden_materia": _orden_materia(materia_legible),
                    "_orden_grupo": grupo_order.get(grupo, 0),
                }
            )

    if not registros:
        return pd.DataFrame()

    salida = pd.DataFrame(registros)
    salida = salida.drop_duplicates(
        subset=["Nivel", "Grado", "Grupo", "Nombre de Clase", "Clase Clave", "Materias"],
        keep="first",
    )
    salida = salida.sort_values(
        by=["_orden_nivel", "_orden_grado", "_orden_materia", "Materias", "_orden_grupo"],
        ascending=[True, True, True, True, True],
    ).drop(columns=["_orden_nivel", "_orden_grado", "_orden_materia", "_orden_grupo"])

    return salida.reset_index(drop=True)


def exportar_excel(
    df: pd.DataFrame,
    plantilla_bytes: Optional[bytes] = None,
    plantilla_path: Optional[Path] = None,
) -> bytes:
    """
    Exporta el DataFrame a Excel usando una plantilla:
    - Mantiene la hoja existente y limpia solo filas de datos (deja encabezados).
    - Respeta el orden de columnas de la plantilla.
    Devuelve los bytes del archivo.
    """
    wb: Workbook
    if plantilla_bytes:
        wb = load_workbook(BytesIO(plantilla_bytes))
    elif plantilla_path and Path(plantilla_path).exists():
        wb = load_workbook(plantilla_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if OUTPUT_SHEET_NAME in wb.sheetnames:
        ws_out = wb[OUTPUT_SHEET_NAME]
        if ws_out.max_row > 1:
            ws_out.delete_rows(2, ws_out.max_row - 1)
        headers = [cell.value for cell in ws_out[1]]
    else:
        ws_out = wb.create_sheet(OUTPUT_SHEET_NAME)
        headers = list(df.columns)
        ws_out.append(headers)

    try:
        df_to_write = df[headers]
    except KeyError as exc:
        raise KeyError(
            f"No coinciden los encabezados de la plantilla con las columnas generadas: {exc}"
        )

    for row in dataframe_to_rows(df_to_write, index=False, header=False):
        ws_out.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_excel(
    excel_input,
    codigo: str,
    columna_codigo: str = CODE_COLUMN_NAME,
    hoja: str = SHEET_NAME,
    plantilla_bytes: Optional[bytes] = None,
    plantilla_path: Optional[Path] = None,
    grupos: Optional[Sequence[str]] = None,
) -> Tuple[bytes, Dict[str, int]]:
    """
    Flujo completo: carga, filtra, transforma y devuelve los bytes del Excel final.
    """
    df = cargar_excel(excel_input, hoja=hoja, columna_codigo=columna_codigo)
    df_filtrado = filtrar_codigo(df, codigo, columna_codigo=columna_codigo)

    if df_filtrado.empty:
        raise ValueError(f"No se encontraron filas para el código {codigo}.")

    df_transformado = transformar(df_filtrado, grupos=grupos)
    if df_transformado.empty:
        raise ValueError(
            "No hay filas que cumplan con 'Compartir Aprendizajes' y reglas de transformación."
        )

    output_bytes = exportar_excel(
        df_transformado,
        plantilla_bytes=plantilla_bytes,
        plantilla_path=plantilla_path,
    )

    summary = {
        "filas_en_hoja": len(df),
        "filas_filtradas": len(df_filtrado),
        "filas_salida": len(df_transformado),
    }
    return output_bytes, summary
