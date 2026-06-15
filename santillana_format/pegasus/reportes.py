from __future__ import annotations

import re
import unicodedata
from datetime import date
from io import BytesIO
from typing import Dict, Iterable, List, Mapping, Optional, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_LEVEL_LABELS = {
    38: "Inicial",
    39: "Primaria",
    40: "Secundaria",
}


def _safe_int(value: object) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(
        character
        for character in text
        if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"\s+", " ", text).strip()


def _to_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return _normalize_text(value) in {
        "1",
        "ACTIVO",
        "ACTIVA",
        "ACTIVE",
        "ENABLED",
        "SI",
        "TRUE",
        "YES",
    }


def _full_name(
    nombre: object,
    apellido_paterno: object,
    apellido_materno: object,
    fallback: object = "",
) -> str:
    fallback_text = str(fallback or "").strip()
    if fallback_text:
        return fallback_text
    return " ".join(
        part
        for part in (
            str(nombre or "").strip(),
            str(apellido_paterno or "").strip(),
            str(apellido_materno or "").strip(),
        )
        if part
    ).strip()


def _level_ids(row: Mapping[str, object]) -> List[int]:
    values: List[int] = []
    for key in (
        "niveles_detalle_activos",
        "niveles_presentes",
        "niveles_detalle",
    ):
        raw_values = row.get(key)
        if not isinstance(raw_values, (list, tuple, set)):
            continue
        for raw_value in raw_values:
            value = _safe_int(raw_value)
            if value is not None and value not in values:
                values.append(value)
    active_map = row.get("niveles_activos")
    if isinstance(active_map, dict):
        for raw_value, active in active_map.items():
            value = _safe_int(raw_value)
            if value is not None and _to_bool(active) and value not in values:
                values.append(value)
    return sorted(values)


def build_colegio_row(row: Mapping[str, object]) -> Dict[str, object]:
    colegio_id = _safe_int(row.get("colegio_id"))
    colegio = str(row.get("colegio") or row.get("label") or "").strip()
    return {
        "Colegio ID": colegio_id or "",
        "Colegio": colegio,
        "Clave": str(row.get("colegio_clave") or "").strip(),
        "SAP ID": str(row.get("sap_id") or "").strip(),
        "CRM ID": str(row.get("crm_id") or "").strip(),
        "Cliente": str(row.get("cliente") or "").strip(),
        "Municipio": str(row.get("municipio") or "").strip(),
        "Departamento": str(row.get("departamento") or "").strip(),
        "Telefono": str(row.get("telefono") or "").strip(),
        "Demo": "Si" if _to_bool(row.get("demo")) else "No",
    }


def build_clase_row(
    row: Mapping[str, object],
    colegio_row: Mapping[str, object],
) -> Dict[str, object]:
    colegio_id = _safe_int(colegio_row.get("colegio_id"))
    colegio = str(
        colegio_row.get("colegio")
        or colegio_row.get("label")
        or f"Colegio {colegio_id or ''}"
    ).strip()
    nivel_id = _safe_int(row.get("nivel_id"))
    grado_id = _safe_int(row.get("grado_id"))
    clase_id = _safe_int(row.get("clase_id"))
    active = _to_bool(row.get("activo")) and not _to_bool(row.get("baja"))
    seccion = str(row.get("seccion") or "").strip()
    return {
        "Colegio ID": colegio_id or "",
        "Colegio": colegio,
        "Clase ID": clase_id or "",
        "Codigo": str(row.get("clase_codigo") or "").strip(),
        "Clase": str(
            row.get("clase_nombre") or row.get("clase") or ""
        ).strip(),
        "Nivel": str(
            row.get("nivel")
            or REPORT_LEVEL_LABELS.get(nivel_id or 0)
            or ""
        ).strip(),
        "Nivel ID": nivel_id or "",
        "Grado": str(row.get("grado") or "").strip(),
        "Grado ID": grado_id or "",
        "Seccion": seccion,
        "Grupo ID": _safe_int(row.get("grupo_id")) or "",
        "Activo": "Si" if active else "No",
        "Baja": "Si" if _to_bool(row.get("baja")) else "No",
        "Alumnos": "",
        "Profesores": "",
        "_colegio_id": colegio_id,
        "_clase_id": clase_id,
        "_nivel_id": nivel_id,
        "_grado_id": grado_id,
        "_seccion": seccion,
        "_activo": active,
        "_alumnos_total": None,
        "_profesores_total": None,
    }


def build_profesor_row(
    row: Mapping[str, object],
    colegio_row: Mapping[str, object],
) -> Dict[str, object]:
    colegio_id = _safe_int(colegio_row.get("colegio_id"))
    colegio = str(
        colegio_row.get("colegio")
        or colegio_row.get("label")
        or f"Colegio {colegio_id or ''}"
    ).strip()
    level_ids = _level_ids(row)
    level_names = [
        REPORT_LEVEL_LABELS.get(level_id, f"Nivel {level_id}")
        for level_id in level_ids
    ]
    login = str(row.get("login") or "").strip()
    status_text = str(row.get("estado") or "").strip()
    active = _to_bool(row.get("login_activo"))
    if row.get("login_activo") is None:
        active = _normalize_text(status_text) in {"ACTIVO", "ACTIVA"}
    return {
        "Colegio ID": colegio_id or "",
        "Colegio": colegio,
        "Persona ID": _safe_int(row.get("persona_id")) or "",
        "Profesor": _full_name(
            row.get("nombre"),
            row.get("apellido_paterno"),
            row.get("apellido_materno"),
        ),
        "Nombre": str(row.get("nombre") or "").strip(),
        "Apellido paterno": str(row.get("apellido_paterno") or "").strip(),
        "Apellido materno": str(row.get("apellido_materno") or "").strip(),
        "DNI": str(row.get("dni") or "").strip(),
        "E-mail": str(row.get("email") or "").strip(),
        "Login": login,
        "Estado": status_text or ("Activo" if active else "Inactivo"),
        "Niveles": ", ".join(level_names),
        "Nivel IDs": ", ".join(str(value) for value in level_ids),
        "_colegio_id": colegio_id,
        "_nivel_ids": level_ids,
        "_activo": active,
        "_login": login,
    }


def build_alumno_row(
    row: Mapping[str, object],
    colegio_row: Mapping[str, object],
) -> Dict[str, object]:
    colegio_id = _safe_int(colegio_row.get("colegio_id"))
    colegio = str(
        colegio_row.get("colegio")
        or colegio_row.get("label")
        or f"Colegio {colegio_id or ''}"
    ).strip()
    nivel_id = _safe_int(row.get("nivel_id"))
    grado_id = _safe_int(row.get("grado_id"))
    seccion = str(
        row.get("seccion_norm")
        or row.get("grupo_clave")
        or row.get("seccion")
        or ""
    ).strip()
    login = str(row.get("login") or "").strip()
    active = _to_bool(row.get("activo"))
    con_pago = _to_bool(row.get("con_pago"))
    return {
        "Colegio ID": colegio_id or "",
        "Colegio": colegio,
        "Alumno ID": _safe_int(row.get("alumno_id")) or "",
        "Persona ID": _safe_int(row.get("persona_id")) or "",
        "Alumno": _full_name(
            row.get("nombre"),
            row.get("apellido_paterno"),
            row.get("apellido_materno"),
            fallback=row.get("nombre_completo"),
        ),
        "Nombre": str(row.get("nombre") or "").strip(),
        "Apellido paterno": str(row.get("apellido_paterno") or "").strip(),
        "Apellido materno": str(row.get("apellido_materno") or "").strip(),
        "DNI": str(row.get("id_oficial") or "").strip(),
        "Login": login,
        "Nivel": str(
            row.get("nivel")
            or REPORT_LEVEL_LABELS.get(nivel_id or 0)
            or ""
        ).strip(),
        "Nivel ID": nivel_id or "",
        "Grado": str(row.get("grado") or "").strip(),
        "Grado ID": grado_id or "",
        "Seccion": seccion,
        "Grupo ID": _safe_int(row.get("grupo_id")) or "",
        "Activo": "Si" if active else "No",
        "Con pago": "Si" if con_pago else "No",
        "Fecha desde": str(row.get("fecha_desde") or "").strip(),
        "_colegio_id": colegio_id,
        "_nivel_id": nivel_id,
        "_grado_id": grado_id,
        "_seccion": seccion,
        "_activo": active,
        "_login": login,
        "_con_pago": con_pago,
    }


def build_profesor_clase_row(
    profesor: Mapping[str, object],
    clase: Mapping[str, object],
) -> Dict[str, object]:
    login = str(profesor.get("login") or "").strip()
    active = _to_bool(profesor.get("activo"))
    return {
        "Colegio ID": clase.get("Colegio ID", ""),
        "Colegio": clase.get("Colegio", ""),
        "Clase ID": clase.get("Clase ID", ""),
        "Clase": clase.get("Clase", ""),
        "Nivel": clase.get("Nivel", ""),
        "Nivel ID": clase.get("Nivel ID", ""),
        "Grado": clase.get("Grado", ""),
        "Grado ID": clase.get("Grado ID", ""),
        "Seccion": clase.get("Seccion", ""),
        "Persona ID": _safe_int(profesor.get("persona_id")) or "",
        "Profesor": _full_name(
            profesor.get("nombre_base"),
            profesor.get("apellido_paterno"),
            profesor.get("apellido_materno"),
            fallback=profesor.get("nombre"),
        ),
        "DNI": str(profesor.get("dni") or "").strip(),
        "Login": login,
        "Activo": "Si" if active else "No",
        "_colegio_id": clase.get("_colegio_id"),
        "_nivel_id": clase.get("_nivel_id"),
        "_grado_id": clase.get("_grado_id"),
        "_seccion": clase.get("_seccion"),
        "_activo": active,
        "_login": login,
    }


def build_alumno_clase_row(
    alumno: Mapping[str, object],
    clase: Mapping[str, object],
) -> Dict[str, object]:
    login = str(alumno.get("login") or "").strip()
    return {
        "Colegio ID": clase.get("Colegio ID", ""),
        "Colegio": clase.get("Colegio", ""),
        "Clase ID": clase.get("Clase ID", ""),
        "Clase": clase.get("Clase", ""),
        "Nivel": clase.get("Nivel", ""),
        "Nivel ID": clase.get("Nivel ID", ""),
        "Grado": clase.get("Grado", ""),
        "Grado ID": clase.get("Grado ID", ""),
        "Seccion": clase.get("Seccion", ""),
        "Alumno ID": _safe_int(alumno.get("alumno_id")) or "",
        "Alumno": str(alumno.get("nombre_completo") or "").strip(),
        "DNI": str(alumno.get("dni") or "").strip(),
        "Login": login,
        "_colegio_id": clase.get("_colegio_id"),
        "_nivel_id": clase.get("_nivel_id"),
        "_grado_id": clase.get("_grado_id"),
        "_seccion": clase.get("_seccion"),
        "_login": login,
    }


def filter_report_rows(
    rows: Sequence[Mapping[str, object]],
    *,
    nivel_ids: Optional[Iterable[int]] = None,
    grado_ids: Optional[Iterable[int]] = None,
    secciones: Optional[Iterable[str]] = None,
    estado: str = "todos",
    login: str = "todos",
    pago: str = "todos",
    alumnos_clase: str = "todos",
    profesores_clase: str = "todos",
    search_text: object = "",
) -> List[Dict[str, object]]:
    selected_levels = {
        int(value)
        for value in (nivel_ids or [])
        if _safe_int(value) is not None
    }
    selected_grades = {
        int(value)
        for value in (grado_ids or [])
        if _safe_int(value) is not None
    }
    selected_sections = {
        _normalize_text(value)
        for value in (secciones or [])
        if _normalize_text(value)
    }
    search_tokens = [
        token
        for token in _normalize_text(search_text).split(" ")
        if token
    ]

    filtered: List[Dict[str, object]] = []
    for raw_row in rows:
        row = dict(raw_row)
        row_level_ids = {
            int(value)
            for value in (row.get("_nivel_ids") or [])
            if _safe_int(value) is not None
        }
        row_level_id = _safe_int(row.get("_nivel_id"))
        if row_level_id is not None:
            row_level_ids.add(row_level_id)
        has_level_dimension = "_nivel_id" in row or "_nivel_ids" in row
        if (
            selected_levels
            and has_level_dimension
            and not row_level_ids.intersection(selected_levels)
        ):
            continue

        row_grade_id = _safe_int(row.get("_grado_id"))
        if (
            selected_grades
            and "_grado_id" in row
            and row_grade_id not in selected_grades
        ):
            continue

        row_section = _normalize_text(row.get("_seccion"))
        if (
            selected_sections
            and "_seccion" in row
            and row_section not in selected_sections
        ):
            continue

        if estado != "todos" and "_activo" in row:
            is_active = bool(row.get("_activo"))
            if estado == "activos" and not is_active:
                continue
            if estado == "inactivos" and is_active:
                continue

        if login != "todos" and "_login" in row:
            has_login = bool(str(row.get("_login") or "").strip())
            if login == "con" and not has_login:
                continue
            if login == "sin" and has_login:
                continue

        if pago != "todos" and "_con_pago" in row:
            has_payment = bool(row.get("_con_pago"))
            if pago == "con" and not has_payment:
                continue
            if pago == "sin" and has_payment:
                continue

        alumnos_total = row.get("_alumnos_total")
        if alumnos_clase != "todos" and alumnos_total is not None:
            has_students = int(alumnos_total or 0) > 0
            if alumnos_clase == "con" and not has_students:
                continue
            if alumnos_clase == "sin" and has_students:
                continue

        profesores_total = row.get("_profesores_total")
        if profesores_clase != "todos" and profesores_total is not None:
            has_teachers = int(profesores_total or 0) > 0
            if profesores_clase == "con" and not has_teachers:
                continue
            if profesores_clase == "sin" and has_teachers:
                continue

        if search_tokens:
            searchable = _normalize_text(
                " ".join(
                    str(value or "")
                    for key, value in row.items()
                    if not str(key).startswith("_")
                )
            )
            if not all(token in searchable for token in search_tokens):
                continue
        filtered.append(row)
    return filtered


def clean_report_rows(
    rows: Sequence[Mapping[str, object]],
) -> List[Dict[str, object]]:
    return [
        {
            str(key): value
            for key, value in row.items()
            if not str(key).startswith("_")
        }
        for row in rows
    ]


def _safe_sheet_name(raw_name: object, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]+", " ", str(raw_name or "Datos")).strip()
    base = re.sub(r"\s+", " ", base)[:31] or "Datos"
    candidate = base
    suffix = 2
    while candidate.lower() in used:
        suffix_text = f" {suffix}"
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def export_report_workbook(
    sheets: Mapping[str, Sequence[Mapping[str, object]]],
    *,
    summary_rows: Sequence[Mapping[str, object]],
    errors: Sequence[Mapping[str, object]],
    config_rows: Sequence[Mapping[str, object]],
) -> bytes:
    output = BytesIO()
    used_names: set[str] = set()
    workbook_sheets: List[tuple[str, pd.DataFrame]] = [
        (
            _safe_sheet_name("Resumen", used_names),
            pd.DataFrame(clean_report_rows(summary_rows)),
        ),
        (
            _safe_sheet_name("Configuracion", used_names),
            pd.DataFrame(clean_report_rows(config_rows)),
        ),
    ]
    for sheet_name, rows in sheets.items():
        workbook_sheets.append(
            (
                _safe_sheet_name(sheet_name, used_names),
                pd.DataFrame(clean_report_rows(rows)),
            )
        )
    if errors:
        workbook_sheets.append(
            (
                _safe_sheet_name("Errores", used_names),
                pd.DataFrame(clean_report_rows(errors)),
            )
        )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in workbook_sheets:
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            if dataframe.columns.size:
                worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
            for column_index, column_name in enumerate(
                dataframe.columns,
                start=1,
            ):
                sample = (
                    dataframe[column_name].astype(str).head(300).tolist()
                    if not dataframe.empty
                    else []
                )
                max_length = max(
                    [len(str(column_name))]
                    + [len(value) for value in sample]
                )
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = min(max(max_length + 2, 10), 55)
            worksheet.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


def build_report_filename(
    colegio_count: int,
    *,
    generated_on: Optional[date] = None,
) -> str:
    report_date = generated_on or date.today()
    scope = "un_colegio" if int(colegio_count) == 1 else f"{int(colegio_count)}_colegios"
    return f"reporte_pegasus_{scope}_{report_date.isoformat()}.xlsx"
