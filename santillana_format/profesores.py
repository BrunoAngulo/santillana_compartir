from io import BytesIO
from typing import Callable, Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_EMPRESA_ID = 11
DEFAULT_CICLO_ID = 207

BASE_URL = (
    "https://www.uno-internacional.com/pegasus-api/censo/empresas/{empresa_id}"
    "/ciclos/{ciclo_id}/colegios/{colegio_id}/niveles/{nivel_id}/profesores"
)
FILTERS_URL = (
    "https://www.uno-internacional.com/pegasus-api/censo/empresas/{empresa_id}"
    "/ciclos/{ciclo_id}/colegios/{colegio_id}/profesoresByFilters"
)

NIVEL_MAP = {
    "Inicial": 38,
    "Primaria": 39,
    "Secundaria": 40,
}
NIVEL_ID_TO_NAME = {nivel_id: nombre for nombre, nivel_id in NIVEL_MAP.items()}

PROFESOR_COLUMNS = [
    "Id",
    "Nombre",
    "Apellido Paterno",
    "Apellido Materno",
    "Estado",
    "Sexo",
    "DNI",
    "E-mail",
    "Login",
    "Password",
    "Inicial",
    "Primaria",
    "Secundaria",
    "I3",
    "I4",
    "I5",
    "P1",
    "P2",
    "P3",
    "P4",
    "P5",
    "P6",
    "S1",
    "S2",
    "S3",
    "S4",
    "S5",
    "Clases",
    "Secciones",
]

ERROR_COLUMNS = [
    "tipo",
    "nivel_id",
    "persona_id",
    "url",
    "status_code",
    "error",
]


def build_profesores_filename(colegio_id: int) -> str:
    if colegio_id:
        return f"profesores_{int(colegio_id)}.xlsx"
    return "profesores.xlsx"


def _build_url(context: Dict[str, int], persona_id: Optional[int] = None) -> str:
    url = BASE_URL.format(
        empresa_id=context["empresa_id"],
        ciclo_id=context["ciclo_id"],
        colegio_id=context["colegio_id"],
        nivel_id=context["nivel_id"],
    )
    if persona_id is not None:
        return f"{url}/{persona_id}"
    return url


def _fetch_profesores_list(
    session: requests.Session,
    token: str,
    context: Dict[str, int],
    timeout: int = 30,
) -> Tuple[List[Dict[str, object]], Optional[str], Optional[int], str]:
    url = _build_url(context)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }
    try:
        response = session.get(url, headers=headers, timeout=timeout)
    except requests.RequestException as exc:
        return [], str(exc), None, url

    status_code = response.status_code
    try:
        payload = response.json()
    except ValueError:
        return [], f"Respuesta no JSON (status {status_code})", status_code, url

    if not response.ok:
        message = payload.get("message") if isinstance(payload, dict) else ""
        return [], message or f"HTTP {status_code}", status_code, url

    if not isinstance(payload, dict) or not payload.get("success", False):
        message = payload.get("message") if isinstance(payload, dict) else "Respuesta invalida"
        return [], message or "Respuesta invalida", status_code, url

    data = payload.get("data") or []
    if not isinstance(data, list):
        return [], "Campo data no es lista", status_code, url
    return data, None, status_code, url


def _fetch_profesores_by_filters(
    session: requests.Session,
    token: str,
    empresa_id: int,
    ciclo_id: int,
    colegio_id: int,
    timeout: int = 30,
) -> Tuple[List[Dict[str, object]], Optional[str], Optional[int], str]:
    url = FILTERS_URL.format(
        empresa_id=int(empresa_id),
        ciclo_id=int(ciclo_id),
        colegio_id=int(colegio_id),
    )
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }
    try:
        response = session.get(url, headers=headers, timeout=timeout)
    except requests.RequestException as exc:
        return [], str(exc), None, url

    status_code = response.status_code
    try:
        payload = response.json()
    except ValueError:
        return [], f"Respuesta no JSON (status {status_code})", status_code, url

    if not response.ok:
        message = payload.get("message") if isinstance(payload, dict) else ""
        return [], message or f"HTTP {status_code}", status_code, url

    if not isinstance(payload, dict) or not payload.get("success", False):
        message = payload.get("message") if isinstance(payload, dict) else "Respuesta invalida"
        return [], message or "Respuesta invalida", status_code, url

    data = payload.get("data") or []
    if not isinstance(data, list):
        return [], "Campo data no es lista", status_code, url
    return data, None, status_code, url


def _fetch_profesor_detail(
    session: requests.Session,
    token: str,
    context: Dict[str, int],
    persona_id: int,
    timeout: int = 30,
) -> Tuple[Optional[Dict[str, object]], Optional[str], Optional[int], str]:
    url = _build_url(context, persona_id=persona_id)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }
    try:
        response = session.get(url, headers=headers, timeout=timeout)
    except requests.RequestException as exc:
        return None, str(exc), None, url

    status_code = response.status_code
    try:
        payload = response.json()
    except ValueError:
        return None, f"Respuesta no JSON (status {status_code})", status_code, url

    if not response.ok:
        message = payload.get("message") if isinstance(payload, dict) else ""
        return None, message or f"HTTP {status_code}", status_code, url

    if not isinstance(payload, dict) or not payload.get("success", False):
        message = payload.get("message") if isinstance(payload, dict) else "Respuesta invalida"
        return None, message or "Respuesta invalida", status_code, url

    data = payload.get("data") or {}
    if not isinstance(data, dict):
        return None, "Campo data no es objeto", status_code, url
    return data, None, status_code, url


def _ensure_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=columns)
    return df.reindex(columns=columns)


def _extract_niveles(detail: Dict[str, object], only_activos: bool = True) -> Set[int]:
    niveles: Set[int] = set()
    for entry in detail.get("niveles") or []:
        if not isinstance(entry, dict):
            continue
        if only_activos and entry.get("activo") is False:
            continue
        nivel = entry.get("nivel") if isinstance(entry.get("nivel"), dict) else {}
        nivel_id = nivel.get("nivelId") or entry.get("nivelId")
        if nivel_id is None:
            continue
        try:
            niveles.add(int(nivel_id))
        except (TypeError, ValueError):
            continue

    if niveles:
        return niveles

    for entry in detail.get("personaRoles") or []:
        if not isinstance(entry, dict):
            continue
        if only_activos and entry.get("activo") is False:
            continue
        nivel = entry.get("nivel") if isinstance(entry.get("nivel"), dict) else {}
        nivel_id = nivel.get("nivelId")
        if nivel_id is None:
            continue
        try:
            niveles.add(int(nivel_id))
        except (TypeError, ValueError):
            continue
    return niveles


def _extract_niveles_activos_map(detail: Dict[str, object]) -> Dict[int, bool]:
    activos: Dict[int, bool] = {}
    for entry in detail.get("niveles") or []:
        if not isinstance(entry, dict):
            continue
        nivel = entry.get("nivel") if isinstance(entry.get("nivel"), dict) else {}
        nivel_id = nivel.get("nivelId") or entry.get("nivelId")
        if nivel_id is None:
            continue
        activo_value = entry.get("activo")
        if activo_value is None:
            continue
        try:
            nivel_id_int = int(nivel_id)
            activo = _parse_activo(activo_value)
            if nivel_id_int in activos:
                activos[nivel_id_int] = activos[nivel_id_int] or activo
            else:
                activos[nivel_id_int] = activo
        except (TypeError, ValueError):
            continue

    for entry in detail.get("personaRoles") or []:
        if not isinstance(entry, dict):
            continue
        nivel = entry.get("nivel") if isinstance(entry.get("nivel"), dict) else {}
        nivel_id = nivel.get("nivelId") or entry.get("nivelId")
        if nivel_id is None:
            continue
        activo_value = entry.get("activo")
        if activo_value is None:
            continue
        try:
            nivel_id_int = int(nivel_id)
            activo = _parse_activo(activo_value)
            if nivel_id_int in activos:
                activos[nivel_id_int] = activos[nivel_id_int] or activo
            else:
                activos[nivel_id_int] = activo
        except (TypeError, ValueError):
            continue
    return activos


def _derive_estado(activos: Dict[int, bool]) -> str:
    if not activos:
        return ""
    if any(activos.values()):
        return "Activo"
    return "Inactivo"


def _pick_value(detail: Optional[Dict[str, object]], persona: Dict[str, object], key: str) -> object:
    if detail:
        value = detail.get(key)
        if value not in (None, ""):
            return value
    value = persona.get(key)
    if value in (None, ""):
        return ""
    return value


def _parse_activo(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "si", "sí", "yes"}
    return False


def export_profesores_excel(profesores: List[Dict[str, object]]) -> bytes:
    output = BytesIO()
    df_profesores = _ensure_columns(pd.DataFrame(profesores), PROFESOR_COLUMNS)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_profesores.to_excel(writer, index=False, sheet_name="Profesores")
        df_profesores.head(0).to_excel(writer, index=False, sheet_name="Profesores_clases")
        ws = writer.book["Profesores"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, col in enumerate(df_profesores.columns, start=1):
            sample = df_profesores[col].astype(str).head(200).tolist()
            max_len = max([len(str(col))] + [len(val) for val in sample])
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        ws_clases = writer.book["Profesores_clases"]
        ws_clases.freeze_panes = "A2"
        ws_clases.auto_filter.ref = ws_clases.dimensions
        for idx, col in enumerate(df_profesores.columns, start=1):
            max_len = len(str(col))
            ws_clases.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        estado_col_index = None
        try:
            estado_col_index = PROFESOR_COLUMNS.index("Estado") + 1
        except ValueError:
            estado_col_index = None

        if estado_col_index is not None:
            catalog_sheet = "Catalogos"
            if catalog_sheet in writer.book.sheetnames:
                writer.book.remove(writer.book[catalog_sheet])
            catalog_ws = writer.book.create_sheet(catalog_sheet)
            catalog_ws["A1"] = "Activo"
            catalog_ws["A2"] = "Inactivo"
            catalog_ws.sheet_state = "hidden"

            estado_col_letter = get_column_letter(estado_col_index)
            dv = DataValidation(
                type="list",
                formula1=f"={catalog_sheet}!$A$1:$A$2",
                allow_blank=True,
            )
            max_row = max(ws.max_row, 2000)
            ws.add_data_validation(dv)
            dv.add(f"{estado_col_letter}2:{estado_col_letter}{max_row}")

            max_row_clases = max(ws_clases.max_row, 2000)
            ws_clases.add_data_validation(dv)
            dv.add(f"{estado_col_letter}2:{estado_col_letter}{max_row_clases}")

    output.seek(0)
    return output.getvalue()


def listar_profesores_data(
    token: str,
    colegio_id: int,
    nivel_ids: Optional[Sequence[int]] = None,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    ciclo_id: int = DEFAULT_CICLO_ID,
    timeout: int = 30,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[List[Dict[str, object]], Dict[str, int], List[Dict[str, object]]]:
    nivel_ids = list(nivel_ids) if nivel_ids else list(NIVEL_MAP.values())
    contexts = [
        {
            "colegio_id": int(colegio_id),
            "nivel_id": int(nivel_id),
            "empresa_id": int(empresa_id),
            "ciclo_id": int(ciclo_id),
        }
        for nivel_id in nivel_ids
    ]

    profesores: Dict[int, Dict[str, object]] = {}
    errores: List[Dict[str, object]] = []
    seen_roles: Set[int] = set()

    with requests.Session() as session:
        for index, context in enumerate(contexts, start=1):
            data, error, status_code, url = _fetch_profesores_list(
                session=session,
                token=token,
                context=context,
                timeout=timeout,
            )
            if error:
                errores.append(
                    {
                        "tipo": "listado",
                        "nivel_id": context["nivel_id"],
                        "persona_id": "",
                        "url": url,
                        "status_code": status_code or "",
                        "error": error,
                    }
                )
                if on_progress:
                    on_progress(index, len(contexts))
                continue

            for item in data:
                if not isinstance(item, dict):
                    continue
                persona_rol_id = item.get("personaRolId")
                if persona_rol_id is not None:
                    try:
                        persona_rol_id = int(persona_rol_id)
                    except (TypeError, ValueError):
                        persona_rol_id = None
                if persona_rol_id is not None:
                    if persona_rol_id in seen_roles:
                        continue
                    seen_roles.add(persona_rol_id)

                persona = item.get("persona") if isinstance(item, dict) else {}
                if not isinstance(persona, dict):
                    persona = {}
                persona_id = persona.get("personaId")
                if persona_id is None:
                    errores.append(
                        {
                            "tipo": "listado",
                            "nivel_id": context["nivel_id"],
                            "persona_id": "",
                            "url": url,
                            "status_code": status_code or "",
                            "error": "Persona sin personaId.",
                        }
                    )
                    continue
                try:
                    persona_id_int = int(persona_id)
                except (TypeError, ValueError):
                    errores.append(
                        {
                            "tipo": "listado",
                            "nivel_id": context["nivel_id"],
                            "persona_id": persona_id,
                            "url": url,
                            "status_code": status_code or "",
                            "error": "personaId invalido.",
                        }
                    )
                    continue

                entry = profesores.get(persona_id_int)
                if entry is None:
                    entry = {
                        "persona": persona,
                        "niveles": {context["nivel_id"]},
                        "niveles_activos": {context["nivel_id"]: _parse_activo(item.get("activo"))},
                        "detalle": None,
                        "nivel_preferido": context["nivel_id"],
                        "login": "",
                        "niveles_detalle": set(),
                        "niveles_detalle_activos": set(),
                        "estado": "",
                    }
                    profesores[persona_id_int] = entry
                else:
                    entry["niveles"].add(context["nivel_id"])
                    entry["niveles_activos"][context["nivel_id"]] = _parse_activo(
                        item.get("activo")
                    )
                    if not entry.get("persona"):
                        entry["persona"] = persona
                    else:
                        for key, value in persona.items():
                            if entry["persona"].get(key) in (None, "") and value not in (
                                None,
                                "",
                            ):
                                entry["persona"][key] = value

            if on_progress:
                on_progress(index, len(contexts))

        for persona_id, entry in profesores.items():
            context = {
                "colegio_id": int(colegio_id),
                "nivel_id": int(entry["nivel_preferido"]),
                "empresa_id": int(empresa_id),
                "ciclo_id": int(ciclo_id),
            }
            detail, error, status_code, url = _fetch_profesor_detail(
                session=session,
                token=token,
                context=context,
                persona_id=persona_id,
                timeout=timeout,
            )
            if error:
                errores.append(
                    {
                        "tipo": "detalle",
                        "nivel_id": context["nivel_id"],
                        "persona_id": persona_id,
                        "url": url,
                        "status_code": status_code or "",
                        "error": error,
                    }
                )
                continue
            entry["detalle"] = detail
            persona_login = detail.get("personaLogin") if isinstance(detail, dict) else None
            if isinstance(persona_login, dict):
                entry["login"] = persona_login.get("login") or ""
            entry["niveles_detalle"] = _extract_niveles(detail, only_activos=False)
            entry["niveles_detalle_activos"] = _extract_niveles(detail, only_activos=True)
            activos_map = _extract_niveles_activos_map(detail)
            for nivel_id, activo in activos_map.items():
                if activo:
                    entry["niveles_activos"][nivel_id] = True
                else:
                    entry["niveles_activos"].setdefault(nivel_id, False)
            entry["estado"] = _derive_estado(entry.get("niveles_activos", {}))

    resultados: List[Dict[str, object]] = []
    for persona_id, entry in profesores.items():
        persona = entry.get("persona") or {}
        detail = entry.get("detalle")
        resultados.append(
            {
                "persona_id": persona_id,
                "nombre": _pick_value(detail, persona, "nombre"),
                "apellido_paterno": _pick_value(detail, persona, "apellidoPaterno"),
                "apellido_materno": _pick_value(detail, persona, "apellidoMaterno"),
                "sexo": _pick_value(detail, persona, "sexoMoral"),
                "dni": _pick_value(detail, persona, "idOficial"),
                "email": _pick_value(detail, persona, "email"),
                "login": entry.get("login", ""),
                "estado": entry.get("estado", ""),
                "niveles_presentes": set(entry.get("niveles", set())),
                "niveles_activos": dict(entry.get("niveles_activos", {})),
                "niveles_detalle": set(entry.get("niveles_detalle", set())),
                "niveles_detalle_activos": set(entry.get("niveles_detalle_activos", set())),
                "detalle": detail,
            }
        )

    summary = {
        "niveles_total": len(contexts),
        "niveles_error": sum(1 for err in errores if err.get("tipo") == "listado"),
        "profesores_total": len(resultados),
        "detalle_error": sum(1 for err in errores if err.get("tipo") == "detalle"),
    }
    return resultados, summary, errores


def listar_profesores(
    token: str,
    colegio_id: int,
    nivel_ids: Optional[Sequence[int]] = None,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    ciclo_id: int = DEFAULT_CICLO_ID,
    timeout: int = 30,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[bytes, Dict[str, int], List[Dict[str, object]]]:
    data, summary, errores = listar_profesores_data(
        token=token,
        colegio_id=colegio_id,
        nivel_ids=nivel_ids,
        empresa_id=empresa_id,
        ciclo_id=ciclo_id,
        timeout=timeout,
        on_progress=on_progress,
    )

    filas: List[Dict[str, object]] = []
    for entry in data:
        niveles = entry.get("niveles_detalle_activos") or set()
        if not niveles:
            niveles = entry.get("niveles_presentes") or set()
        filas.append(
            {
                "Id": entry.get("persona_id", ""),
                "Nombre": entry.get("nombre", ""),
                "Apellido Paterno": entry.get("apellido_paterno", ""),
                "Apellido Materno": entry.get("apellido_materno", ""),
                "Estado": entry.get("estado", ""),
                "Sexo": entry.get("sexo", ""),
                "DNI": entry.get("dni", ""),
                "E-mail": entry.get("email", ""),
                "Login": entry.get("login", ""),
                "Password": "",
                "Inicial": "SI" if NIVEL_MAP["Inicial"] in niveles else "",
                "Primaria": "SI" if NIVEL_MAP["Primaria"] in niveles else "",
                "Secundaria": "SI" if NIVEL_MAP["Secundaria"] in niveles else "",
                "I3": "",
                "I4": "",
                "I5": "",
                "P1": "",
                "P2": "",
                "P3": "",
                "P4": "",
                "P5": "",
                "P6": "",
                "S1": "",
                "S2": "",
                "S3": "",
                "S4": "",
                "S5": "",
                "Clases": "",
                "Secciones": "",
            }
        )

    output_bytes = export_profesores_excel(filas)
    return output_bytes, summary, errores
