from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROCESS_NAMES = ("RECORDAR", "COMPRENDER", "APLICAR", "ANALIZAR", "EVALUAR", "CREAR")
GENERIC_ITINERARY_TITLES = {"DETALLE", "HOJA", "SHEET", "TAB", "TABLA", "MATRIZ"}

STANDARD_HEADERS = [
    "ITINERARIO",
    "ESTACIÓN",
    "COMPETENCIA",
    "MACROHABILIDAD",
    "MICROHABILIDAD",
    "CONOCIMIENTOS",
    *PROCESS_NAMES,
    "EVIDENCIAS",
    "INSTRUMENTOS DE EVALUACIÓN",
    "CRITERIOS DE EVALUACIÓN",
]

OUTPUT_HEADERS = [
    "ID MICRO HABILIDAD ESPEC\u00cdFICA",
    "\u00c1REA",
    "GRADO",
    "Nivel",
    "BIMESTRE",
    "# ITINERARIO",
    "ITINERARIO",
    "COMPETENCIA",
    "# MACROHABILIDAD",
    "MACROHABILIDAD",
    "# MICROHABILIDAD",
    "MICROHABILIDAD",
    "# ESTACI\u00d3N",
    "ESTACI\u00d3N",
    "CONOCIMIENTOS",
    "CONTENIDO ESPEC\u00cdFICO EVALUADO",
    "# MICROHABILIDADES ESPEC\u00cdFICAS",
    "MICROHABILIDADES ESPEC\u00cdFICAS",
    "PROCESO COGNITIVO",
    "# MICRO TEST",
]

DEFAULT_AREA_BY_CODE = {
    "CO": "Comunicaci\u00f3n",
    "COM": "Comunicaci\u00f3n",
    "MA": "Matem\u00e1tica",
    "MAT": "Matem\u00e1tica",
    "CCSS": "Ciencias sociales",
    "CT": "Ciencia y Tecnolog\u00eda",
    "CCT": "Ciencia y Tecnolog\u00eda",
    "PS": "Personal Social",
}

COURSE_CODES = tuple(DEFAULT_AREA_BY_CODE.keys())
NON_COURSE_CODES = {"SUMUN", "MATRIZ", "ITI", "ITINERARIO", "HITO", "HOJA"}

DEFAULT_LEVEL_BY_NAME = {
    "Primaria": "PRI",
    "Secundaria": "SEC",
}

DEFAULT_COLUMN_WIDTHS = {
    "A": 34,
    "B": 24,
    "C": 10,
    "D": 14,
    "E": 14,
    "F": 12,
    "G": 34,
    "H": 54,
    "I": 18,
    "J": 54,
    "K": 18,
    "L": 54,
    "M": 12,
    "N": 34,
    "O": 54,
    "P": 34,
    "Q": 20,
    "R": 62,
    "S": 20,
    "T": 14,
}


@dataclass(frozen=True)
class MatrixLayout:
    sheet_name: str
    header_row: int
    data_start_row: int
    itinerary_cols: tuple[int, ...]
    itinerary_number_cols: tuple[int, ...] | None
    competence_cols: tuple[int, ...]
    macro_cols: tuple[int, ...]
    micro_cols: tuple[int, ...]
    station_cols: tuple[int, ...]
    knowledge_cols: tuple[int, ...]
    process_cols: dict[str, tuple[int, ...]]
    evidence_cols: tuple[int, ...] | None = None
    instrument_cols: tuple[int, ...] | None = None
    criteria_cols: tuple[int, ...] | None = None


@dataclass(frozen=True)
class SumunTemplateSummary:
    prefix: str
    area: str
    grade: int
    level: str
    output_sheet: str
    generated_rows: int
    macro_count: int
    micro_count: int
    unique_micro_count: int
    processed_sheets: list[str]
    skipped_sheets: list[str]
    rows_by_sheet: dict[str, int]
    specific_rows_by_itinerary: list[dict[str, Any]]
    specific_rows_by_knowledge: list[dict[str, Any]]
    nonnumber_station_rows: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class SumunSheetInspection:
    index: int
    sheet_name: str
    detected: bool
    estimated_rows: int
    reason: str = ""
    empty_field_rows: tuple[tuple[int, tuple[str, ...]], ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class SumunStandardizationSummary:
    generated_rows: int
    processed_sheets: list[str]
    skipped_sheets: list[str]
    rows_by_sheet: dict[str, int]
    output_sheet: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class SheetScanStats:
    estimated_rows: int
    missing_station_rows: tuple[int, ...]
    empty_field_rows: tuple[tuple[int, tuple[str, ...]], ...] = ()


def inspect_sumun_workbook_sheets(excel_bytes: bytes) -> list[SumunSheetInspection]:
    workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
    result: list[SumunSheetInspection] = []
    for index, ws in enumerate(workbook.worksheets, start=1):
        if ws.sheet_state != "visible":
            result.append(
                SumunSheetInspection(
                    index=index,
                    sheet_name=ws.title,
                    detected=False,
                    estimated_rows=0,
                    reason="Hoja oculta.",
                    empty_field_rows=(),
                )
            )
            continue
        values, merged_sources = _fill_merged_values_with_sources(ws)
        layout = _detect_matrix_layout(ws.title, values)
        scan_stats = (
            _scan_sheet_rows(values, layout, merged_sources=merged_sources)
            if layout
            else SheetScanStats(0, ())
        )
        estimated_rows = scan_stats.estimated_rows
        reason = _inspection_reason(layout, scan_stats)
        result.append(
            SumunSheetInspection(
                index=index,
                sheet_name=ws.title,
                detected=bool(layout and estimated_rows),
                estimated_rows=estimated_rows,
                reason=reason,
                empty_field_rows=scan_stats.empty_field_rows,
            )
        )
    return result


def standardize_sumun_workbook_from_excel(
    excel_bytes: bytes,
    *,
    sheet_names: list[str] | None = None,
) -> tuple[bytes, SumunStandardizationSummary]:
    """Flatten a wide SUMUN matrix into one column per standard field."""
    workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
    selected_sheets = set(sheet_names or [])
    itinerary_titles = _collect_itinerary_titles(workbook, sheet_names=sheet_names)
    station_titles = _collect_station_titles(
        workbook,
        sheet_names=sheet_names,
        itinerary_titles=itinerary_titles,
    )
    standard_rows: list[list[Any]] = []
    processed_sheets: list[str] = []
    skipped_sheets: list[str] = []
    rows_by_sheet: dict[str, int] = {}

    for ws in workbook.worksheets:
        if selected_sheets and ws.title not in selected_sheets:
            continue
        if ws.sheet_state != "visible":
            skipped_sheets.append(ws.title)
            continue

        values, merged_sources = _fill_merged_values_with_sources(ws)
        layout = _detect_matrix_layout(ws.title, values)
        if layout is None:
            skipped_sheets.append(ws.title)
            continue

        before_count = len(standard_rows)
        sheet_itinerary = _infer_sheet_itinerary_context(values, layout, ws.title)
        current_itinerary = sheet_itinerary
        last_station_by_itinerary: dict[int, tuple[int, str]] = {}
        station_aliases: dict[tuple[int, str], tuple[int, str]] = {}
        next_station_number_by_itinerary: dict[int, int] = {}

        for row_number in range(layout.data_start_row, len(values) + 1):
            row = values[row_number - 1]
            if not _row_has_standardizable_content(
                row,
                layout,
                merged_sources=merged_sources,
                row_number=row_number,
            ):
                continue

            cell_itinerary = _parse_row_itinerary(row, layout)
            if cell_itinerary:
                current_itinerary = cell_itinerary
            itinerary = _resolve_itinerary_context(
                cell_itinerary,
                current_itinerary,
                itinerary_titles=itinerary_titles,
            )

            itinerary_number = itinerary[0] if itinerary else None
            itinerary_title = itinerary[1] if itinerary else None
            itinerary_value = _standard_itinerary_value(
                itinerary_number,
                itinerary_title,
            )

            station_value = _column_group_value(
                row,
                layout.station_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            )
            standard_station = _clean_text(station_value)
            if itinerary_number is not None:
                station_context, _station_source = _resolve_station_context(
                    station_value,
                    itinerary_number,
                    last_station_by_itinerary=last_station_by_itinerary,
                    station_aliases=station_aliases,
                    next_station_number_by_itinerary=next_station_number_by_itinerary,
                    station_titles=station_titles,
                )
                if station_context:
                    standard_station = _standard_station_value(*station_context)

            process_values = {
                process: "\n\n".join(
                    _specific_skill_group_values(
                        row,
                        layout.process_cols.get(process, ()),
                        merged_sources=merged_sources,
                        row_number=row_number,
                    )
                )
                or None
                for process in PROCESS_NAMES
            }

            standard_rows.append(
                [
                    itinerary_value,
                    standard_station,
                    _standard_group_value(
                        row,
                        layout.competence_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    _standard_group_value(
                        row,
                        layout.macro_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    _standard_group_value(
                        row,
                        layout.micro_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    _standard_group_value(
                        row,
                        layout.knowledge_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    *(process_values[process] for process in PROCESS_NAMES),
                    _standard_optional_group_value(
                        row,
                        layout.evidence_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    _standard_optional_group_value(
                        row,
                        layout.instrument_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                    _standard_optional_group_value(
                        row,
                        layout.criteria_cols,
                        merged_sources=merged_sources,
                        row_number=row_number,
                    ),
                ]
            )

        generated_for_sheet = len(standard_rows) - before_count
        if generated_for_sheet:
            processed_sheets.append(ws.title)
            rows_by_sheet[ws.title] = generated_for_sheet
        else:
            skipped_sheets.append(ws.title)

    if not standard_rows:
        raise ValueError(
            "No se encontraron filas SUMMUN para estandarizar en las hojas seleccionadas."
        )

    output_sheet = "SUMUN_Estandar"
    output_bytes = _write_standard_workbook(
        standard_rows,
        output_sheet=output_sheet,
    )
    summary = SumunStandardizationSummary(
        generated_rows=len(standard_rows),
        processed_sheets=processed_sheets,
        skipped_sheets=skipped_sheets,
        rows_by_sheet=rows_by_sheet,
        output_sheet=output_sheet,
    )
    return output_bytes, summary


def generate_sumun_template_from_excel(
    excel_bytes: bytes,
    *,
    source_name: str = "",
    area: str | None = None,
    grade: int | None = None,
    level: str = "Secundaria",
    course_code: str | None = None,
    sheet_names: list[str] | None = None,
) -> tuple[bytes, SumunTemplateSummary]:
    """Build the SUMUN load template from an uploaded matrix workbook.

    The reader scans every visible worksheet and processes sheets that look like
    SUMUN matrices. The current input format can keep every itinerary in one
    sheet, separated by a numeric itinerary column.
    """
    workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
    inference_sheet_names = sheet_names or workbook.sheetnames
    inferred_code, inferred_grade = _infer_course_code_and_grade(
        source_name,
        inference_sheet_names,
    )
    final_course_code = (course_code or inferred_code or "CT").strip().upper()
    final_grade = int(grade or inferred_grade or 1)
    final_level = (level or "Secundaria").strip() or "Secundaria"
    final_area = (
        (area or "").strip()
        or DEFAULT_AREA_BY_CODE.get(final_course_code)
        or DEFAULT_AREA_BY_CODE.get(final_course_code.replace("CCT", "CT"))
        or DEFAULT_AREA_BY_CODE["CT"]
    )
    level_code = DEFAULT_LEVEL_BY_NAME.get(final_level, _level_code(final_level))
    prefix = f"{level_code}{final_course_code}{final_grade}"

    rows, build_stats = _build_output_rows(
        workbook=workbook,
        area=final_area,
        grade=final_grade,
        level=final_level,
        prefix=prefix,
        sheet_names=sheet_names,
    )
    if not rows:
        inspected = inspect_sumun_workbook_sheets(excel_bytes)
        selected = set(sheet_names or [])
        inspected = [
            item for item in inspected if not selected or item.sheet_name in selected
        ]
        detail = "; ".join(
            f"{item.index}-{item.sheet_name}: {item.reason or 'sin detalle'}"
            for item in inspected[:8]
        )
        raise ValueError(
            "No se pudieron generar filas de la plantilla SUMUN. "
            f"Revision: {detail or 'sin hojas seleccionadas'}"
        )

    output_sheet = _safe_sheet_title(f"{final_course_code}{final_grade}{level_code}")
    output_bytes = _write_template_workbook(rows, output_sheet=output_sheet)
    summary = SumunTemplateSummary(
        prefix=prefix,
        area=final_area,
        grade=final_grade,
        level=final_level,
        output_sheet=output_sheet,
        generated_rows=len(rows),
        macro_count=build_stats["macro_count"],
        micro_count=build_stats["micro_count"],
        unique_micro_count=build_stats["unique_micro_count"],
        processed_sheets=build_stats["processed_sheets"],
        skipped_sheets=build_stats["skipped_sheets"],
        rows_by_sheet=build_stats["rows_by_sheet"],
        specific_rows_by_itinerary=build_stats["specific_rows_by_itinerary"],
        specific_rows_by_knowledge=build_stats["specific_rows_by_knowledge"],
        nonnumber_station_rows=build_stats["nonnumber_station_rows"],
    )
    return output_bytes, summary


def generate_sumun_template_file(
    input_path: str | Path,
    output_path: str | Path,
    *,
    area: str | None = None,
    grade: int | None = None,
    level: str = "Secundaria",
    course_code: str | None = None,
    sheet_names: list[str] | None = None,
) -> SumunTemplateSummary:
    input_file = Path(input_path)
    output_file = Path(output_path)
    output_bytes, summary = generate_sumun_template_from_excel(
        input_file.read_bytes(),
        source_name=input_file.name,
        area=area,
        grade=grade,
        level=level,
        course_code=course_code,
        sheet_names=sheet_names,
    )
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_bytes(output_bytes)
    return summary


def _build_output_rows(
    *,
    workbook,
    area: str,
    grade: int,
    level: str,
    prefix: str,
    sheet_names: list[str] | None = None,
) -> tuple[list[list[Any]], dict[str, Any]]:
    macro_ids: dict[str, int] = {}
    micro_ids: dict[str, int] = {}
    specific_counters: defaultdict[tuple[int, int], int] = defaultdict(int)
    last_station_by_itinerary: dict[int, tuple[int, str]] = {}
    station_aliases: dict[tuple[int, str], tuple[int, str]] = {}
    next_station_number_by_itinerary: dict[int, int] = {}
    itinerary_titles = _collect_itinerary_titles(workbook, sheet_names=sheet_names)
    station_titles = _collect_station_titles(
        workbook,
        sheet_names=sheet_names,
        itinerary_titles=itinerary_titles,
    )
    output_rows: list[list[Any]] = []
    processed_sheets: list[str] = []
    skipped_sheets: list[str] = []
    rows_by_sheet: dict[str, int] = {}
    specific_rows_by_itinerary: defaultdict[tuple[int, str], int] = defaultdict(int)
    specific_rows_by_knowledge: defaultdict[tuple[int, str, int, str, str], int] = defaultdict(int)
    nonnumber_station_rows: list[str] = []
    micro_row_count = 0
    selected_sheets = set(sheet_names or [])

    for ws in workbook.worksheets:
        if selected_sheets and ws.title not in selected_sheets:
            continue
        if ws.sheet_state != "visible":
            skipped_sheets.append(ws.title)
            continue
        values, merged_sources = _fill_merged_values_with_sources(ws)
        layout = _detect_matrix_layout(ws.title, values)
        if layout is None:
            skipped_sheets.append(ws.title)
            continue

        before_count = len(output_rows)
        sheet_itinerary = _infer_sheet_itinerary_context(values, layout, ws.title)
        current_itinerary = sheet_itinerary
        for row_number in range(layout.data_start_row, len(values) + 1):
            row = values[row_number - 1]
            cell_itinerary = _parse_row_itinerary(row, layout)
            if cell_itinerary:
                current_itinerary = cell_itinerary
            itinerary = _resolve_itinerary_context(
                cell_itinerary,
                current_itinerary,
                itinerary_titles=itinerary_titles,
            )
            if not itinerary:
                continue
            itinerary_number, itinerary_name = itinerary
            itinerary_display_name = _itinerary_output_name(itinerary_number, itinerary_name)

            station_value = _column_group_value(
                row,
                layout.station_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            )
            station_context, _station_source = _resolve_station_context(
                station_value,
                itinerary_number,
                last_station_by_itinerary=last_station_by_itinerary,
                station_aliases=station_aliases,
                next_station_number_by_itinerary=next_station_number_by_itinerary,
                station_titles=station_titles,
            )

            competence = _clean_text(
                _column_group_value(
                    row,
                    layout.competence_cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                )
            )
            macro = _clean_text(
                _column_group_value(
                    row,
                    layout.macro_cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                ),
                strip_bullet=True,
            )
            micro = _clean_text(
                _column_group_value(
                    row,
                    layout.micro_cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                ),
                strip_bullet=True,
            )
            if not competence or not macro or not micro:
                continue

            process_items = [
                (process, skill)
                for process, cols in layout.process_cols.items()
                for skill in _specific_skill_group_values(
                    row,
                    cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                )
            ]
            if not process_items:
                continue
            if not station_context:
                nonnumber_station_rows.append(f"{ws.title}!R{row_number}")
                continue

            micro_row_count += 1
            if macro not in macro_ids:
                macro_ids[macro] = len(macro_ids) + 1
            if micro not in micro_ids:
                micro_ids[micro] = len(micro_ids) + 1

            station_number, station_name = station_context
            knowledge = _normalize_knowledge_text(
                _column_group_value(
                    row,
                    layout.knowledge_cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                )
            ) or ""

            for process, skill in process_items:
                specific_counters[(itinerary_number, station_number)] += 1
                specific_number = specific_counters[(itinerary_number, station_number)]
                output_id = (
                    f"{prefix}_I{itinerary_number:02d}_E{station_number:02d}"
                    f"_MA{macro_ids[macro]:02d}_MI{micro_ids[micro]:02d}"
                    f"_ME{specific_number:02d}"
                )
                output_rows.append(
                    [
                        output_id,
                        area,
                        grade,
                        level,
                        None,
                        itinerary_number,
                        itinerary_display_name,
                        competence,
                        macro_ids[macro],
                        macro,
                        micro_ids[micro],
                        micro,
                        station_number,
                        station_name,
                        knowledge,
                        None,
                        specific_number,
                        skill,
                        process,
                        None,
                    ]
                )
                specific_rows_by_itinerary[(itinerary_number, itinerary_display_name)] += 1
                specific_rows_by_knowledge[
                    (
                        itinerary_number,
                        itinerary_display_name,
                        station_number,
                        station_name,
                        knowledge,
                    )
                ] += 1

        generated_for_sheet = len(output_rows) - before_count
        if generated_for_sheet:
            processed_sheets.append(ws.title)
            rows_by_sheet[ws.title] = generated_for_sheet
        else:
            skipped_sheets.append(ws.title)

    stats = {
        "macro_count": len(macro_ids),
        "micro_count": micro_row_count,
        "unique_micro_count": len(micro_ids),
        "processed_sheets": processed_sheets,
        "skipped_sheets": skipped_sheets,
        "rows_by_sheet": rows_by_sheet,
        "specific_rows_by_itinerary": [
            {
                "itinerary_number": itinerary_number,
                "itinerary": itinerary_name,
                "specific_rows": total_rows,
            }
            for (itinerary_number, itinerary_name), total_rows in specific_rows_by_itinerary.items()
        ],
        "specific_rows_by_knowledge": [
            {
                "itinerary_number": itinerary_number,
                "itinerary": itinerary_name,
                "station_number": station_number,
                "station": station_name,
                "knowledge": knowledge,
                "specific_rows": total_rows,
            }
            for (
                itinerary_number,
                itinerary_name,
                station_number,
                station_name,
                knowledge,
            ), total_rows in specific_rows_by_knowledge.items()
        ],
        "nonnumber_station_rows": nonnumber_station_rows,
    }
    return output_rows, stats


def _detect_matrix_layout(sheet_name: str, values: list[list[Any]]) -> MatrixLayout | None:
    max_rows = min(len(values), 40)
    for header_row in range(1, max_rows + 1):
        header_values = values[header_row - 1]
        header_candidates = [(header_row, header_values)]
        for row_number in range(header_row + 1, min(header_row + 3, len(values)) + 1):
            header_candidates.append(
                (
                    row_number,
                    _combine_header_rows(header_values, values[row_number - 1]),
                )
            )

        context_groups: dict[str, tuple[int, ...]] | None = None
        context_header_row = header_row
        for row_number, candidate_values in header_candidates:
            context_groups = _find_context_column_groups(candidate_values)
            if context_groups:
                context_header_row = row_number
                break
        if not context_groups:
            continue

        process_cols: dict[str, tuple[int, ...]] = {}
        process_header_row = header_row
        for row_number, candidate_values in header_candidates:
            candidate_process_cols = _find_process_column_groups(candidate_values)
            if len(candidate_process_cols) > len(process_cols):
                process_cols = candidate_process_cols
                process_header_row = row_number

        if len(process_cols) >= 2:
            ordered_process_cols = {
                process: process_cols[process]
                for process in PROCESS_NAMES
                if process in process_cols
            }
            return MatrixLayout(
                sheet_name=sheet_name,
                header_row=header_row,
                data_start_row=max(context_header_row, process_header_row) + 1,
                itinerary_cols=context_groups["itinerary"],
                itinerary_number_cols=context_groups.get("itinerary_number"),
                competence_cols=context_groups["competence"],
                macro_cols=context_groups["macro"],
                micro_cols=context_groups["micro"],
                station_cols=context_groups["station"],
                knowledge_cols=context_groups["knowledge"],
                process_cols=ordered_process_cols,
                evidence_cols=context_groups.get("evidence"),
                instrument_cols=context_groups.get("instrument"),
                criteria_cols=context_groups.get("criteria"),
            )

    return _fallback_layout_from_data(sheet_name, values)


def _inspection_reason(
    layout: MatrixLayout | None,
    scan_stats: SheetScanStats,
) -> str:
    empty_fields_detail = _format_empty_field_detail(scan_stats.empty_field_rows)
    if layout is None:
        return "No se reconocieron encabezados de matriz o columnas A:L con itinerarios."
    if scan_stats.estimated_rows <= 0 and scan_stats.missing_station_rows:
        rows = ", ".join(str(row) for row in scan_stats.missing_station_rows[:3])
        if len(scan_stats.missing_station_rows) > 3:
            rows = f"{rows}, ..."
        reason = (
            "Se reconocio la estructura y microhabilidades, pero faltan estaciones "
            "identificables para generar filas. "
            "La columna ESTACION puede venir numerada o con nombre; "
            "las filas sin dato util no se pueden procesar "
            f"(primeras filas: {rows})."
        )
        if empty_fields_detail:
            reason = f"{reason} {empty_fields_detail}"
        return reason
    if scan_stats.estimated_rows <= 0:
        reason = (
            "Se reconocio la estructura, pero no se encontraron microhabilidades "
            "en RECORDAR/COMPRENDER/APLICAR/ANALIZAR/EVALUAR/CREAR."
        )
        if empty_fields_detail:
            reason = f"{reason} {empty_fields_detail}"
        return reason
    process_list = ", ".join(layout.process_cols.keys())
    reason = (
        f"Matriz detectada. Encabezado fila {layout.header_row}; "
        f"procesos: {process_list}; filas estimadas: {scan_stats.estimated_rows}."
    )
    if empty_fields_detail:
        reason = f"{reason} {empty_fields_detail}"
    return reason


def _estimate_sheet_rows(values: list[list[Any]], layout: MatrixLayout | None) -> int:
    return _scan_sheet_rows(values, layout).estimated_rows


def _format_empty_field_detail(
    empty_field_rows: tuple[tuple[int, tuple[str, ...]], ...],
    *,
    limit: int = 5,
) -> str:
    if not empty_field_rows:
        return ""
    parts: list[str] = []
    for row_number, fields in empty_field_rows[:limit]:
        field_txt = ", ".join(str(field) for field in fields if str(field).strip())
        if not field_txt:
            continue
        parts.append(f"R{row_number}: {field_txt}")
    if not parts:
        return ""
    extra = len(empty_field_rows) - len(parts)
    detail = "Campos vacios detectados: " + "; ".join(parts)
    if extra > 0:
        detail += f"; +{extra} filas mas"
    return detail + "."


def _scan_sheet_rows(
    values: list[list[Any]],
    layout: MatrixLayout | None,
    *,
    merged_sources: list[list[tuple[int, int]]] | None = None,
) -> SheetScanStats:
    if layout is None:
        return SheetScanStats(0, ())
    estimated = 0
    missing_station_rows: list[int] = []
    empty_field_rows: list[tuple[int, tuple[str, ...]]] = []
    last_station_by_itinerary: dict[int, tuple[int, str]] = {}
    station_aliases: dict[tuple[int, str], tuple[int, str]] = {}
    next_station_number_by_itinerary: dict[int, int] = {}
    itinerary_titles: dict[int, str] = {}
    sheet_itinerary = _infer_sheet_itinerary_context(values, layout, layout.sheet_name)
    current_itinerary = sheet_itinerary
    for row_number in range(layout.data_start_row, len(values) + 1):
        row = values[row_number - 1]
        if not _row_has_source_content(row, layout, merged_sources, row_number):
            continue
        cell_itinerary = _parse_row_itinerary(row, layout)
        if cell_itinerary:
            current_itinerary = cell_itinerary
        itinerary = _resolve_itinerary_context(
            cell_itinerary,
            current_itinerary,
            itinerary_titles=itinerary_titles,
        )
        if not itinerary:
            continue
        itinerary_number, _ = itinerary
        competence = _clean_text(
            _column_group_value(
                row,
                layout.competence_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            )
        )
        macro = _clean_text(
            _column_group_value(
                row,
                layout.macro_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            ),
            strip_bullet=True,
        )
        micro = _clean_text(
            _column_group_value(
                row,
                layout.micro_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            ),
            strip_bullet=True,
        )
        station_raw_value = _column_group_value(
            row,
            layout.station_cols,
            merged_sources=merged_sources,
            row_number=row_number,
        )
        station_value = _clean_text(station_raw_value)
        knowledge = _normalize_knowledge_text(
            _column_group_value(
                row,
                layout.knowledge_cols,
                merged_sources=merged_sources,
                row_number=row_number,
            )
        )
        station_context, _station_source = _resolve_station_context(
            station_raw_value,
            itinerary_number,
            last_station_by_itinerary=last_station_by_itinerary,
            station_aliases=station_aliases,
            next_station_number_by_itinerary=next_station_number_by_itinerary,
        )
        process_count = sum(
            len(
                _specific_skill_group_values(
                    row,
                    cols,
                    merged_sources=merged_sources,
                    row_number=row_number,
                )
            )
            for cols in layout.process_cols.values()
        )
        missing_fields: list[str] = []
        if not competence:
            missing_fields.append("COMPETENCIA")
        if not macro:
            missing_fields.append("MACROHABILIDAD")
        if not micro:
            missing_fields.append("MICROHABILIDAD")
        if not station_value:
            missing_fields.append("ESTACION")
        if not knowledge:
            missing_fields.append("CONOCIMIENTOS")
        if not process_count:
            missing_fields.append("MICROHABILIDAD ESPECIFICA")
        if missing_fields:
            empty_field_rows.append((row_number, tuple(missing_fields)))
        if not competence or not macro or not micro:
            continue
        if not process_count:
            continue

        if not station_context:
            missing_station_rows.append(row_number)
            continue

        estimated += process_count
    return SheetScanStats(
        estimated,
        tuple(missing_station_rows),
        tuple(empty_field_rows),
    )


def _find_context_column_groups(
    header_values: list[Any],
) -> dict[str, tuple[int, ...]] | None:
    normalized = {idx: _label_key(value) for idx, value in enumerate(header_values, start=1)}

    def find(*aliases: str, avoid_hash: bool = False, exclude: tuple[str, ...] = ()) -> int | None:
        alias_set = {_label_key(alias) for alias in aliases}
        excluded = tuple(_label_key(item) for item in exclude)
        for idx, label in normalized.items():
            if not label:
                continue
            if avoid_hash and str(label).startswith("#"):
                continue
            if any(item and item in label for item in excluded):
                continue
            if label in alias_set or any(alias and alias in label for alias in alias_set):
                return idx
        return None

    itinerary_number_col = next(
        (
            idx
            for idx, label in normalized.items()
            if _is_itinerary_number_label(label)
        ),
        None,
    )
    itinerary_col = next(
        (
            idx
            for idx, label in normalized.items()
            if idx != itinerary_number_col and _is_itinerary_name_label(label)
        ),
        None,
    )
    if itinerary_col is None:
        itinerary_col = itinerary_number_col or find("ITINERARIO")

    context_cols = {
        "itinerary": itinerary_col,
        "itinerary_number": itinerary_number_col,
        "competence": find("COMPETENCIA"),
        "macro": find(
            "MACROHABILIDAD",
            "MACRO HABILIDAD",
            avoid_hash=True,
            exclude=("MICROHABILIDAD", "MICRO HABILIDAD"),
        ),
        "micro": find(
            "MICROHABILIDAD",
            "MICRO HABILIDAD",
            avoid_hash=True,
            exclude=("ESPECIFICA", "ESPECIFICO"),
        ),
        "station": find("ESTACION", "ESTACI\u00d3N", avoid_hash=True),
        "knowledge": find("CONOCIMIENTOS"),
        "evidence": find("EVIDENCIAS", "EVIDENCIA"),
        "instrument": find(
            "INSTRUMENTOS DE EVALUACION",
            "INSTRUMENTO DE EVALUACION",
            "INSTRUMENTOS",
        ),
        "criteria": find(
            "CRITERIOS DE EVALUACION",
            "CRITERIO DE EVALUACION",
            "CRITERIOS",
        ),
    }
    required_context_cols = {
        key: value
        for key, value in context_cols.items()
        if key
        not in {
            "itinerary_number",
            "evidence",
            "instrument",
            "criteria",
        }
    }
    if all(required_context_cols.values()):
        return {
            key: _header_column_group(header_values, int(value))
            for key, value in context_cols.items()
            if value is not None
        }
    return None


def _find_process_column_groups(
    header_values: list[Any],
) -> dict[str, tuple[int, ...]]:
    result: dict[str, tuple[int, ...]] = {}
    for col_idx, value in enumerate(header_values, start=1):
        process = _process_from_label(_label_key(value))
        if process and process not in result:
            result[process] = _header_column_group(header_values, col_idx)
    return result


def _header_column_group(
    header_values: list[Any],
    target_col: int,
) -> tuple[int, ...]:
    blocks = _header_column_blocks(header_values)
    for start_col, end_col in blocks:
        if start_col <= target_col <= end_col:
            return tuple(range(start_col, end_col + 1))
    return (target_col,)


def _header_column_blocks(header_values: list[Any]) -> list[tuple[int, int]]:
    anchors: list[tuple[int, str]] = []
    for col_idx, value in enumerate(header_values, start=1):
        label = _label_key(value)
        if not label:
            continue
        if anchors and anchors[-1][1] == label:
            continue
        anchors.append((col_idx, label))

    blocks: list[tuple[int, int]] = []
    for index, (start_col, _label) in enumerate(anchors):
        if index + 1 < len(anchors):
            end_col = anchors[index + 1][0] - 1
        else:
            end_col = len(header_values)
        blocks.append((start_col, max(start_col, end_col)))
    return blocks


def _is_itinerary_number_label(label: str) -> bool:
    clean = str(label or "").strip()
    if not clean:
        return False
    return bool(
        clean == "# ITINERARIO"
        or re.fullmatch(
            r"(?:N|NO|NRO|NUM|NUMERO)(?: DE)? ITINERARIO",
            clean,
        )
    )


def _is_itinerary_name_label(label: str) -> bool:
    clean = str(label or "").strip()
    if not clean or _is_itinerary_number_label(clean):
        return False
    return clean in {"ITINERARIO", "NOMBRE ITINERARIO", "NOMBRE DEL ITINERARIO"}


def _row_has_source_content(
    row: list[Any],
    layout: MatrixLayout,
    merged_sources: list[list[tuple[int, int]]] | None,
    row_number: int,
) -> bool:
    relevant_cols = [
        *layout.itinerary_cols,
        *(layout.itinerary_number_cols or ()),
        *layout.competence_cols,
        *layout.macro_cols,
        *layout.micro_cols,
        *layout.station_cols,
        *layout.knowledge_cols,
        *(col for cols in layout.process_cols.values() for col in cols),
    ]
    for col in relevant_cols:
        source = _cell_source(merged_sources, row_number, col)
        if source is not None and source != (row_number, col):
            continue
        if _clean_text(_cell(row, col)) is not None:
            return True
    return False


def _row_has_standardizable_content(
    row: list[Any],
    layout: MatrixLayout,
    *,
    merged_sources: list[list[tuple[int, int]]] | None,
    row_number: int,
) -> bool:
    relevant_cols = [
        *layout.itinerary_cols,
        *(layout.itinerary_number_cols or ()),
        *layout.competence_cols,
        *layout.macro_cols,
        *layout.micro_cols,
        *layout.station_cols,
        *layout.knowledge_cols,
        *(col for cols in layout.process_cols.values() for col in cols),
        *(layout.evidence_cols or ()),
        *(layout.instrument_cols or ()),
        *(layout.criteria_cols or ()),
    ]
    for col in relevant_cols:
        source = _cell_source(merged_sources, row_number, col)
        if source is not None and source != (row_number, col):
            continue
        if _clean_text(_cell(row, col)) is not None:
            return True
    return False


def _fallback_layout_from_data(sheet_name: str, values: list[list[Any]]) -> MatrixLayout | None:
    sheet_itinerary = _parse_itinerary_from_sheet_name(sheet_name)
    for row_number, row in enumerate(values, start=1):
        if (_parse_itinerary(_cell(row, 1)) or sheet_itinerary) and any(
            _cell(row, col) for col in range(7, 13)
        ):
            return MatrixLayout(
                sheet_name=sheet_name,
                header_row=max(1, row_number - 1),
                data_start_row=row_number,
                itinerary_cols=(1,),
                itinerary_number_cols=None,
                competence_cols=(2,),
                macro_cols=(3,),
                micro_cols=(4,),
                station_cols=(5,),
                knowledge_cols=(6,),
                process_cols={
                    process: (idx,)
                    for process, idx in zip(PROCESS_NAMES, range(7, 13))
                },
            )
    return None


def _write_standard_workbook(
    rows: list[list[Any]],
    *,
    output_sheet: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = output_sheet
    ws.append(STANDARD_HEADERS)
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(STANDARD_HEADERS))}{ws.max_row}"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 34,
        "B": 34,
        "C": 54,
        "D": 48,
        "E": 48,
        "F": 54,
        "G": 48,
        "H": 48,
        "I": 48,
        "J": 48,
        "K": 48,
        "L": 48,
        "M": 48,
        "N": 48,
        "O": 48,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    table_ref = f"A1:{get_column_letter(len(STANDARD_HEADERS))}{ws.max_row}"
    table = Table(displayName="TablaSUMUNEstandar", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_template_workbook(rows: list[list[Any]], *, output_sheet: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = output_sheet
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for letter, width in DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 72

    table_ref = f"A1:{get_column_letter(len(OUTPUT_HEADERS))}{ws.max_row}"
    table = Table(displayName="Tabla1", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    _add_datos_sheet(wb)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _add_datos_sheet(wb) -> None:
    ws = wb.create_sheet("Datos")
    rows = [
        [
            "\u00c1reas",
            "Abreviatura de grado",
            "Nivel",
            "Abreviatura de nivel",
            "Grado",
            "Jerarqu\u00eda",
            "Jerarqu\u00eda ejemplo",
        ],
        [
            "Comunicaci\u00f3n",
            "CO",
            "Primaria",
            "PRI",
            1,
            "\u00c1reas + Grado + Nivel",
            "Ciencia y Tecnolog\u00eda 1 SEC",
        ],
        [
            "Matem\u00e1tica",
            "MA",
            "Secundaria",
            "SEC",
            2,
            'ID Macrohabilidad + " - " + Macrohabilidad',
            "1 - Comprende y usa conocimientos sobre los seres vivos, materia y energia, biodiversidad, Tierra y universo.",
        ],
        [
            "Ciencias sociales",
            "CCSS",
            None,
            None,
            3,
            '"E" + ID Estaci\u00f3n + Estaci\u00f3n',
            "E1 - Los seres vivos y la c\u00e9lula",
        ],
        [
            "Ciencia y Tecnolog\u00eda",
            "CT",
            None,
            None,
            4,
            'Proceso cognitivo + " - " + Microhabilidad',
            "RECORDAR - Explicar que las sustancias inorganicas y biomoleculas conforman la estructura de la celula.",
        ],
        [
            "Personal Social",
            "PS",
            None,
            None,
            5,
            "C\u00f3digo de Nanohabilidad + Nanohabilidad",
            "Reconocer las caracteristicas de los seres vivos y su relacion con la estructura celular.",
        ],
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = {"A": 26, "B": 20, "C": 16, "D": 20, "E": 10, "F": 48, "G": 90}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _fill_merged_values_with_sources(ws) -> tuple[list[list[Any]], list[list[tuple[int, int]]]]:
    values = [
        [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        for row in range(1, ws.max_row + 1)
    ]
    sources = [
        [(row, col) for col in range(1, ws.max_column + 1)]
        for row in range(1, ws.max_row + 1)
    ]
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        source = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[row - 1][col - 1] = value
                sources[row - 1][col - 1] = source
    return values, sources


def _fill_merged_values(ws) -> list[list[Any]]:
    values, _sources = _fill_merged_values_with_sources(ws)
    return values


def _cell(row: list[Any], col: int) -> Any:
    if col <= 0 or col > len(row):
        return None
    return row[col - 1]


def _cell_source(
    sources: list[list[tuple[int, int]]] | None,
    row_number: int,
    col: int,
) -> tuple[int, int] | None:
    if sources is None or row_number <= 0 or row_number > len(sources):
        return None
    row = sources[row_number - 1]
    if col <= 0 or col > len(row):
        return None
    return row[col - 1]


def _column_group_values(
    row: list[Any],
    cols: tuple[int, ...],
    *,
    merged_sources: list[list[tuple[int, int]]] | None = None,
    row_number: int | None = None,
) -> tuple[Any, ...]:
    result: list[Any] = []
    seen_sources: set[tuple[int, int]] = set()
    seen_texts: set[str] = set()
    for col in cols:
        source = (
            _cell_source(merged_sources, row_number, col)
            if row_number is not None
            else None
        )
        if source is not None and source in seen_sources:
            continue
        if source is not None:
            seen_sources.add(source)

        value = _cell(row, col)
        clean = _clean_text(value)
        if not clean or clean in seen_texts:
            continue
        seen_texts.add(clean)
        result.append(value)
    return tuple(result)


def _column_group_value(
    row: list[Any],
    cols: tuple[int, ...],
    *,
    merged_sources: list[list[tuple[int, int]]] | None = None,
    row_number: int | None = None,
) -> Any:
    values = _column_group_values(
        row,
        cols,
        merged_sources=merged_sources,
        row_number=row_number,
    )
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    return "\n".join(str(value).strip() for value in values)


def _standard_group_value(
    row: list[Any],
    cols: tuple[int, ...],
    *,
    merged_sources: list[list[tuple[int, int]]] | None,
    row_number: int,
) -> str | None:
    return _clean_text(
        _column_group_value(
            row,
            cols,
            merged_sources=merged_sources,
            row_number=row_number,
        )
    )


def _standard_optional_group_value(
    row: list[Any],
    cols: tuple[int, ...] | None,
    *,
    merged_sources: list[list[tuple[int, int]]] | None,
    row_number: int,
) -> str | None:
    if not cols:
        return None
    return _standard_group_value(
        row,
        cols,
        merged_sources=merged_sources,
        row_number=row_number,
    )


def _standard_itinerary_value(
    itinerary_number: int | None,
    itinerary_title: str | None,
) -> str | None:
    title = _clean_text(itinerary_title)
    if itinerary_number is None:
        return title
    if _is_meaningful_itinerary_title(itinerary_number, title):
        return f"Itinerario {itinerary_number}. {title}"
    return str(itinerary_number)


def _standard_station_value(
    station_number: int,
    station_name: str,
) -> str:
    name = _normalize_station_text(station_name)
    if name:
        return f"E{station_number} - {name}"
    return f"E{station_number}"


def _clean_text(value: Any, *, strip_bullet: bool = False) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if strip_bullet:
        text = re.sub(r"^[\s\-\u2022]+", "", text).strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text or None


def _lstrip_to_first_letter(text: str) -> str:
    for index, char in enumerate(str(text)):
        if char.isalpha():
            return str(text)[index:].strip()
    return ""


def _starts_with_lowercase_letter(text: str) -> bool:
    for char in str(text).strip():
        if char.isalpha():
            return char.islower()
    return False


def _normalize_station_text(value: Any) -> str | None:
    if value is None:
        return None
    raw_text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw_text:
        return None

    candidate = raw_text
    if ":" in candidate:
        _prefix, suffix = candidate.split(":", 1)
        if suffix.strip():
            candidate = suffix

    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in candidate.split("\n") if line.strip()]
    if not lines:
        return None

    text = " ".join(lines).strip()
    text = re.sub(r"[ \t]+", " ", text)
    if ("\n" in raw_text or ":" in raw_text) and text and not re.search(r"[.!?]$", text):
        text += "."
    return text or None


def _knowledge_line_separator(current_text: str, next_line: str, *, had_blank_line: bool) -> str:
    current = current_text.rstrip()
    if not current:
        return ""
    last_char = current[-1]
    if last_char in ".!?":
        return " "
    if had_blank_line:
        return ". "
    if last_char in ",;:(" or _starts_with_lowercase_letter(next_line):
        return " "
    return ". "


def _normalize_knowledge_text(value: Any) -> str | None:
    if value is None:
        return None
    raw_text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw_text:
        return None

    result = ""
    had_blank_line = False
    for raw_line in raw_text.split("\n"):
        line = re.sub(r"[ \t]+", " ", raw_line).strip()
        if not line:
            had_blank_line = True
            continue
        line = _lstrip_to_first_letter(line)
        if not line:
            continue
        if not result:
            result = line
        else:
            result += _knowledge_line_separator(result, line, had_blank_line=had_blank_line)
            result += line
        had_blank_line = False

    result = re.sub(r"[ \t]+", " ", result).strip()
    result = _lstrip_to_first_letter(result)
    return result or None


def _specific_skill_cell_values(
    value: Any,
    *,
    source: tuple[int, int] | None = None,
    row_number: int | None = None,
    col: int | None = None,
) -> tuple[str, ...]:
    if (
        source is not None
        and row_number is not None
        and col is not None
        and source != (row_number, col)
    ):
        return ()
    if value is None:
        return ()
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ()
    blocks = [block.strip() for block in re.split(r"\n[ \t]*\n+", text)]
    return tuple(block for block in blocks if block)


def _specific_skill_group_values(
    row: list[Any],
    cols: tuple[int, ...],
    *,
    merged_sources: list[list[tuple[int, int]]] | None = None,
    row_number: int,
) -> tuple[str, ...]:
    result: list[str] = []
    seen_skills: set[str] = set()
    for col in cols:
        for skill in _specific_skill_cell_values(
            _cell(row, col),
            source=_cell_source(merged_sources, row_number, col),
            row_number=row_number,
            col=col,
        ):
            if skill in seen_skills:
                continue
            seen_skills.add(skill)
            result.append(skill)
    return tuple(result)


def _parse_itinerary(value: Any) -> tuple[int, str] | None:
    text = _clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.0)?", text):
        number = int(float(text))
        return number, str(number)
    match = re.search(r"Itinerario\s*(\d+)\s*\.?\s*(.*)", text, flags=re.I)
    if not match:
        return None
    return int(match.group(1)), match.group(2).strip()


def _parse_row_itinerary(
    row: list[Any],
    layout: MatrixLayout,
) -> tuple[int, str] | None:
    itinerary_values = _column_group_values(row, layout.itinerary_cols)
    number_values = (
        _column_group_values(row, layout.itinerary_number_cols)
        if layout.itinerary_number_cols
        else ()
    )
    parsed_values = [
        parsed
        for value in (*number_values, *itinerary_values)
        if (parsed := _parse_itinerary(value)) is not None
    ]
    if not parsed_values:
        return None

    itinerary_number = parsed_values[0][0]
    title_candidates: list[str | None] = [
        parsed[1]
        for parsed in parsed_values
        if parsed[0] == itinerary_number
    ]
    for value in itinerary_values:
        clean = _clean_text(value)
        parsed = _parse_itinerary(value)
        if clean and parsed is None:
            title_candidates.append(clean)
    return (
        itinerary_number,
        _best_itinerary_title(
            itinerary_number,
            *title_candidates,
        ),
    )


def _is_meaningful_itinerary_title(number: int | None, title: str | None) -> bool:
    text = _clean_text(title)
    if not text:
        return False
    if number is not None and text == str(number):
        return False
    normalized = _strip_accents(text).upper()
    if number is not None and re.fullmatch(
        rf"(?:ITI|ITINERARIO|HITO)\s*0*{number}",
        normalized,
    ):
        return False
    if normalized in GENERIC_ITINERARY_TITLES:
        return False
    return True


def _best_itinerary_title(number: int, *candidates: str | None) -> str:
    best: str | None = None
    for candidate in candidates:
        clean = _clean_text(candidate)
        if not _is_meaningful_itinerary_title(number, clean):
            continue
        if best is None or len(clean or "") > len(best):
            best = clean
    if best:
        return best
    return str(number)


def _register_itinerary_title(
    titles_by_number: dict[int, str],
    number: int,
    title: str | None,
) -> None:
    clean = _clean_text(title)
    if not _is_meaningful_itinerary_title(number, clean):
        return
    current = titles_by_number.get(number)
    if current is None or len(clean or "") > len(current):
        titles_by_number[number] = str(clean)


def _collect_itinerary_titles(workbook, *, sheet_names: list[str] | None = None) -> dict[int, str]:
    titles_by_number: dict[int, str] = {}
    selected_sheets = set(sheet_names or [])
    for ws in workbook.worksheets:
        if selected_sheets and ws.title not in selected_sheets:
            continue
        if ws.sheet_state != "visible":
            continue
        values = _fill_merged_values(ws)
        layout = _detect_matrix_layout(ws.title, values)
        if layout is None:
            continue

        sheet_itinerary = _infer_sheet_itinerary_context(values, layout, ws.title)
        if sheet_itinerary:
            _register_itinerary_title(
                titles_by_number,
                sheet_itinerary[0],
                sheet_itinerary[1],
            )

        for row_number in range(layout.data_start_row, len(values) + 1):
            parsed = _parse_row_itinerary(values[row_number - 1], layout)
            if parsed:
                _register_itinerary_title(titles_by_number, parsed[0], parsed[1])
    return titles_by_number


def _register_station_title(
    titles_by_key: dict[tuple[int, int], str],
    itinerary_number: int,
    station_number: int,
    title: str | None,
) -> None:
    clean = _normalize_station_text(title)
    if not clean:
        return
    key = (itinerary_number, station_number)
    current = titles_by_key.get(key)
    if current is None or len(clean) > len(current):
        titles_by_key[key] = clean


def _collect_station_titles(
    workbook,
    *,
    sheet_names: list[str] | None = None,
    itinerary_titles: dict[int, str] | None = None,
) -> dict[tuple[int, int], str]:
    titles_by_key: dict[tuple[int, int], str] = {}
    selected_sheets = set(sheet_names or [])
    resolved_itinerary_titles = itinerary_titles or _collect_itinerary_titles(
        workbook,
        sheet_names=sheet_names,
    )
    for ws in workbook.worksheets:
        if selected_sheets and ws.title not in selected_sheets:
            continue
        if ws.sheet_state != "visible":
            continue
        values = _fill_merged_values(ws)
        layout = _detect_matrix_layout(ws.title, values)
        if layout is None:
            continue

        sheet_itinerary = _infer_sheet_itinerary_context(values, layout, ws.title)
        current_itinerary = sheet_itinerary
        for row_number in range(layout.data_start_row, len(values) + 1):
            row = values[row_number - 1]
            cell_itinerary = _parse_row_itinerary(row, layout)
            if cell_itinerary:
                current_itinerary = cell_itinerary
            itinerary = _resolve_itinerary_context(
                cell_itinerary,
                current_itinerary,
                itinerary_titles=resolved_itinerary_titles,
            )
            if not itinerary:
                continue
            parsed_station = _parse_station(
                _column_group_value(row, layout.station_cols)
            )
            if not parsed_station:
                continue
            station_number, station_name = parsed_station
            _register_station_title(
                titles_by_key,
                itinerary[0],
                station_number,
                station_name,
            )
    return titles_by_key


def _resolve_itinerary_context(
    cell_itinerary: tuple[int, str] | None,
    sheet_itinerary: tuple[int, str] | None,
    *,
    itinerary_titles: dict[int, str],
) -> tuple[int, str] | None:
    if cell_itinerary:
        itinerary_number = cell_itinerary[0]
    elif sheet_itinerary:
        itinerary_number = sheet_itinerary[0]
    else:
        return None
    return (
        itinerary_number,
        _best_itinerary_title(
            itinerary_number,
            itinerary_titles.get(itinerary_number),
            cell_itinerary[1] if cell_itinerary else None,
            sheet_itinerary[1] if sheet_itinerary else None,
        ),
    )


def _itinerary_output_name(number: int, title: str | None) -> str | None:
    clean = _clean_text(title)
    if not _is_meaningful_itinerary_title(number, clean):
        return None
    return clean


def _merge_itinerary_context(
    cell_itinerary: tuple[int, str] | None,
    sheet_itinerary: tuple[int, str] | None,
) -> tuple[int, str] | None:
    if cell_itinerary and sheet_itinerary and cell_itinerary[1] == str(cell_itinerary[0]):
        return sheet_itinerary[0], str(sheet_itinerary[0])
    if cell_itinerary:
        return cell_itinerary
    if sheet_itinerary and sheet_itinerary[1]:
        return sheet_itinerary
    return cell_itinerary or sheet_itinerary


def _infer_sheet_itinerary_context(
    values: list[list[Any]],
    layout: MatrixLayout,
    sheet_name: str,
) -> tuple[int, str] | None:
    sheet_itinerary = _parse_itinerary_from_sheet_name(sheet_name)
    if sheet_itinerary:
        itinerary_number = sheet_itinerary[0]
        itinerary_title = ""
    else:
        itinerary_number = None
        itinerary_title = ""

    row_itineraries: dict[int, str] = {}
    for row_number in range(layout.data_start_row, len(values) + 1):
        parsed = _parse_row_itinerary(values[row_number - 1], layout)
        if not parsed:
            continue
        current_title = row_itineraries.get(parsed[0])
        row_itineraries[parsed[0]] = _best_itinerary_title(
            parsed[0],
            current_title,
            parsed[1],
        )

    if itinerary_number is None and len(row_itineraries) == 1:
        itinerary_number, itinerary_title = next(iter(row_itineraries.items()))
    elif itinerary_number is not None:
        itinerary_title = row_itineraries.get(itinerary_number, "")

    if not itinerary_title and sheet_itinerary and _is_meaningful_itinerary_title(
        sheet_itinerary[0], sheet_itinerary[1]
    ):
        itinerary_title = sheet_itinerary[1]
    if itinerary_number is None:
        return None
    return itinerary_number, itinerary_title


def _parse_itinerary_from_sheet_name(sheet_name: str) -> tuple[int, str] | None:
    text = _clean_text(sheet_name)
    if not text:
        return None
    match = re.search(r"\b(?:ITI|ITINERARIO|HITO)\s*([0-9]{1,2})\b", text, flags=re.I)
    if not match:
        return None
    suffix = text[match.end() :].strip(" ._-")
    return int(match.group(1)), suffix


def _combine_header_rows(first_row: list[Any], second_row: list[Any]) -> list[str]:
    max_len = max(len(first_row), len(second_row))
    result: list[str] = []
    for index in range(max_len):
        top = _clean_text(first_row[index]) if index < len(first_row) else None
        bottom = _clean_text(second_row[index]) if index < len(second_row) else None
        if top and bottom and _label_key(top) != _label_key(bottom):
            result.append(f"{top} {bottom}")
        else:
            result.append(top or bottom or "")
    return result


def _first_line(value: str | None) -> str:
    if not value:
        return ""
    for line in str(value).splitlines():
        clean = line.strip()
        if clean:
            return clean
    return ""


def _title_from_knowledge(value: str | None) -> str:
    title = _first_line(value)
    if not title:
        return ""
    title = re.sub(r"^\s*Texto\s*:\s*", "", title, flags=re.I).strip()
    return title


def _station_text_key(value: Any) -> str:
    text = _normalize_station_text(value)
    if not text:
        return ""
    parsed_station = _parse_station(text)
    if parsed_station and parsed_station[1]:
        text = parsed_station[1]
    else:
        text = _strip_accents(text)
        text = re.sub(r"^\s*ESTACION\s*[:\-]?\s*", "", text, flags=re.I).strip()
    text = _strip_accents(text).upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _register_station_alias(
    itinerary_number: int,
    station: tuple[int, str],
    *,
    station_aliases: dict[tuple[int, str], tuple[int, str]],
    next_station_number_by_itinerary: dict[int, int],
    raw_text: str | None = None,
) -> None:
    station_number, station_name = station
    next_station_number_by_itinerary[itinerary_number] = max(
        next_station_number_by_itinerary.get(itinerary_number, 1),
        station_number + 1,
    )
    for candidate in (station_name, raw_text):
        key = _station_text_key(candidate)
        if key:
            station_aliases.setdefault((itinerary_number, key), station)


def _resolve_station_context(
    value: Any,
    itinerary_number: int,
    *,
    last_station_by_itinerary: dict[int, tuple[int, str]],
    station_aliases: dict[tuple[int, str], tuple[int, str]],
    next_station_number_by_itinerary: dict[int, int],
    station_titles: dict[tuple[int, int], str] | None = None,
) -> tuple[tuple[int, str] | None, str]:
    raw_text = _normalize_station_text(value)
    parsed_station = _parse_station(value)
    if parsed_station:
        station_number, station_name = parsed_station
        station_name = _normalize_station_text(station_name) or ""
        canonical_station_name = ""
        if station_titles:
            canonical_station_name = station_titles.get(
                (itinerary_number, station_number),
                "",
            )
        if canonical_station_name and not station_name:
            station_name = canonical_station_name
        parsed_station = (station_number, station_name)
        alias_key = _station_text_key(station_name or raw_text)
        station = station_aliases.get((itinerary_number, alias_key)) if alias_key else None
        if station is None:
            station = parsed_station
            _register_station_alias(
                itinerary_number,
                station,
                station_aliases=station_aliases,
                next_station_number_by_itinerary=next_station_number_by_itinerary,
                raw_text=raw_text,
            )
        elif canonical_station_name and not station[1]:
            station = (station[0], canonical_station_name)
            _register_station_alias(
                itinerary_number,
                station,
                station_aliases=station_aliases,
                next_station_number_by_itinerary=next_station_number_by_itinerary,
                raw_text=raw_text,
            )
        last_station_by_itinerary[itinerary_number] = station
        return station, "parsed"
    if raw_text:
        alias_key = _station_text_key(raw_text)
        station = station_aliases.get((itinerary_number, alias_key)) if alias_key else None
        if station is None:
            station_number = next_station_number_by_itinerary.get(itinerary_number, 1)
            station = (station_number, raw_text)
            _register_station_alias(
                itinerary_number,
                station,
                station_aliases=station_aliases,
                next_station_number_by_itinerary=next_station_number_by_itinerary,
                raw_text=raw_text,
            )
        last_station_by_itinerary[itinerary_number] = station
        return station, "text"
    inherited_station = last_station_by_itinerary.get(itinerary_number)
    if inherited_station:
        return inherited_station, "inherited"
    return None, "missing"


def _parse_station(value: Any) -> tuple[int, str] | None:
    text = _clean_text(value)
    if not text:
        return None
    normalized = _strip_accents(text).upper()
    match = re.match(r"^(?:E|ESTACI.*?N)\s*[:\-]?\s*0*(\d+)\s*[\.\-:]?\s*", normalized)
    if not match:
        match = re.match(r"^0*(\d+)\s*[\.\-:]?\s*", normalized)
    if not match:
        return None
    return int(match.group(1)), text[match.end() :].strip()


def _infer_course_code_and_grade(
    source_name: str,
    sheet_names: list[str],
) -> tuple[str | None, int | None]:
    candidates = [source_name, *sheet_names]
    for candidate in candidates:
        text = _strip_accents(str(candidate or "")).upper()
        matches = re.findall(r"(?<![A-Z])([A-Z]{2,6})[\s_\-]*([1-6])(?!\d)", text)
        for code, grade in matches:
            if code in COURSE_CODES:
                return code, int(grade)
        for code, grade in matches:
            if code not in NON_COURSE_CODES:
                return code, int(grade)
    return None, None


def _level_code(level: str) -> str:
    text = _strip_accents(level).upper()
    if text.startswith("PRI"):
        return "PRI"
    if text.startswith("SEC"):
        return "SEC"
    return re.sub(r"[^A-Z]", "", text)[:3] or "SEC"


def _label_key(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    text = _strip_accents(text).upper()
    text = re.sub(r"[^A-Z0-9#]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _process_from_label(label: str) -> str | None:
    clean = str(label or "").strip()
    if not clean:
        return None
    for process in PROCESS_NAMES:
        if clean == process:
            return process
    for process in PROCESS_NAMES:
        if re.search(rf"(^|\s){re.escape(process)}(\s|$)", clean):
            return process
    return None


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(text))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _safe_sheet_title(title: str) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "_", str(title or "SUMUN"))
    clean = clean.strip() or "SUMUN"
    return clean[:31]
