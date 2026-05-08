from io import BytesIO
import unittest

from openpyxl import Workbook, load_workbook

from santillana_format.sumun import (
    generate_sumun_template_from_excel,
    inspect_sumun_workbook_sheets,
)


def _build_sumun_workbook(station_value: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "1-Ma4_Iti1"
    ws.append(
        [
            "ITINERARIO",
            "COMPETENCIA",
            "MACROHABILIDAD",
            "MICROHABILIDAD",
            "ESTACIÓN",
            "CONOCIMIENTOS",
            "RECORDAR",
            "COMPRENDER",
            "APLICAR",
            "ANALIZAR",
            "EVALUAR",
            "CREAR",
        ]
    )
    ws.append(
        [
            "Itinerario 1. La célula",
            "Competencia base",
            "Macro base",
            "Micro base",
            station_value,
            "Texto: conocimiento base",
            "Primer skill\n\nSegundo skill",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_sumun_workbook_without_station() -> bytes:
    return _build_sumun_workbook(None)


def _build_sumun_workbook_with_repeated_skills() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "1-Ma4_Iti1"
    ws.append(
        [
            "ITINERARIO",
            "COMPETENCIA",
            "MACROHABILIDAD",
            "MICROHABILIDAD",
            "ESTACIÓN",
            "CONOCIMIENTOS",
            "RECORDAR",
            "COMPRENDER",
            "APLICAR",
            "ANALIZAR",
            "EVALUAR",
            "CREAR",
        ]
    )
    ws.append(
        [
            "Itinerario 1. La célula",
            "Competencia base",
            "Macro compartida",
            "Micro compartida",
            "E1 - Estación 1",
            "Texto: conocimiento base",
            "Skill 1",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    ws.append(
        [
            "Itinerario 1. La célula",
            "Competencia base",
            "Macro compartida",
            "Micro compartida",
            "E2 - Estación 2",
            "Texto: conocimiento base",
            "Skill 2",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _generated_rows(workbook_bytes: bytes) -> list[tuple]:
    wb = load_workbook(BytesIO(workbook_bytes))
    ws = wb[wb.sheetnames[0]]
    return [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]


def _build_sumun_workbook_with_process_value(process_value: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "1-Ma4_Iti1"
    ws.append(
        [
            "ITINERARIO",
            "COMPETENCIA",
            "MACROHABILIDAD",
            "MICROHABILIDAD",
            "ESTACIÓN",
            "CONOCIMIENTOS",
            "RECORDAR",
            "COMPRENDER",
            "APLICAR",
            "ANALIZAR",
            "EVALUAR",
            "CREAR",
        ]
    )
    ws.append(
        [
            "Itinerario 1. La célula",
            "Competencia base",
            "Macro base",
            "Micro base",
            "E1 - Estación 1",
            "Texto: conocimiento base",
            process_value,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


class SumunStationParsingTests(unittest.TestCase):
    def test_generate_template_accepts_common_station_formats(self) -> None:
        for station_value in (
            "1. Celula",
            "E1 - Celula",
            "Estación 1 - Celula",
            "Primera estacion",
        ):
            with self.subTest(station_value=station_value):
                output_bytes, summary = generate_sumun_template_from_excel(
                    _build_sumun_workbook(station_value),
                    source_name="MA4.xlsx",
                )
                self.assertTrue(output_bytes)
                self.assertEqual(summary.generated_rows, 1)

    def test_inspection_reports_missing_station_value(self) -> None:
        sheets = inspect_sumun_workbook_sheets(_build_sumun_workbook_without_station())
        self.assertEqual(len(sheets), 1)
        self.assertFalse(sheets[0].detected)
        self.assertEqual(sheets[0].estimated_rows, 0)
        self.assertIn("faltan estaciones", sheets[0].reason)

    def test_macro_and_micro_ids_are_reused_across_station_changes(self) -> None:
        output_bytes, summary = generate_sumun_template_from_excel(
            _build_sumun_workbook_with_repeated_skills(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 2)
        self.assertEqual(summary.micro_count, 2)
        self.assertEqual(summary.unique_micro_count, 1)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0][5], 1)
        self.assertEqual(rows[1][5], 1)
        self.assertEqual(rows[0][8], 1)
        self.assertEqual(rows[1][8], 1)
        self.assertEqual(rows[0][10], 1)
        self.assertEqual(rows[1][10], 1)

    def test_text_only_station_is_reused_by_content(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "1-Ma4_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÓN",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "Itinerario 1. La célula",
                "Competencia base",
                "Macro 1",
                "Micro 1",
                "Primera estación",
                "Texto: conocimiento base",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Itinerario 1. La célula",
                "Competencia base",
                "Macro 2",
                "Micro 2",
                "Primera estación",
                "Texto: conocimiento base",
                "Skill 2",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 2)
        self.assertEqual(rows[0][5], 1)
        self.assertEqual(rows[1][5], 1)
        self.assertEqual(rows[0][12], 1)
        self.assertEqual(rows[1][12], 1)
        self.assertEqual(rows[0][13], "Primera estación")
        self.assertEqual(rows[1][13], "Primera estación")

    def test_station_name_is_shared_when_numeric_and_descriptive_forms_mix(self) -> None:
        cases = (
            (1, "E1 - Primera estaciÃ³n"),
            ("E1 - Primera estaciÃ³n", 1),
        )
        for first_station_value, second_station_value in cases:
            with self.subTest(
                first_station_value=first_station_value,
                second_station_value=second_station_value,
            ):
                wb = Workbook()
                ws = wb.active
                ws.title = "1-Ma4_Iti1"
                ws.append(
                    [
                        "ITINERARIO",
                        "COMPETENCIA",
                        "MACROHABILIDAD",
                        "MICROHABILIDAD",
                        "ESTACIÃ“N",
                        "CONOCIMIENTOS",
                        "RECORDAR",
                        "COMPRENDER",
                        "APLICAR",
                        "ANALIZAR",
                        "EVALUAR",
                        "CREAR",
                    ]
                )
                ws.append(
                    [
                        "Itinerario 1. La cÃ©lula",
                        "Competencia base",
                        "Macro 1",
                        "Micro 1",
                        first_station_value,
                        "Texto: conocimiento base",
                        "Skill 1",
                        None,
                        None,
                        None,
                        None,
                        None,
                    ]
                )
                ws.append(
                    [
                        "Itinerario 1. La cÃ©lula",
                        "Competencia base",
                        "Macro 2",
                        "Micro 2",
                        second_station_value,
                        "Texto: conocimiento base",
                        "Skill 2",
                        None,
                        None,
                        None,
                        None,
                        None,
                    ]
                )
                output = BytesIO()
                wb.save(output)

                output_bytes, summary = generate_sumun_template_from_excel(
                    output.getvalue(),
                    source_name="MA4.xlsx",
                )
                rows = _generated_rows(output_bytes)

                self.assertEqual(summary.generated_rows, 2)
                self.assertEqual(rows[0][12], 1)
                self.assertEqual(rows[1][12], 1)
                self.assertEqual(rows[0][13], "Primera estaciÃ³n")
                self.assertEqual(rows[1][13], "Primera estaciÃ³n")

    def test_itinerary_title_is_reused_when_numeric_and_descriptive_forms_mix(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mat3_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÓN",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "1",
                "Competencia base",
                "Macro 1",
                "Micro 1",
                "E1 - Estación 1",
                "Texto: conocimiento base",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Itinerario 1. La célula",
                "Competencia base",
                "Macro 2",
                "Micro 2",
                "E1 - Estación 1",
                "Texto: conocimiento base",
                "Skill 2",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 2)
        self.assertEqual(rows[0][6], "La célula")
        self.assertEqual(rows[1][6], "La célula")
        self.assertEqual(rows[0][5], 1)
        self.assertEqual(rows[1][5], 1)

    def test_itinerary_title_is_shared_across_sheets_by_number(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Mat3_Iti1"
        ws1.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÓN",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws1.append(
            [
                "1",
                "Competencia base",
                "Macro 1",
                "Micro 1",
                "E1 - Estación 1",
                "Texto: conocimiento base",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws2 = wb.create_sheet("Mat3_Iti1_detalle")
        ws2.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÓN",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws2.append(
            [
                "Itinerario 1. La célula",
                "Competencia base",
                "Macro 2",
                "Micro 2",
                "E2 - Estación 2",
                "Texto: conocimiento base",
                "Skill 2",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 2)
        self.assertEqual(rows[0][6], "La célula")
        self.assertEqual(rows[1][6], "La célula")
        self.assertEqual(rows[0][5], 1)
        self.assertEqual(rows[1][5], 1)

    def test_specific_skill_cell_value_is_kept_as_one_row(self) -> None:
        for process_value in (
            "Bullet 1\nBullet 2",
            "- Skill 1\n- Skill 2",
            "1) Skill 1\n2) Skill 2",
        ):
            with self.subTest(process_value=process_value):
                output_bytes, summary = generate_sumun_template_from_excel(
                    _build_sumun_workbook_with_process_value(process_value),
                    source_name="MA4.xlsx",
                )
                rows = _generated_rows(output_bytes)
                self.assertEqual(summary.generated_rows, 1)
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0][17], process_value)

    def test_specific_skills_plain_multiline_cell_stays_as_one_row(self) -> None:
        output_bytes, summary = generate_sumun_template_from_excel(
            _build_sumun_workbook_with_process_value("Skill 1\nSkill 2"),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 1)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][17], "Skill 1\nSkill 2")

    def test_summary_reports_specific_rows_by_itinerary_and_knowledge(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mat3_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACION",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "Itinerario 1. La celula",
                "Competencia base",
                "Macro base",
                "Micro base",
                "E1 - Estacion 1",
                "Texto: conocimiento base",
                "Ubica informacion literal",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Itinerario 1. La celula",
                "Competencia base",
                "Macro base",
                "Micro base",
                "E1 - Estacion 1",
                "Texto: conocimiento base",
                None,
                "Describir informacion explicita en distintas partes del texto.",
                None,
                None,
                None,
                None,
            ]
        )
        ws.append(
            [
                "Itinerario 1. La celula",
                "Competencia base",
                "Macro base",
                "Micro base",
                "E2 - Estacion 2",
                "Texto: conocimiento complementario",
                None,
                "Relaciona informacion del texto con el contexto.",
                None,
                None,
                None,
                None,
            ]
        )
        ws2 = wb.create_sheet("Mat3_Iti2")
        ws2.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACION",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws2.append(
            [
                "Itinerario 2. Otra ruta",
                "Competencia base",
                "Macro base",
                "Micro base",
                "E1 - Estacion 1",
                "Texto: conocimiento base",
                "Recuerda una idea central",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 4)
        self.assertEqual(len(rows), 4)
        self.assertEqual(
            summary.specific_rows_by_itinerary,
            [
                {
                    "itinerary_number": 1,
                    "itinerary": "La celula",
                    "specific_rows": 3,
                },
                {
                    "itinerary_number": 2,
                    "itinerary": "Otra ruta",
                    "specific_rows": 1,
                },
            ],
        )
        self.assertEqual(
            summary.specific_rows_by_knowledge,
            [
                {
                    "itinerary_number": 1,
                    "itinerary": "La celula",
                    "station_number": 1,
                    "station": "Estacion 1",
                    "knowledge": "Texto: conocimiento base",
                    "specific_rows": 2,
                },
                {
                    "itinerary_number": 1,
                    "itinerary": "La celula",
                    "station_number": 2,
                    "station": "Estacion 2",
                    "knowledge": "Texto: conocimiento complementario",
                    "specific_rows": 1,
                },
                {
                    "itinerary_number": 2,
                    "itinerary": "Otra ruta",
                    "station_number": 1,
                    "station": "Estacion 1",
                    "knowledge": "Texto: conocimiento base",
                    "specific_rows": 1,
                },
            ],
        )
    def test_two_row_header_layout_is_detected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mat3_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "ESTACIÓN",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "CONOCIMIENTOS",
                "NANOHABILIDADES",
                "NANOHABILIDADES",
                "NANOHABILIDADES",
                "NANOHABILIDADES",
                "NANOHABILIDADES",
                "NANOHABILIDADES",
            ]
        )
        ws.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "Itinerario 1. La célula",
                "E1 - Estación 1",
                "Competencia base",
                "Macro base",
                "Micro base",
                "Texto: conocimiento base",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 1)
        self.assertEqual(rows[0][5], 1)
        self.assertEqual(rows[0][6], "La célula")
        self.assertEqual(rows[0][17], "Skill 1")


    def test_station_text_drops_prefix_before_colon_and_fixes_line_breaks(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mat3_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÃ“N",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "Itinerario 1. La cÃ©lula",
                "Competencia base",
                "Macro base",
                "Micro base",
                "EstaciÃ³n 17:\nMedidas de\ndispersiÃ³n y distribuciones\nestadÃ­sticas",
                "Texto: conocimiento base",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 1)
        self.assertEqual(rows[0][12], 17)
        self.assertEqual(
            rows[0][13],
            "Medidas de dispersiÃ³n y distribuciones estadÃ­sticas.",
        )

    def test_knowledge_text_starts_with_letter_and_joins_blocks_with_period(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mat3_Iti1"
        ws.append(
            [
                "ITINERARIO",
                "COMPETENCIA",
                "MACROHABILIDAD",
                "MICROHABILIDAD",
                "ESTACIÃ“N",
                "CONOCIMIENTOS",
                "RECORDAR",
                "COMPRENDER",
                "APLICAR",
                "ANALIZAR",
                "EVALUAR",
                "CREAR",
            ]
        )
        ws.append(
            [
                "Itinerario 1. La cÃ©lula",
                "Competencia base",
                "Macro base",
                "Micro base",
                "E1 - EstaciÃ³n 1",
                ". 1 Medidas de centralizaciÃ³n para datos agrupados: media, moda y mediana\n"
                "Medidas de localizaciÃ³n: terciles, cuartiles,\n"
                "quintiles y percentiles",
                "Skill 1",
                None,
                None,
                None,
                None,
                None,
            ]
        )
        output = BytesIO()
        wb.save(output)

        output_bytes, summary = generate_sumun_template_from_excel(
            output.getvalue(),
            source_name="MA4.xlsx",
        )
        rows = _generated_rows(output_bytes)

        self.assertEqual(summary.generated_rows, 1)
        self.assertEqual(
            rows[0][14],
            "Medidas de centralizaciÃ³n para datos agrupados: media, moda y mediana. "
            "Medidas de localizaciÃ³n: terciles, cuartiles, quintiles y percentiles",
        )


if __name__ == "__main__":
    unittest.main()
